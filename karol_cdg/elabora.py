"""
Script di elaborazione completa del CDG.
Legge i dati dal Master, calcola CE Industriale, allocazione sede, CE Gestionale,
KPI e scrive tutto nei fogli di output dell'Excel Master.
"""

import logging
from pathlib import Path
from collections import defaultdict

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from karol_cdg.config import (
    UNITA_OPERATIVE, UO_OPERATIVE, EXCEL_MASTER,
    VOCI_RICAVI, VOCI_COSTI_DIRETTI, VOCI_COSTI_SEDE, VOCI_ALTRI_COSTI,
    BENCHMARK, ALERT_CONFIG, SOGLIE_SEMAFORO,
    MESI_IT, MESI_BREVI_IT, StatoUO,
    CategoriaCostoSede, DriverAllocazione,
)
from karol_cdg.excel.reader import leggi_foglio

logger = logging.getLogger(__name__)

# ============================================================================
# STILI EXCEL
# ============================================================================

FONT_TITOLO = Font(name="Calibri", size=14, bold=True, color="1F4E79")
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_SEZIONE = Font(name="Calibri", size=11, bold=True, color="1F4E79")
FONT_NORMALE = Font(name="Calibri", size=10)
FONT_TOTALE = Font(name="Calibri", size=10, bold=True)
FONT_POSITIVO = Font(name="Calibri", size=10, bold=True, color="006600")
FONT_NEGATIVO = Font(name="Calibri", size=10, bold=True, color="CC0000")

FILL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FILL_SEZIONE = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_TOTALE = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_MOL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_ROSSO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_GIALLO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_MOL_GEST = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_RISULTATO = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

BORDO = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

FMT_EURO = '#,##0'
FMT_EURO_DEC = '#,##0.00'
FMT_PCT = '0.0%'


def _stile_header(ws, riga, n_col):
    for c in range(1, n_col + 1):
        cell = ws.cell(row=riga, column=c)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDO


def _stile_sezione(ws, riga, n_col):
    for c in range(1, n_col + 1):
        cell = ws.cell(row=riga, column=c)
        cell.font = FONT_SEZIONE
        cell.fill = FILL_SEZIONE
        cell.border = BORDO


def _scrivi_riga_valori(ws, riga, valori_dict, col_uo_map, fmt=FMT_EURO, font=FONT_NORMALE, fill=None):
    """Scrive valori per UO nelle colonne corrispondenti."""
    totale = 0
    for uo, col in col_uo_map.items():
        val = valori_dict.get(uo, 0)
        cell = ws.cell(row=riga, column=col)
        cell.value = round(val)
        cell.number_format = fmt
        cell.font = font
        cell.border = BORDO
        if fill:
            cell.fill = fill
        totale += val
    # Colonna totale
    col_tot = max(col_uo_map.values()) + 1
    cell = ws.cell(row=riga, column=col_tot)
    cell.value = round(totale)
    cell.number_format = fmt
    cell.font = FONT_TOTALE
    cell.border = BORDO
    if fill:
        cell.fill = fill
    return totale


# ============================================================================
# 1. LETTURA DATI
# ============================================================================

def leggi_dati_master(file_path: Path):
    """Legge tutti i dati necessari dal file Excel Master."""
    logger.info("Lettura dati dal Master...")

    df_prod = leggi_foglio(file_path, "Produzione_Mensile")
    df_costi = leggi_foglio(file_path, "Costi_Mensili")
    df_sede = leggi_foglio(file_path, "Costi_Sede_Dettaglio")
    df_driver = leggi_foglio(file_path, "Driver_Allocazione")
    df_pers = leggi_foglio(file_path, "Anagrafiche_Personale")
    df_scad = leggi_foglio(file_path, "Scadenzario")

    logger.info(f"  Produzione: {len(df_prod)} righe")
    logger.info(f"  Costi: {len(df_costi)} righe")
    logger.info(f"  Costi sede: {len(df_sede)} righe")
    logger.info(f"  Personale: {len(df_pers)} righe")
    logger.info(f"  Scadenzario: {len(df_scad)} righe")

    return {
        "produzione": df_prod,
        "costi": df_costi,
        "costi_sede": df_sede,
        "driver": df_driver,
        "personale": df_pers,
        # Chiavi con nomi originali per compatibilita' con pagina Cash Flow
        "Produzione_Mensile": df_prod,
        "Costi_Mensili": df_costi,
        "Costi_Sede_Dettaglio": df_sede,
        "Driver_Allocazione": df_driver,
        "Anagrafiche_Personale": df_pers,
        "Scadenzario": df_scad,
    }


# ============================================================================
# 2. CALCOLO CE INDUSTRIALE
# ============================================================================

def calcola_ce_industriale_da_dati(dati: dict):
    """Calcola il CE Industriale per ogni UO dal dataframe."""
    df_prod = dati["produzione"]
    df_costi = dati["costi"]

    risultati = {}

    for uo in UO_OPERATIVE:
        # Ricavi per voce
        ricavi = {}
        mask_prod = df_prod.iloc[:, 0] == uo
        df_uo_prod = df_prod.loc[mask_prod]

        for _, row in df_uo_prod.iterrows():
            codice = str(row.iloc[4])  # codice_prestazione
            importo = float(row.iloc[8]) if pd.notna(row.iloc[8]) else 0
            ricavi[codice] = ricavi.get(codice, 0) + importo

        totale_ricavi = sum(ricavi.values())

        # Costi diretti per voce
        costi = {}
        mask_costi = df_costi.iloc[:, 0] == uo
        df_uo_costi = df_costi.loc[mask_costi]

        for _, row in df_uo_costi.iterrows():
            codice = str(row.iloc[3])  # codice_voce
            importo = float(row.iloc[6]) if pd.notna(row.iloc[6]) else 0
            costi[codice] = costi.get(codice, 0) + importo

        totale_costi = sum(costi.values())

        # Costi per sotto-categoria
        costi_personale = sum(v for k, v in costi.items() if k.startswith("CD0"))
        costi_acquisti = sum(v for k, v in costi.items() if k.startswith("CD1"))
        costi_servizi = sum(v for k, v in costi.items() if k.startswith("CD2"))
        costi_ammort = sum(v for k, v in costi.items() if k.startswith("CD3"))

        mol_i = totale_ricavi - totale_costi
        mol_pct = (mol_i / totale_ricavi) if totale_ricavi > 0 else 0

        # Giornate degenza
        giornate = 0
        for _, row in df_uo_prod.iterrows():
            g = row.iloc[9] if pd.notna(row.iloc[9]) else 0  # giornate_degenza
            giornate += int(g)

        risultati[uo] = {
            "ricavi": ricavi,
            "totale_ricavi": totale_ricavi,
            "costi": costi,
            "costi_personale": costi_personale,
            "costi_acquisti": costi_acquisti,
            "costi_servizi": costi_servizi,
            "costi_ammort": costi_ammort,
            "totale_costi": totale_costi,
            "mol_industriale": mol_i,
            "mol_pct": mol_pct,
            "giornate_degenza": giornate,
        }

        logger.info(
            f"  {uo}: Ricavi={totale_ricavi:>12,.0f}  Costi={totale_costi:>12,.0f}  "
            f"MOL-I={mol_i:>12,.0f} ({mol_pct:.1%})"
        )

    return risultati


# ============================================================================
# 3. CALCOLO ALLOCAZIONE COSTI SEDE
# ============================================================================

def calcola_allocazione_sede(dati: dict, ce_industriale: dict):
    """Alloca i costi sede sulle UO in base ai driver."""
    df_sede = dati["costi_sede"]
    df_driver = dati["driver"]

    # Ricavi per UO (per allocazione pro-quota)
    ricavi_uo = {uo: ce["totale_ricavi"] for uo, ce in ce_industriale.items()}
    ricavi_totali = sum(ricavi_uo.values())

    allocazione = {uo: {} for uo in UO_OPERATIVE}
    riepilogo_categorie = defaultdict(float)
    non_allocati = 0

    for _, row in df_sede.iterrows():
        codice = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
        desc = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ""
        importo = float(row.iloc[2]) if pd.notna(row.iloc[2]) else 0
        categoria = str(row.iloc[3]) if pd.notna(row.iloc[3]) else ""
        driver = str(row.iloc[5]) if pd.notna(row.iloc[5]) else ""

        if importo <= 0:
            continue

        riepilogo_categorie[categoria] += importo

        # Sviluppo e Storico non allocabili non vengono ribaltati
        if driver == "NON_ALLOCABILE" or categoria in ("SVILUPPO", "STORICO"):
            non_allocati += importo
            continue

        # Trova i valori driver per UO
        valori_driver = {}
        if driver == "RICAVI":
            valori_driver = ricavi_uo
        elif driver == "QUOTA_FISSA":
            # Quota uguale per tutte le UO operative
            for uo in UO_OPERATIVE:
                valori_driver[uo] = 1
        else:
            # Cerca nella tabella driver
            mask = df_driver.iloc[:, 0] == codice
            if mask.any():
                row_driver = df_driver.loc[mask].iloc[0]
                uo_list = list(UNITA_OPERATIVE.keys())
                for i, uo_cod in enumerate(uo_list):
                    if uo_cod in UO_OPERATIVE:
                        col_idx = 3 + i  # colonne UO partono da indice 3
                        if col_idx < len(row_driver):
                            val = row_driver.iloc[col_idx]
                            valori_driver[uo_cod] = float(val) if pd.notna(val) and val != 0 else 0
            else:
                # Fallback: alloca per ricavi
                valori_driver = ricavi_uo

        # Calcola percentuali e alloca
        totale_driver = sum(valori_driver.values())
        if totale_driver > 0:
            for uo in UO_OPERATIVE:
                quota = valori_driver.get(uo, 0) / totale_driver
                importo_allocato = importo * quota
                allocazione[uo][codice] = allocazione[uo].get(codice, 0) + importo_allocato

    # Log riepilogo
    logger.info("Allocazione costi sede:")
    for cat, tot in riepilogo_categorie.items():
        logger.info(f"  {cat}: {tot:>12,.0f}")
    logger.info(f"  Non allocati: {non_allocati:>12,.0f}")

    for uo in UO_OPERATIVE:
        tot_uo = sum(allocazione[uo].values())
        logger.info(f"  Allocato a {uo}: {tot_uo:>12,.0f}")

    return allocazione, non_allocati, dict(riepilogo_categorie)


# ============================================================================
# 4. CALCOLO CE GESTIONALE
# ============================================================================

def calcola_ce_gestionale_da_dati(ce_industriale: dict, allocazione: dict, non_allocati: float):
    """Calcola CE Gestionale = CE Industriale - Costi Sede Allocati."""
    risultati = {}

    for uo in UO_OPERATIVE:
        ce_ind = ce_industriale[uo]
        costi_sede_uo = sum(allocazione[uo].values())

        mol_g = ce_ind["mol_industriale"] - costi_sede_uo
        mol_g_pct = (mol_g / ce_ind["totale_ricavi"]) if ce_ind["totale_ricavi"] > 0 else 0

        risultati[uo] = {
            "mol_industriale": ce_ind["mol_industriale"],
            "costi_sede_allocati": costi_sede_uo,
            "dettaglio_sede": allocazione[uo],
            "mol_gestionale": mol_g,
            "mol_gestionale_pct": mol_g_pct,
            "totale_ricavi": ce_ind["totale_ricavi"],
        }

        logger.info(
            f"  {uo}: MOL-I={ce_ind['mol_industriale']:>10,.0f}  "
            f"Sede={costi_sede_uo:>10,.0f}  MOL-G={mol_g:>10,.0f} ({mol_g_pct:.1%})"
        )

    return risultati


# ============================================================================
# 5. CALCOLO KPI
# ============================================================================

def calcola_kpi(ce_industriale: dict, ce_gestionale: dict, dati: dict):
    """Calcola tutti i KPI operativi, economici e finanziari."""
    kpi_list = []

    # KPI per UO
    for uo in UO_OPERATIVE:
        ce_i = ce_industriale[uo]
        ce_g = ce_gestionale[uo]
        uo_info = UNITA_OPERATIVE[uo]

        # MOL % Industriale
        kpi_list.append({
            "kpi": "MOL % Industriale",
            "unita_operativa": uo,
            "valore": ce_i["mol_pct"],
            "target": 0.15,
            "formula": "MOL-I / Ricavi",
        })

        # MOL % Gestionale
        kpi_list.append({
            "kpi": "MOL % Gestionale",
            "unita_operativa": uo,
            "valore": ce_g["mol_gestionale_pct"],
            "target": 0.08,
            "formula": "MOL-G / Ricavi",
        })

        # Costo personale / Ricavi
        cp_pct = ce_i["costi_personale"] / ce_i["totale_ricavi"] if ce_i["totale_ricavi"] > 0 else 0
        kpi_list.append({
            "kpi": "Costo Personale %",
            "unita_operativa": uo,
            "valore": cp_pct,
            "target": 0.55,
            "formula": "Costi Personale / Ricavi",
        })

        # Occupancy (solo per strutture con PL)
        if uo_info.posti_letto > 0 and ce_i["giornate_degenza"] > 0:
            gg_disponibili = uo_info.posti_letto * 365
            occupancy = ce_i["giornate_degenza"] / gg_disponibili
            kpi_list.append({
                "kpi": "Occupancy %",
                "unita_operativa": uo,
                "valore": occupancy,
                "target": 0.90,
                "formula": "Giornate erogate / Giornate disponibili",
            })

            # Ricavo medio/giornata
            ricavo_gg = ce_i["totale_ricavi"] / ce_i["giornate_degenza"]
            kpi_list.append({
                "kpi": "Ricavo medio/giornata",
                "unita_operativa": uo,
                "valore": ricavo_gg,
                "target": uo_info.posti_letto * 3.5,  # stima
                "formula": "Ricavi / Giornate degenza",
            })

            # Costo personale/giornata
            costo_gg = ce_i["costi_personale"] / ce_i["giornate_degenza"]
            kpi_list.append({
                "kpi": "Costo personale/giornata",
                "unita_operativa": uo,
                "valore": costo_gg,
                "target": ricavo_gg * 0.55,
                "formula": "Costi personale / Giornate degenza",
            })

    # KPI Consolidati
    totale_ricavi = sum(ce_industriale[uo]["totale_ricavi"] for uo in UO_OPERATIVE)
    totale_mol_i = sum(ce_industriale[uo]["mol_industriale"] for uo in UO_OPERATIVE)
    totale_mol_g = sum(ce_gestionale[uo]["mol_gestionale"] for uo in UO_OPERATIVE)
    totale_costi_pers = sum(ce_industriale[uo]["costi_personale"] for uo in UO_OPERATIVE)
    totale_costi_sede = sum(ce_gestionale[uo]["costi_sede_allocati"] for uo in UO_OPERATIVE)

    kpi_list.append({
        "kpi": "MOL % Consolidato",
        "unita_operativa": "GRUPPO",
        "valore": totale_mol_g / totale_ricavi if totale_ricavi > 0 else 0,
        "target": 0.12,
        "formula": "MOL-G Totale / Ricavi Totali",
    })

    kpi_list.append({
        "kpi": "Peso Costi Sede %",
        "unita_operativa": "GRUPPO",
        "valore": totale_costi_sede / totale_ricavi if totale_ricavi > 0 else 0,
        "target": 0.15,
        "formula": "Costi Sede Allocati / Ricavi Totali",
    })

    kpi_list.append({
        "kpi": "Costo Personale % Consolidato",
        "unita_operativa": "GRUPPO",
        "valore": totale_costi_pers / totale_ricavi if totale_ricavi > 0 else 0,
        "target": 0.55,
        "formula": "Costi Personale Totali / Ricavi Totali",
    })

    # Assegna livello alert
    for kpi in kpi_list:
        kpi["alert"] = _livello_alert(kpi["kpi"], kpi["valore"])

    return kpi_list


def _livello_alert(nome_kpi, valore):
    """Determina il livello alert per un KPI."""
    if "MOL" in nome_kpi and "%" in nome_kpi:
        if valore >= 0.15:
            return "VERDE"
        elif valore >= 0.08:
            return "GIALLO"
        else:
            return "ROSSO"
    elif "Personale %" in nome_kpi or "Costi Sede" in nome_kpi:
        # Invertito: sotto è meglio
        if valore <= 0.55:
            return "VERDE"
        elif valore <= 0.60:
            return "GIALLO"
        else:
            return "ROSSO"
    elif "Occupancy" in nome_kpi:
        if valore >= 0.90:
            return "VERDE"
        elif valore >= 0.80:
            return "GIALLO"
        else:
            return "ROSSO"
    return "VERDE"


# ============================================================================
# 6. SCRITTURA FOGLI EXCEL
# ============================================================================

def _pulisci_foglio(ws):
    """Pulisce un foglio rimuovendo merge e dati."""
    # Prima rimuovi tutti i merge
    for merge_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge_range))
    # Poi pulisci i valori
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None
            cell.font = FONT_NORMALE


def scrivi_ce_industriale(wb, ce_industriale: dict):
    """Scrive il CE Industriale nel foglio Excel."""
    ws = wb["CE_Industriale"]

    _pulisci_foglio(ws)

    # Titolo
    ws.cell(row=1, column=1, value="CONTO ECONOMICO INDUSTRIALE - ANNO 2025")
    ws.cell(row=1, column=1).font = FONT_TITOLO
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(UO_OPERATIVE) + 1)

    # Header colonne
    uo_list = list(UO_OPERATIVE)
    col_uo_map = {}
    ws.cell(row=3, column=1, value="Codice")
    ws.cell(row=3, column=2, value="Voce")
    for i, uo in enumerate(uo_list):
        col = 3 + i
        col_uo_map[uo] = col
        ws.cell(row=3, column=col, value=uo)
    col_tot = 3 + len(uo_list)
    ws.cell(row=3, column=col_tot, value="TOTALE")
    n_col = col_tot
    _stile_header(ws, 3, n_col)

    riga = 4

    # --- RICAVI ---
    ws.cell(row=riga, column=2, value="RICAVI")
    _stile_sezione(ws, riga, n_col)
    riga += 1

    totali_ricavi = {}
    for codice, desc in VOCI_RICAVI.items():
        ws.cell(row=riga, column=1, value=codice).font = FONT_NORMALE
        ws.cell(row=riga, column=1).border = BORDO
        ws.cell(row=riga, column=2, value=desc).font = FONT_NORMALE
        ws.cell(row=riga, column=2).border = BORDO
        valori = {uo: ce_industriale[uo]["ricavi"].get(codice, 0) for uo in uo_list}
        _scrivi_riga_valori(ws, riga, valori, col_uo_map)
        riga += 1

    # Totale Ricavi
    ws.cell(row=riga, column=2, value="TOTALE RICAVI").font = FONT_TOTALE
    ws.cell(row=riga, column=2).border = BORDO
    valori_tot_r = {uo: ce_industriale[uo]["totale_ricavi"] for uo in uo_list}
    _scrivi_riga_valori(ws, riga, valori_tot_r, col_uo_map, font=FONT_TOTALE, fill=FILL_TOTALE)
    riga += 2

    # --- COSTI DIRETTI ---
    ws.cell(row=riga, column=2, value="COSTI DIRETTI")
    _stile_sezione(ws, riga, n_col)
    riga += 1

    # Personale
    ws.cell(row=riga, column=2, value="Personale diretto").font = FONT_SEZIONE
    ws.cell(row=riga, column=2).border = BORDO
    riga += 1
    for codice in ["CD01", "CD02", "CD03", "CD04", "CD05"]:
        if codice in VOCI_COSTI_DIRETTI:
            ws.cell(row=riga, column=1, value=codice).font = FONT_NORMALE
            ws.cell(row=riga, column=1).border = BORDO
            ws.cell(row=riga, column=2, value=VOCI_COSTI_DIRETTI[codice]).font = FONT_NORMALE
            ws.cell(row=riga, column=2).border = BORDO
            valori = {uo: ce_industriale[uo]["costi"].get(codice, 0) for uo in uo_list}
            _scrivi_riga_valori(ws, riga, valori, col_uo_map)
            riga += 1

    # Subtotale personale
    ws.cell(row=riga, column=2, value="Subtotale Personale").font = FONT_TOTALE
    ws.cell(row=riga, column=2).border = BORDO
    valori = {uo: ce_industriale[uo]["costi_personale"] for uo in uo_list}
    _scrivi_riga_valori(ws, riga, valori, col_uo_map, font=FONT_TOTALE)
    riga += 1

    # Acquisti
    ws.cell(row=riga, column=2, value="Acquisti diretti").font = FONT_SEZIONE
    ws.cell(row=riga, column=2).border = BORDO
    riga += 1
    for codice in ["CD10", "CD11", "CD12", "CD13"]:
        if codice in VOCI_COSTI_DIRETTI:
            ws.cell(row=riga, column=1, value=codice).font = FONT_NORMALE
            ws.cell(row=riga, column=1).border = BORDO
            ws.cell(row=riga, column=2, value=VOCI_COSTI_DIRETTI[codice]).font = FONT_NORMALE
            ws.cell(row=riga, column=2).border = BORDO
            valori = {uo: ce_industriale[uo]["costi"].get(codice, 0) for uo in uo_list}
            _scrivi_riga_valori(ws, riga, valori, col_uo_map)
            riga += 1

    # Servizi
    ws.cell(row=riga, column=2, value="Servizi diretti").font = FONT_SEZIONE
    ws.cell(row=riga, column=2).border = BORDO
    riga += 1
    for codice in ["CD20", "CD21", "CD22", "CD23", "CD24"]:
        if codice in VOCI_COSTI_DIRETTI:
            ws.cell(row=riga, column=1, value=codice).font = FONT_NORMALE
            ws.cell(row=riga, column=1).border = BORDO
            ws.cell(row=riga, column=2, value=VOCI_COSTI_DIRETTI[codice]).font = FONT_NORMALE
            ws.cell(row=riga, column=2).border = BORDO
            valori = {uo: ce_industriale[uo]["costi"].get(codice, 0) for uo in uo_list}
            _scrivi_riga_valori(ws, riga, valori, col_uo_map)
            riga += 1

    # Ammortamenti
    ws.cell(row=riga, column=1, value="CD30").font = FONT_NORMALE
    ws.cell(row=riga, column=1).border = BORDO
    ws.cell(row=riga, column=2, value=VOCI_COSTI_DIRETTI["CD30"]).font = FONT_NORMALE
    ws.cell(row=riga, column=2).border = BORDO
    valori = {uo: ce_industriale[uo]["costi"].get("CD30", 0) for uo in uo_list}
    _scrivi_riga_valori(ws, riga, valori, col_uo_map)
    riga += 1

    # Totale Costi Diretti
    ws.cell(row=riga, column=2, value="TOTALE COSTI DIRETTI").font = FONT_TOTALE
    ws.cell(row=riga, column=2).border = BORDO
    valori_tot_c = {uo: ce_industriale[uo]["totale_costi"] for uo in uo_list}
    _scrivi_riga_valori(ws, riga, valori_tot_c, col_uo_map, font=FONT_TOTALE, fill=FILL_TOTALE)
    riga += 2

    # --- MOL INDUSTRIALE ---
    ws.cell(row=riga, column=2, value="MOL INDUSTRIALE (MOL-I)").font = Font(
        name="Calibri", size=12, bold=True, color="006600"
    )
    ws.cell(row=riga, column=2).border = BORDO
    valori_mol = {uo: ce_industriale[uo]["mol_industriale"] for uo in uo_list}
    for uo, col in col_uo_map.items():
        val = valori_mol[uo]
        cell = ws.cell(row=riga, column=col)
        cell.value = round(val)
        cell.number_format = FMT_EURO
        cell.font = FONT_POSITIVO if val >= 0 else FONT_NEGATIVO
        cell.fill = FILL_MOL
        cell.border = BORDO
    # Totale
    tot_mol = sum(valori_mol.values())
    cell = ws.cell(row=riga, column=col_tot)
    cell.value = round(tot_mol)
    cell.number_format = FMT_EURO
    cell.font = FONT_POSITIVO if tot_mol >= 0 else FONT_NEGATIVO
    cell.fill = FILL_MOL
    cell.border = BORDO
    riga += 1

    # Margine %
    ws.cell(row=riga, column=2, value="MARGINE % INDUSTRIALE").font = FONT_TOTALE
    ws.cell(row=riga, column=2).border = BORDO
    for uo, col in col_uo_map.items():
        pct = ce_industriale[uo]["mol_pct"]
        cell = ws.cell(row=riga, column=col)
        cell.value = pct
        cell.number_format = FMT_PCT
        cell.font = FONT_POSITIVO if pct >= 0.15 else (FONT_NEGATIVO if pct < 0.08 else FONT_TOTALE)
        cell.border = BORDO
    # Totale %
    pct_tot = tot_mol / sum(valori_tot_r.values()) if sum(valori_tot_r.values()) > 0 else 0
    cell = ws.cell(row=riga, column=col_tot)
    cell.value = pct_tot
    cell.number_format = FMT_PCT
    cell.font = FONT_TOTALE
    cell.border = BORDO

    # Larghezza colonne
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 38
    for i in range(3, n_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    logger.info("Foglio CE_Industriale scritto")


def scrivi_ce_gestionale(wb, ce_industriale: dict, ce_gestionale: dict, allocazione: dict, non_allocati: float, riepilogo_cat: dict):
    """Scrive il CE Gestionale nel foglio Excel."""
    ws = wb["CE_Gestionale"]

    _pulisci_foglio(ws)

    # Titolo
    ws.cell(row=1, column=1, value="CONTO ECONOMICO GESTIONALE - ANNO 2025")
    ws.cell(row=1, column=1).font = FONT_TITOLO
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(UO_OPERATIVE) + 1)

    uo_list = list(UO_OPERATIVE)
    col_uo_map = {}
    ws.cell(row=3, column=1, value="Codice")
    ws.cell(row=3, column=2, value="Voce")
    for i, uo in enumerate(uo_list):
        col = 3 + i
        col_uo_map[uo] = col
        ws.cell(row=3, column=col, value=uo)
    col_tot = 3 + len(uo_list)
    ws.cell(row=3, column=col_tot, value="TOTALE")
    n_col = col_tot
    _stile_header(ws, 3, n_col)

    riga = 4

    # MOL Industriale (riportato)
    ws.cell(row=riga, column=2, value="MOL INDUSTRIALE (da CE Industriale)").font = Font(
        name="Calibri", size=11, bold=True, color="006600"
    )
    ws.cell(row=riga, column=2).border = BORDO
    valori = {uo: ce_industriale[uo]["mol_industriale"] for uo in uo_list}
    _scrivi_riga_valori(ws, riga, valori, col_uo_map, font=FONT_TOTALE, fill=FILL_MOL)
    riga += 1

    # Margine % Industriale
    ws.cell(row=riga, column=2, value="Margine % Industriale").font = FONT_TOTALE
    ws.cell(row=riga, column=2).border = BORDO
    for uo, col in col_uo_map.items():
        cell = ws.cell(row=riga, column=col)
        cell.value = ce_industriale[uo]["mol_pct"]
        cell.number_format = FMT_PCT
        cell.font = FONT_TOTALE
        cell.border = BORDO
    riga += 2

    # --- COSTI SEDE ALLOCATI ---
    ws.cell(row=riga, column=2, value="COSTI SEDE ALLOCATI")
    _stile_sezione(ws, riga, n_col)
    riga += 1

    # Dettaglio per voce sede (solo quelle con importi)
    voci_con_dati = set()
    for uo in uo_list:
        voci_con_dati.update(allocazione[uo].keys())

    for codice in sorted(voci_con_dati):
        desc = VOCI_COSTI_SEDE.get(codice, codice)
        ws.cell(row=riga, column=1, value=codice).font = FONT_NORMALE
        ws.cell(row=riga, column=1).border = BORDO
        ws.cell(row=riga, column=2, value=desc).font = FONT_NORMALE
        ws.cell(row=riga, column=2).border = BORDO
        valori = {uo: allocazione[uo].get(codice, 0) for uo in uo_list}
        _scrivi_riga_valori(ws, riga, valori, col_uo_map)
        riga += 1

    # Totale costi sede allocati
    ws.cell(row=riga, column=2, value="TOTALE COSTI SEDE ALLOCATI").font = FONT_TOTALE
    ws.cell(row=riga, column=2).border = BORDO
    valori_sede = {uo: ce_gestionale[uo]["costi_sede_allocati"] for uo in uo_list}
    _scrivi_riga_valori(ws, riga, valori_sede, col_uo_map, font=FONT_TOTALE, fill=FILL_TOTALE)
    riga += 1

    # Costi sede non allocati
    ws.cell(row=riga, column=2, value="Costi sede non allocati (Sviluppo + Storici)").font = Font(
        italic=True, size=10, color="666666"
    )
    ws.cell(row=riga, column=2).border = BORDO
    ws.cell(row=riga, column=col_tot, value=round(non_allocati)).font = Font(italic=True, size=10)
    ws.cell(row=riga, column=col_tot).number_format = FMT_EURO
    ws.cell(row=riga, column=col_tot).border = BORDO
    riga += 2

    # --- MOL GESTIONALE ---
    ws.cell(row=riga, column=2, value="MOL GESTIONALE (MOL-G)").font = Font(
        name="Calibri", size=12, bold=True, color="1F4E79"
    )
    ws.cell(row=riga, column=2).border = BORDO
    valori_mol_g = {uo: ce_gestionale[uo]["mol_gestionale"] for uo in uo_list}
    for uo, col in col_uo_map.items():
        val = valori_mol_g[uo]
        cell = ws.cell(row=riga, column=col)
        cell.value = round(val)
        cell.number_format = FMT_EURO
        cell.font = FONT_POSITIVO if val >= 0 else FONT_NEGATIVO
        cell.fill = FILL_MOL_GEST
        cell.border = BORDO
    tot_mol_g = sum(valori_mol_g.values())
    cell = ws.cell(row=riga, column=col_tot)
    cell.value = round(tot_mol_g)
    cell.number_format = FMT_EURO
    cell.font = FONT_POSITIVO if tot_mol_g >= 0 else FONT_NEGATIVO
    cell.fill = FILL_MOL_GEST
    cell.border = BORDO
    riga += 1

    # Margine % Gestionale
    ws.cell(row=riga, column=2, value="MARGINE % GESTIONALE").font = FONT_TOTALE
    ws.cell(row=riga, column=2).border = BORDO
    for uo, col in col_uo_map.items():
        pct = ce_gestionale[uo]["mol_gestionale_pct"]
        cell = ws.cell(row=riga, column=col)
        cell.value = pct
        cell.number_format = FMT_PCT
        cell.font = FONT_POSITIVO if pct >= 0.08 else FONT_NEGATIVO
        cell.border = BORDO
    tot_ricavi = sum(ce_industriale[uo]["totale_ricavi"] for uo in uo_list)
    pct_tot = tot_mol_g / tot_ricavi if tot_ricavi > 0 else 0
    cell = ws.cell(row=riga, column=col_tot)
    cell.value = pct_tot
    cell.number_format = FMT_PCT
    cell.font = FONT_TOTALE
    cell.border = BORDO
    riga += 2

    # --- RIEPILOGO COSTI SEDE PER CATEGORIA ---
    ws.cell(row=riga, column=2, value="RIEPILOGO COSTI SEDE PER CATEGORIA")
    _stile_sezione(ws, riga, 5)
    riga += 1
    ws.cell(row=riga, column=2, value="Categoria").font = FONT_TOTALE
    ws.cell(row=riga, column=3, value="Importo").font = FONT_TOTALE
    ws.cell(row=riga, column=4, value="% su Sede").font = FONT_TOTALE
    ws.cell(row=riga, column=5, value="% su Ricavi").font = FONT_TOTALE
    riga += 1

    tot_sede = sum(riepilogo_cat.values())
    for cat, importo in riepilogo_cat.items():
        ws.cell(row=riga, column=2, value=cat).font = FONT_NORMALE
        ws.cell(row=riga, column=2).border = BORDO
        ws.cell(row=riga, column=3, value=round(importo)).number_format = FMT_EURO
        ws.cell(row=riga, column=3).border = BORDO
        ws.cell(row=riga, column=4, value=importo / tot_sede if tot_sede > 0 else 0).number_format = FMT_PCT
        ws.cell(row=riga, column=4).border = BORDO
        ws.cell(row=riga, column=5, value=importo / tot_ricavi if tot_ricavi > 0 else 0).number_format = FMT_PCT
        ws.cell(row=riga, column=5).border = BORDO
        riga += 1

    ws.cell(row=riga, column=2, value="TOTALE SEDE").font = FONT_TOTALE
    ws.cell(row=riga, column=2).border = BORDO
    ws.cell(row=riga, column=3, value=round(tot_sede)).number_format = FMT_EURO
    ws.cell(row=riga, column=3).font = FONT_TOTALE
    ws.cell(row=riga, column=3).border = BORDO

    # Larghezza colonne
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 45
    for i in range(3, n_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    logger.info("Foglio CE_Gestionale scritto")


def scrivi_kpi(wb, kpi_list: list):
    """Scrive i KPI nel foglio Excel."""
    ws = wb["KPI_Calcolati"]

    _pulisci_foglio(ws)

    # Titolo
    ws.cell(row=1, column=1, value="KPI - INDICATORI CHIAVE DI PERFORMANCE - ANNO 2025")
    ws.cell(row=1, column=1).font = FONT_TITOLO
    ws.merge_cells("A1:H1")

    # Header
    intestazioni = ["KPI", "Unità Operativa", "Valore", "Target", "Scostamento", "Alert", "Formula"]
    for col, titolo in enumerate(intestazioni, 1):
        ws.cell(row=3, column=col, value=titolo)
    _stile_header(ws, 3, len(intestazioni))

    riga = 4
    for kpi in kpi_list:
        ws.cell(row=riga, column=1, value=kpi["kpi"]).font = FONT_NORMALE
        ws.cell(row=riga, column=1).border = BORDO
        ws.cell(row=riga, column=2, value=kpi["unita_operativa"]).font = FONT_NORMALE
        ws.cell(row=riga, column=2).border = BORDO

        # Valore
        cell_val = ws.cell(row=riga, column=3)
        if "%" in kpi["kpi"] or "Occupancy" in kpi["kpi"]:
            cell_val.value = kpi["valore"]
            cell_val.number_format = FMT_PCT
        else:
            cell_val.value = round(kpi["valore"], 2)
            cell_val.number_format = FMT_EURO_DEC
        cell_val.font = FONT_NORMALE
        cell_val.border = BORDO

        # Target
        cell_tgt = ws.cell(row=riga, column=4)
        if "%" in kpi["kpi"] or "Occupancy" in kpi["kpi"]:
            cell_tgt.value = kpi["target"]
            cell_tgt.number_format = FMT_PCT
        else:
            cell_tgt.value = round(kpi["target"], 2)
            cell_tgt.number_format = FMT_EURO_DEC
        cell_tgt.font = FONT_NORMALE
        cell_tgt.border = BORDO

        # Scostamento
        scost = kpi["valore"] - kpi["target"]
        ws.cell(row=riga, column=5, value=round(scost, 4)).font = FONT_NORMALE
        ws.cell(row=riga, column=5).border = BORDO

        # Alert con colore
        cell_alert = ws.cell(row=riga, column=6)
        cell_alert.value = kpi["alert"]
        cell_alert.border = BORDO
        if kpi["alert"] == "VERDE":
            cell_alert.fill = FILL_VERDE
        elif kpi["alert"] == "GIALLO":
            cell_alert.fill = FILL_GIALLO
        else:
            cell_alert.fill = FILL_ROSSO

        ws.cell(row=riga, column=7, value=kpi["formula"]).font = Font(size=9, italic=True)
        ws.cell(row=riga, column=7).border = BORDO

        riga += 1

    # Larghezza
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 40

    logger.info(f"Foglio KPI_Calcolati scritto ({len(kpi_list)} KPI)")


def scrivi_dashboard_economica(wb, ce_industriale: dict, ce_gestionale: dict, kpi_list: list):
    """Scrive la Dashboard Economica con riepilogo visuale."""
    ws = wb["Dashboard_Economica"]

    _pulisci_foglio(ws)

    # Titolo
    ws.cell(row=1, column=1, value="DASHBOARD ECONOMICA - GRUPPO KAROL - ANNO 2025")
    ws.cell(row=1, column=1).font = FONT_TITOLO
    ws.merge_cells("A1:I1")

    # --- SEZIONE 1: RIEPILOGO PER UO ---
    riga = 3
    ws.cell(row=riga, column=1, value="RIEPILOGO CE PER UNITA' OPERATIVA")
    _stile_sezione(ws, riga, 9)
    riga += 1

    headers = ["UO", "Nome", "Ricavi", "Costi Dir.", "MOL-I", "MOL-I %",
               "Costi Sede", "MOL-G", "MOL-G %"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=riga, column=col, value=h)
    _stile_header(ws, riga, len(headers))
    riga += 1

    for uo in UO_OPERATIVE:
        ce_i = ce_industriale[uo]
        ce_g = ce_gestionale[uo]
        nome = UNITA_OPERATIVE[uo].nome

        ws.cell(row=riga, column=1, value=uo).font = FONT_TOTALE
        ws.cell(row=riga, column=1).border = BORDO
        ws.cell(row=riga, column=2, value=nome).font = FONT_NORMALE
        ws.cell(row=riga, column=2).border = BORDO

        for col_idx, val in [
            (3, ce_i["totale_ricavi"]),
            (4, ce_i["totale_costi"]),
            (5, ce_i["mol_industriale"]),
        ]:
            cell = ws.cell(row=riga, column=col_idx)
            cell.value = round(val)
            cell.number_format = FMT_EURO
            cell.font = FONT_NORMALE
            cell.border = BORDO

        # MOL-I %
        cell = ws.cell(row=riga, column=6)
        cell.value = ce_i["mol_pct"]
        cell.number_format = FMT_PCT
        cell.border = BORDO
        cell.fill = FILL_VERDE if ce_i["mol_pct"] >= 0.15 else (FILL_GIALLO if ce_i["mol_pct"] >= 0.08 else FILL_ROSSO)

        # Costi Sede
        cell = ws.cell(row=riga, column=7)
        cell.value = round(ce_g["costi_sede_allocati"])
        cell.number_format = FMT_EURO
        cell.font = FONT_NORMALE
        cell.border = BORDO

        # MOL-G
        cell = ws.cell(row=riga, column=8)
        cell.value = round(ce_g["mol_gestionale"])
        cell.number_format = FMT_EURO
        cell.font = FONT_POSITIVO if ce_g["mol_gestionale"] >= 0 else FONT_NEGATIVO
        cell.border = BORDO

        # MOL-G %
        cell = ws.cell(row=riga, column=9)
        cell.value = ce_g["mol_gestionale_pct"]
        cell.number_format = FMT_PCT
        cell.border = BORDO
        cell.fill = FILL_VERDE if ce_g["mol_gestionale_pct"] >= 0.08 else (FILL_GIALLO if ce_g["mol_gestionale_pct"] >= 0 else FILL_ROSSO)

        riga += 1

    # Riga totale
    tot_r = sum(ce_industriale[uo]["totale_ricavi"] for uo in UO_OPERATIVE)
    tot_c = sum(ce_industriale[uo]["totale_costi"] for uo in UO_OPERATIVE)
    tot_mol_i = sum(ce_industriale[uo]["mol_industriale"] for uo in UO_OPERATIVE)
    tot_sede = sum(ce_gestionale[uo]["costi_sede_allocati"] for uo in UO_OPERATIVE)
    tot_mol_g = sum(ce_gestionale[uo]["mol_gestionale"] for uo in UO_OPERATIVE)

    ws.cell(row=riga, column=1, value="TOTALE").font = FONT_TOTALE
    ws.cell(row=riga, column=1).fill = FILL_TOTALE
    ws.cell(row=riga, column=1).border = BORDO
    ws.cell(row=riga, column=2, value="").fill = FILL_TOTALE
    ws.cell(row=riga, column=2).border = BORDO

    for col_idx, val in [(3, tot_r), (4, tot_c), (5, tot_mol_i), (7, tot_sede), (8, tot_mol_g)]:
        cell = ws.cell(row=riga, column=col_idx)
        cell.value = round(val)
        cell.number_format = FMT_EURO
        cell.font = FONT_TOTALE
        cell.fill = FILL_TOTALE
        cell.border = BORDO

    cell = ws.cell(row=riga, column=6)
    cell.value = tot_mol_i / tot_r if tot_r > 0 else 0
    cell.number_format = FMT_PCT
    cell.font = FONT_TOTALE
    cell.fill = FILL_TOTALE
    cell.border = BORDO

    cell = ws.cell(row=riga, column=9)
    cell.value = tot_mol_g / tot_r if tot_r > 0 else 0
    cell.number_format = FMT_PCT
    cell.font = FONT_TOTALE
    cell.fill = FILL_TOTALE
    cell.border = BORDO

    riga += 3

    # --- SEZIONE 2: KPI CON SEMAFORI ---
    ws.cell(row=riga, column=1, value="KPI PRINCIPALI CON SEMAFORI")
    _stile_sezione(ws, riga, 6)
    riga += 1

    kpi_headers = ["KPI", "UO", "Valore", "Target", "Alert"]
    for col, h in enumerate(kpi_headers, 1):
        ws.cell(row=riga, column=col, value=h)
    _stile_header(ws, riga, len(kpi_headers))
    riga += 1

    for kpi in kpi_list:
        ws.cell(row=riga, column=1, value=kpi["kpi"]).font = FONT_NORMALE
        ws.cell(row=riga, column=1).border = BORDO
        ws.cell(row=riga, column=2, value=kpi["unita_operativa"]).font = FONT_NORMALE
        ws.cell(row=riga, column=2).border = BORDO

        cell = ws.cell(row=riga, column=3)
        if "%" in kpi["kpi"] or "Occupancy" in kpi["kpi"]:
            cell.value = kpi["valore"]
            cell.number_format = FMT_PCT
        else:
            cell.value = round(kpi["valore"], 2)
            cell.number_format = FMT_EURO_DEC
        cell.font = FONT_NORMALE
        cell.border = BORDO

        cell = ws.cell(row=riga, column=4)
        if "%" in kpi["kpi"] or "Occupancy" in kpi["kpi"]:
            cell.value = kpi["target"]
            cell.number_format = FMT_PCT
        else:
            cell.value = round(kpi["target"], 2)
        cell.font = FONT_NORMALE
        cell.border = BORDO

        cell_a = ws.cell(row=riga, column=5)
        cell_a.value = kpi["alert"]
        cell_a.border = BORDO
        cell_a.font = FONT_TOTALE
        if kpi["alert"] == "VERDE":
            cell_a.fill = FILL_VERDE
        elif kpi["alert"] == "GIALLO":
            cell_a.fill = FILL_GIALLO
        else:
            cell_a.fill = FILL_ROSSO

        riga += 1

    # Larghezza
    for col_idx, w in {1: 30, 2: 22, 3: 16, 4: 16, 5: 16, 6: 14, 7: 14, 8: 14, 9: 14}.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    logger.info("Dashboard Economica scritta")


# ============================================================================
# FUNZIONE PRINCIPALE
# ============================================================================

def elabora_completo(file_path: Path = None, scrivi_excel: bool = True):
    """Esegue l'elaborazione completa e opzionalmente scrive i risultati nell'Excel Master."""
    if file_path is None:
        file_path = EXCEL_MASTER

    print("\n" + "=" * 70)
    print("ELABORAZIONE CONTROLLO DI GESTIONE - GRUPPO KAROL")
    print("=" * 70)

    # 1. Lettura dati
    dati = leggi_dati_master(file_path)

    # 2. CE Industriale
    print("\n--- CALCOLO CE INDUSTRIALE ---")
    ce_industriale = calcola_ce_industriale_da_dati(dati)

    # 3. Allocazione costi sede
    print("\n--- ALLOCAZIONE COSTI SEDE ---")
    allocazione, non_allocati, riepilogo_cat = calcola_allocazione_sede(dati, ce_industriale)

    # 4. CE Gestionale
    print("\n--- CALCOLO CE GESTIONALE ---")
    ce_gestionale = calcola_ce_gestionale_da_dati(ce_industriale, allocazione, non_allocati)

    # 5. KPI
    print("\n--- CALCOLO KPI ---")
    kpi_list = calcola_kpi(ce_industriale, ce_gestionale, dati)
    logger.info(f"Calcolati {len(kpi_list)} KPI")

    # 6. Scrittura Excel (opzionale)
    if scrivi_excel:
        print("\n--- SCRITTURA RISULTATI SU EXCEL ---")
        wb = openpyxl.load_workbook(file_path)

        scrivi_ce_industriale(wb, ce_industriale)
        scrivi_ce_gestionale(wb, ce_industriale, ce_gestionale, allocazione, non_allocati, riepilogo_cat)
        scrivi_kpi(wb, kpi_list)
        scrivi_dashboard_economica(wb, ce_industriale, ce_gestionale, kpi_list)

        wb.save(file_path)
        print(f"\nRisultati salvati in: {file_path}")
        print(f"Dimensione file: {file_path.stat().st_size / 1024:.1f} KB")

    # Riepilogo finale
    print("\n" + "=" * 70)
    print("RIEPILOGO RISULTATI")
    print("=" * 70)

    tot_r = sum(ce_industriale[uo]["totale_ricavi"] for uo in UO_OPERATIVE)
    tot_mol_i = sum(ce_industriale[uo]["mol_industriale"] for uo in UO_OPERATIVE)
    tot_sede_alloc = sum(ce_gestionale[uo]["costi_sede_allocati"] for uo in UO_OPERATIVE)
    tot_mol_g = sum(ce_gestionale[uo]["mol_gestionale"] for uo in UO_OPERATIVE)

    print(f"\nRicavi Totali:           {tot_r:>14,.0f}")
    print(f"MOL Industriale:         {tot_mol_i:>14,.0f}  ({tot_mol_i/tot_r:.1%})")
    print(f"Costi Sede Allocati:    -{tot_sede_alloc:>14,.0f}  ({tot_sede_alloc/tot_r:.1%})")
    print(f"Costi Sede Non Allocati:  {non_allocati:>14,.0f}  (Sviluppo+Storici)")
    print(f"MOL Gestionale:          {tot_mol_g:>14,.0f}  ({tot_mol_g/tot_r:.1%})")

    print(f"\nFogli aggiornati: CE_Industriale, CE_Gestionale, KPI_Calcolati, Dashboard_Economica")
    print("=" * 70)

    return {
        "ce_industriale": ce_industriale,
        "ce_gestionale": ce_gestionale,
        "allocazione": allocazione,
        "kpi": kpi_list,
        "non_allocati": non_allocati,
        "riepilogo_cat": riepilogo_cat,
    }


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
    )
    elabora_completo()
