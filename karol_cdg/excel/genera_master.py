"""
Generazione del file Excel Master (KAROL_CDG_MASTER.xlsx).
Crea la struttura completa con tutti i fogli, intestazioni, formattazione e dati iniziali.
"""

import logging
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from karol_cdg.config import (
    UNITA_OPERATIVE, VOCI_RICAVI, VOCI_COSTI_DIRETTI, VOCI_COSTI_SEDE,
    VOCI_ALTRI_COSTI, DRIVER_PREDEFINITI, BENCHMARK, ALERT_CONFIG,
    SOGLIE_SEMAFORO, SCENARI_CASH_FLOW, MESI_BREVI_IT, MESI_IT,
    CategoriaCostoSede, DriverAllocazione, QualificaPersonale,
    EXCEL_MASTER, DATA_DIR, BACKUP_DIR, OUTPUT_DIR,
)

logger = logging.getLogger(__name__)

# ============================================================================
# STILI
# ============================================================================

FONT_TITOLO = Font(name="Calibri", size=14, bold=True, color="1F4E79")
FONT_INTESTAZIONE = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_SEZIONE = Font(name="Calibri", size=11, bold=True, color="1F4E79")
FONT_NORMALE = Font(name="Calibri", size=10)
FONT_TOTALE = Font(name="Calibri", size=10, bold=True)

FILL_INTESTAZIONE = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FILL_SEZIONE = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_TOTALE = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_ALERT_ROSSO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_ALERT_GIALLO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_ALERT_VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_INPUT = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

BORDO_SOTTILE = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

ALLINEAMENTO_CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALLINEAMENTO_DESTRA = Alignment(horizontal="right", vertical="center")
ALLINEAMENTO_SINISTRA = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _applica_stile_intestazione(ws, riga, num_colonne):
    """Applica stile intestazione a una riga."""
    for col in range(1, num_colonne + 1):
        cella = ws.cell(row=riga, column=col)
        cella.font = FONT_INTESTAZIONE
        cella.fill = FILL_INTESTAZIONE
        cella.alignment = ALLINEAMENTO_CENTRO
        cella.border = BORDO_SOTTILE


def _applica_stile_sezione(ws, riga, num_colonne):
    """Applica stile sezione a una riga."""
    for col in range(1, num_colonne + 1):
        cella = ws.cell(row=riga, column=col)
        cella.font = FONT_SEZIONE
        cella.fill = FILL_SEZIONE
        cella.border = BORDO_SOTTILE


def _imposta_larghezza_colonne(ws, larghezze: dict):
    """Imposta la larghezza delle colonne."""
    for col_idx, larghezza in larghezze.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = larghezza


# ============================================================================
# FOGLI DATI
# ============================================================================

def _crea_foglio_anagrafiche_uo(wb):
    """Crea il foglio Anagrafiche_UO con i dati delle unità operative."""
    ws = wb.create_sheet("Anagrafiche_UO")
    intestazioni = [
        "Codice", "Nome", "Tipologia", "Regione", "Posti Letto",
        "Società", "Attiva", "Note"
    ]

    for col, titolo in enumerate(intestazioni, 1):
        ws.cell(row=1, column=col, value=titolo)
    _applica_stile_intestazione(ws, 1, len(intestazioni))

    for riga, (codice, uo) in enumerate(UNITA_OPERATIVE.items(), 2):
        ws.cell(row=riga, column=1, value=uo.codice)
        ws.cell(row=riga, column=2, value=uo.nome)
        ws.cell(row=riga, column=3, value=", ".join(t.value for t in uo.tipologia))
        ws.cell(row=riga, column=4, value=uo.regione.value)
        ws.cell(row=riga, column=5, value=uo.posti_letto)
        ws.cell(row=riga, column=6, value=uo.societa)
        ws.cell(row=riga, column=7, value="Sì" if uo.attiva else "No")
        ws.cell(row=riga, column=8, value=uo.note)
        for col in range(1, len(intestazioni) + 1):
            ws.cell(row=riga, column=col).font = FONT_NORMALE
            ws.cell(row=riga, column=col).border = BORDO_SOTTILE

    _imposta_larghezza_colonne(ws, {1: 10, 2: 30, 3: 40, 4: 12, 5: 12, 6: 20, 7: 8, 8: 50})
    logger.info("Foglio Anagrafiche_UO creato")


def _crea_foglio_piano_conti(wb):
    """Crea il foglio Piano_Conti (vuoto, da importare da E-Solver)."""
    ws = wb.create_sheet("Piano_Conti")
    intestazioni = [
        "Codice Conto", "Descrizione", "Tipo (Ricavo/Costo)",
        "Gruppo", "Sottogruppo", "Centro di Costo", "Note"
    ]
    for col, titolo in enumerate(intestazioni, 1):
        ws.cell(row=1, column=col, value=titolo)
    _applica_stile_intestazione(ws, 1, len(intestazioni))

    # Riga esempio
    ws.cell(row=2, column=1, value="[Import da E-Solver]")
    ws.cell(row=2, column=1).font = Font(italic=True, color="999999")

    _imposta_larghezza_colonne(ws, {1: 15, 2: 40, 3: 18, 4: 15, 5: 15, 6: 20, 7: 30})
    logger.info("Foglio Piano_Conti creato")


def _crea_foglio_personale(wb):
    """Crea il foglio Anagrafiche_Personale."""
    ws = wb.create_sheet("Anagrafiche_Personale")
    intestazioni = [
        "Matricola", "Cognome", "Nome", "Qualifica", "Unità Operativa",
        "Costo Lordo", "Contributi", "TFR", "Costo Totale",
        "Ore Ordinarie", "Ore Straordinarie", "FTE", "Mese", "Anno"
    ]
    for col, titolo in enumerate(intestazioni, 1):
        ws.cell(row=1, column=col, value=titolo)
    _applica_stile_intestazione(ws, 1, len(intestazioni))

    ws.cell(row=2, column=1, value="[Import da Zucchetti]")
    ws.cell(row=2, column=1).font = Font(italic=True, color="999999")

    _imposta_larghezza_colonne(ws, {
        1: 12, 2: 18, 3: 15, 4: 20, 5: 15,
        6: 14, 7: 14, 8: 12, 9: 14,
        10: 14, 11: 16, 12: 8, 13: 8, 14: 8
    })
    logger.info("Foglio Anagrafiche_Personale creato")


def _crea_foglio_produzione_mensile(wb):
    """Crea il foglio Produzione_Mensile."""
    ws = wb.create_sheet("Produzione_Mensile")
    intestazioni = [
        "Unità Operativa", "Mese", "Anno",
        "Tipo Prestazione", "Codice Prestazione", "Descrizione",
        "Quantità", "Tariffa Unitaria", "Importo Totale",
        "Giornate Degenza", "Posti Letto Occupati", "Fonte Dati"
    ]
    for col, titolo in enumerate(intestazioni, 1):
        ws.cell(row=1, column=col, value=titolo)
    _applica_stile_intestazione(ws, 1, len(intestazioni))

    ws.cell(row=2, column=1, value="[Import da Caremed / HT Sang / Template manuale]")
    ws.cell(row=2, column=1).font = Font(italic=True, color="999999")

    _imposta_larghezza_colonne(ws, {
        1: 18, 2: 8, 3: 8, 4: 20, 5: 18, 6: 35,
        7: 12, 8: 15, 9: 15, 10: 16, 11: 18, 12: 15
    })
    logger.info("Foglio Produzione_Mensile creato")


def _crea_foglio_costi_mensili(wb):
    """Crea il foglio Costi_Mensili."""
    ws = wb.create_sheet("Costi_Mensili")
    intestazioni = [
        "Unità Operativa", "Mese", "Anno",
        "Codice Voce", "Descrizione Voce", "Categoria",
        "Importo", "Note", "Fonte Dati"
    ]
    for col, titolo in enumerate(intestazioni, 1):
        ws.cell(row=1, column=col, value=titolo)
    _applica_stile_intestazione(ws, 1, len(intestazioni))

    # Pre-popola le voci di costo per ogni UO come guida
    riga = 2
    ws.cell(row=riga, column=1, value="--- STRUTTURA VOCI COSTI DIRETTI ---")
    ws.cell(row=riga, column=1).font = FONT_SEZIONE
    riga += 1
    for codice, desc in VOCI_COSTI_DIRETTI.items():
        ws.cell(row=riga, column=4, value=codice)
        ws.cell(row=riga, column=5, value=desc)
        ws.cell(row=riga, column=6, value="Costo Diretto")
        for col in range(4, 7):
            ws.cell(row=riga, column=col).font = Font(italic=True, color="666666", size=9)
        riga += 1

    _imposta_larghezza_colonne(ws, {
        1: 18, 2: 8, 3: 8, 4: 12, 5: 35, 6: 18, 7: 15, 8: 30, 9: 15
    })
    logger.info("Foglio Costi_Mensili creato")


def _crea_foglio_costi_sede(wb):
    """Crea il foglio Costi_Sede_Dettaglio con classificazione per categoria."""
    ws = wb.create_sheet("Costi_Sede_Dettaglio")
    intestazioni = [
        "Codice Conto", "Descrizione", "Importo Annuo",
        "Categoria", "Sotto-Categoria",
        "Driver Allocazione", "Note", "Da Verificare"
    ]
    for col, titolo in enumerate(intestazioni, 1):
        ws.cell(row=1, column=col, value=titolo)
    _applica_stile_intestazione(ws, 1, len(intestazioni))

    # Pre-popola le voci di costo sede
    riga = 2
    for codice, desc in VOCI_COSTI_SEDE.items():
        ws.cell(row=riga, column=1, value=codice)
        ws.cell(row=riga, column=2, value=desc)
        ws.cell(row=riga, column=3).fill = FILL_INPUT
        ws.cell(row=riga, column=3).number_format = '#.##0,00 €'

        # Categoria automatica
        if codice.startswith("CS0"):
            cat = CategoriaCostoSede.SERVIZI.value
        elif codice.startswith("CS1"):
            cat = CategoriaCostoSede.GOVERNANCE.value
        else:
            cat = CategoriaCostoSede.DA_CLASSIFICARE.value
        ws.cell(row=riga, column=4, value=cat)

        # Driver predefinito
        driver = DRIVER_PREDEFINITI.get(codice)
        if driver:
            ws.cell(row=riga, column=6, value=driver.value)

        for col in range(1, len(intestazioni) + 1):
            ws.cell(row=riga, column=col).font = FONT_NORMALE
            ws.cell(row=riga, column=col).border = BORDO_SOTTILE

        riga += 1

    # Spazio per voci aggiuntive
    for i in range(20):
        for col in range(1, len(intestazioni) + 1):
            ws.cell(row=riga + i, column=col).fill = FILL_INPUT
            ws.cell(row=riga + i, column=col).border = BORDO_SOTTILE

    _imposta_larghezza_colonne(ws, {
        1: 15, 2: 35, 3: 18, 4: 25, 5: 20, 6: 20, 7: 30, 8: 14
    })

    # Aggiungi legenda categorie
    riga_legenda = riga + 22
    ws.cell(row=riga_legenda, column=1, value="LEGENDA CATEGORIE")
    ws.cell(row=riga_legenda, column=1).font = FONT_SEZIONE
    for i, cat in enumerate(CategoriaCostoSede):
        ws.cell(row=riga_legenda + 1 + i, column=1, value=cat.value)

    riga_driver = riga_legenda + 1 + len(CategoriaCostoSede)
    ws.cell(row=riga_driver + 1, column=1, value="LEGENDA DRIVER")
    ws.cell(row=riga_driver + 1, column=1).font = FONT_SEZIONE
    for i, drv in enumerate(DriverAllocazione):
        ws.cell(row=riga_driver + 2 + i, column=1, value=drv.value)

    logger.info("Foglio Costi_Sede_Dettaglio creato")


def _crea_foglio_scadenzario(wb):
    """Crea il foglio Scadenzario per incassi e pagamenti."""
    ws = wb.create_sheet("Scadenzario")
    intestazioni = [
        "Data Scadenza", "Tipo (Incasso/Pagamento)", "Categoria",
        "Importo", "Controparte", "Unità Operativa",
        "Stato (Previsto/Confermato/Pagato)", "Note"
    ]
    for col, titolo in enumerate(intestazioni, 1):
        ws.cell(row=1, column=col, value=titolo)
    _applica_stile_intestazione(ws, 1, len(intestazioni))

    # Righe input vuote formattate
    for riga in range(2, 102):
        ws.cell(row=riga, column=1).number_format = "DD/MM/YYYY"
        ws.cell(row=riga, column=4).number_format = '#.##0,00 €'
        for col in range(1, len(intestazioni) + 1):
            ws.cell(row=riga, column=col).fill = FILL_INPUT
            ws.cell(row=riga, column=col).border = BORDO_SOTTILE

    _imposta_larghezza_colonne(ws, {
        1: 15, 2: 22, 3: 20, 4: 15, 5: 25, 6: 18, 7: 28, 8: 30
    })
    logger.info("Foglio Scadenzario creato")


# ============================================================================
# FOGLI ELABORAZIONI
# ============================================================================

def _crea_foglio_ce_industriale(wb):
    """Crea il foglio CE_Industriale (struttura per output calcoli)."""
    ws = wb.create_sheet("CE_Industriale")

    # Titolo
    ws.cell(row=1, column=1, value="CONTO ECONOMICO INDUSTRIALE PER UNITÀ OPERATIVA")
    ws.cell(row=1, column=1).font = FONT_TITOLO
    ws.merge_cells("A1:N1")

    # Intestazione colonne: Voce + una colonna per ogni UO + Totale
    uo_list = list(UNITA_OPERATIVE.keys())
    intestazioni = ["Codice", "Voce"] + [uo for uo in uo_list] + ["TOTALE"]

    riga = 3
    for col, titolo in enumerate(intestazioni, 1):
        ws.cell(row=riga, column=col, value=titolo)
    _applica_stile_intestazione(ws, riga, len(intestazioni))

    # Sezione RICAVI
    riga += 1
    ws.cell(row=riga, column=1, value="")
    ws.cell(row=riga, column=2, value="RICAVI")
    _applica_stile_sezione(ws, riga, len(intestazioni))

    for codice, desc in VOCI_RICAVI.items():
        riga += 1
        ws.cell(row=riga, column=1, value=codice)
        ws.cell(row=riga, column=2, value=desc)
        ws.cell(row=riga, column=1).font = FONT_NORMALE
        ws.cell(row=riga, column=2).font = FONT_NORMALE

    riga += 1
    ws.cell(row=riga, column=2, value="TOTALE RICAVI")
    ws.cell(row=riga, column=2).font = FONT_TOTALE
    for col in range(1, len(intestazioni) + 1):
        ws.cell(row=riga, column=col).fill = FILL_TOTALE
        ws.cell(row=riga, column=col).border = BORDO_SOTTILE

    # Sezione COSTI DIRETTI
    riga += 2
    ws.cell(row=riga, column=2, value="COSTI DIRETTI")
    _applica_stile_sezione(ws, riga, len(intestazioni))

    for codice, desc in VOCI_COSTI_DIRETTI.items():
        riga += 1
        ws.cell(row=riga, column=1, value=codice)
        ws.cell(row=riga, column=2, value=desc)
        ws.cell(row=riga, column=1).font = FONT_NORMALE
        ws.cell(row=riga, column=2).font = FONT_NORMALE

    riga += 1
    ws.cell(row=riga, column=2, value="TOTALE COSTI DIRETTI")
    ws.cell(row=riga, column=2).font = FONT_TOTALE
    for col in range(1, len(intestazioni) + 1):
        ws.cell(row=riga, column=col).fill = FILL_TOTALE
        ws.cell(row=riga, column=col).border = BORDO_SOTTILE

    # MOL Industriale
    riga += 2
    ws.cell(row=riga, column=2, value="MOL INDUSTRIALE (MOL-I)")
    ws.cell(row=riga, column=2).font = Font(name="Calibri", size=12, bold=True, color="006600")
    for col in range(1, len(intestazioni) + 1):
        ws.cell(row=riga, column=col).fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        ws.cell(row=riga, column=col).border = BORDO_SOTTILE

    riga += 1
    ws.cell(row=riga, column=2, value="MARGINE % INDUSTRIALE")
    ws.cell(row=riga, column=2).font = FONT_TOTALE

    _imposta_larghezza_colonne(ws, {1: 10, 2: 40})
    for i in range(3, 3 + len(uo_list) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    logger.info("Foglio CE_Industriale creato")


def _crea_foglio_ce_gestionale(wb):
    """Crea il foglio CE_Gestionale (struttura per output calcoli)."""
    ws = wb.create_sheet("CE_Gestionale")

    ws.cell(row=1, column=1, value="CONTO ECONOMICO GESTIONALE PER UNITÀ OPERATIVA")
    ws.cell(row=1, column=1).font = FONT_TITOLO
    ws.merge_cells("A1:N1")

    uo_list = list(UNITA_OPERATIVE.keys())
    intestazioni = ["Codice", "Voce"] + [uo for uo in uo_list] + ["TOTALE"]

    riga = 3
    for col, titolo in enumerate(intestazioni, 1):
        ws.cell(row=riga, column=col, value=titolo)
    _applica_stile_intestazione(ws, riga, len(intestazioni))

    # MOL Industriale (da CE Industriale)
    riga += 1
    ws.cell(row=riga, column=2, value="MOL INDUSTRIALE (da CE Industriale)")
    ws.cell(row=riga, column=2).font = Font(name="Calibri", size=11, bold=True, color="006600")

    # Costi Sede Allocati
    riga += 2
    ws.cell(row=riga, column=2, value="COSTI SEDE ALLOCATI")
    _applica_stile_sezione(ws, riga, len(intestazioni))

    for codice, desc in VOCI_COSTI_SEDE.items():
        riga += 1
        ws.cell(row=riga, column=1, value=codice)
        ws.cell(row=riga, column=2, value=desc)

    riga += 1
    ws.cell(row=riga, column=2, value="TOTALE COSTI SEDE ALLOCATI")
    ws.cell(row=riga, column=2).font = FONT_TOTALE
    for col in range(1, len(intestazioni) + 1):
        ws.cell(row=riga, column=col).fill = FILL_TOTALE

    # MOL Gestionale
    riga += 2
    ws.cell(row=riga, column=2, value="MOL GESTIONALE (MOL-G)")
    ws.cell(row=riga, column=2).font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    for col in range(1, len(intestazioni) + 1):
        ws.cell(row=riga, column=col).fill = PatternFill(
            start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"
        )

    riga += 1
    ws.cell(row=riga, column=2, value="MARGINE % GESTIONALE")
    ws.cell(row=riga, column=2).font = FONT_TOTALE

    # Altri Costi
    riga += 2
    ws.cell(row=riga, column=2, value="ALTRI COSTI")
    _applica_stile_sezione(ws, riga, len(intestazioni))

    for codice, desc in VOCI_ALTRI_COSTI.items():
        riga += 1
        ws.cell(row=riga, column=1, value=codice)
        ws.cell(row=riga, column=2, value=desc)

    # Risultato Netto
    riga += 2
    ws.cell(row=riga, column=2, value="RISULTATO NETTO UNITÀ OPERATIVA")
    ws.cell(row=riga, column=2).font = Font(name="Calibri", size=12, bold=True)
    for col in range(1, len(intestazioni) + 1):
        ws.cell(row=riga, column=col).fill = PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        )
        ws.cell(row=riga, column=col).border = BORDO_SOTTILE

    _imposta_larghezza_colonne(ws, {1: 10, 2: 40})
    for i in range(3, 3 + len(uo_list) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    logger.info("Foglio CE_Gestionale creato")


def _crea_foglio_ce_consolidato(wb):
    """Crea il foglio CE_Consolidato."""
    ws = wb.create_sheet("CE_Consolidato")
    ws.cell(row=1, column=1, value="CONTO ECONOMICO CONSOLIDATO - GRUPPO KAROL")
    ws.cell(row=1, column=1).font = FONT_TITOLO
    ws.merge_cells("A1:G1")

    intestazioni = [
        "Voce", "Anno Corrente", "Anno Precedente",
        "Delta €", "Delta %", "Budget", "Scost. Budget %"
    ]
    riga = 3
    for col, titolo in enumerate(intestazioni, 1):
        ws.cell(row=riga, column=col, value=titolo)
    _applica_stile_intestazione(ws, riga, len(intestazioni))

    ws.cell(row=4, column=1, value="[Generato automaticamente dall'elaborazione]")
    ws.cell(row=4, column=1).font = Font(italic=True, color="999999")

    _imposta_larghezza_colonne(ws, {1: 40, 2: 18, 3: 18, 4: 15, 5: 12, 6: 18, 7: 15})
    logger.info("Foglio CE_Consolidato creato")


def _crea_foglio_cash_flow_operativo(wb):
    """Crea il foglio Cash_Flow_Operativo (1-3 mesi, settimanale/mensile)."""
    ws = wb.create_sheet("Cash_Flow_Operativo")
    ws.cell(row=1, column=1, value="CASH FLOW OPERATIVO - ROLLING 1-3 MESI")
    ws.cell(row=1, column=1).font = FONT_TITOLO
    ws.merge_cells("A1:H1")

    intestazioni = [
        "Voce", "Sett.1", "Sett.2", "Sett.3", "Sett.4",
        "Mese 2", "Mese 3", "Totale"
    ]
    riga = 3
    for col, titolo in enumerate(intestazioni, 1):
        ws.cell(row=riga, column=col, value=titolo)
    _applica_stile_intestazione(ws, riga, len(intestazioni))

    voci = [
        ("", "CASSA INIZIALE", True),
        ("", "", False),
        ("", "INCASSI", True),
        ("", "Incassi ASP/SSN", False),
        ("", "Incassi privati", False),
        ("", "Incassi crediti pregressi", False),
        ("", "Altri incassi", False),
        ("", "TOTALE INCASSI", True),
        ("", "", False),
        ("", "PAGAMENTI", True),
        ("", "Stipendi (27 del mese)", False),
        ("", "Contributi (16 del mese)", False),
        ("", "Fornitori", False),
        ("", "Rate mutui/leasing", False),
        ("", "Tributi (F24)", False),
        ("", "Altri pagamenti", False),
        ("", "TOTALE PAGAMENTI", True),
        ("", "", False),
        ("", "SALDO PERIODO", True),
        ("", "CASSA FINALE", True),
    ]

    for i, (_, voce, is_bold) in enumerate(voci):
        riga_corrente = riga + 1 + i
        ws.cell(row=riga_corrente, column=1, value=voce)
        if is_bold:
            ws.cell(row=riga_corrente, column=1).font = FONT_TOTALE
            if voce in ("TOTALE INCASSI", "TOTALE PAGAMENTI", "CASSA FINALE"):
                for col in range(1, len(intestazioni) + 1):
                    ws.cell(row=riga_corrente, column=col).fill = FILL_TOTALE

    _imposta_larghezza_colonne(ws, {1: 30})
    for i in range(2, len(intestazioni) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    logger.info("Foglio Cash_Flow_Operativo creato")


def _crea_foglio_cash_flow_strategico(wb):
    """Crea il foglio Cash_Flow_Strategico (12-60 mesi, annuale)."""
    ws = wb.create_sheet("Cash_Flow_Strategico")
    ws.cell(row=1, column=1, value="CASH FLOW STRATEGICO - PIANO 5 ANNI")
    ws.cell(row=1, column=1).font = FONT_TITOLO
    ws.merge_cells("A1:F1")

    intestazioni = ["Voce", "Anno 1", "Anno 2", "Anno 3", "Anno 4", "Anno 5"]
    riga = 3
    for col, titolo in enumerate(intestazioni, 1):
        ws.cell(row=riga, column=col, value=titolo)
    _applica_stile_intestazione(ws, riga, len(intestazioni))

    voci = [
        "EBITDA",
        "+/- Variazione CCN",
        "- CAPEX",
        "= FREE CASH FLOW OPERATIVO",
        "",
        "- Servizio debito",
        "- Imposte",
        "= FREE CASH FLOW NETTO",
        "",
        "PFN iniziale",
        "PFN finale",
        "PFN/EBITDA",
    ]

    for i, voce in enumerate(voci):
        riga_corrente = riga + 1 + i
        ws.cell(row=riga_corrente, column=1, value=voce)
        if voce.startswith("=") or voce.startswith("PFN"):
            ws.cell(row=riga_corrente, column=1).font = FONT_TOTALE
            for col in range(1, len(intestazioni) + 1):
                ws.cell(row=riga_corrente, column=col).fill = FILL_TOTALE

    _imposta_larghezza_colonne(ws, {1: 35})
    for i in range(2, 7):
        ws.column_dimensions[get_column_letter(i)].width = 16

    logger.info("Foglio Cash_Flow_Strategico creato")


def _crea_foglio_kpi(wb):
    """Crea il foglio KPI_Calcolati."""
    ws = wb.create_sheet("KPI_Calcolati")
    ws.cell(row=1, column=1, value="KPI - INDICATORI CHIAVE DI PERFORMANCE")
    ws.cell(row=1, column=1).font = FONT_TITOLO
    ws.merge_cells("A1:H1")

    intestazioni = [
        "KPI", "Unità Operativa", "Valore", "Target",
        "Scostamento", "Alert", "Formula", "Periodo"
    ]
    riga = 3
    for col, titolo in enumerate(intestazioni, 1):
        ws.cell(row=riga, column=col, value=titolo)
    _applica_stile_intestazione(ws, riga, len(intestazioni))

    ws.cell(row=4, column=1, value="[Generato automaticamente dall'elaborazione]")
    ws.cell(row=4, column=1).font = Font(italic=True, color="999999")

    _imposta_larghezza_colonne(ws, {
        1: 25, 2: 18, 3: 14, 4: 14, 5: 14, 6: 10, 7: 40, 8: 12
    })
    logger.info("Foglio KPI_Calcolati creato")


# ============================================================================
# FOGLI CONFIGURAZIONE
# ============================================================================

def _crea_foglio_driver_allocazione(wb):
    """Crea il foglio Driver_Allocazione con regole di ribaltamento."""
    ws = wb.create_sheet("Driver_Allocazione")
    ws.cell(row=1, column=1, value="REGOLE DI ALLOCAZIONE COSTI SEDE")
    ws.cell(row=1, column=1).font = FONT_TITOLO
    ws.merge_cells("A1:L1")

    # Intestazione: Voce + Driver + valore driver per ogni UO
    uo_list = list(UNITA_OPERATIVE.keys())
    intestazioni = ["Codice Voce", "Descrizione", "Driver"] + \
                   [f"{uo}" for uo in uo_list]

    riga = 3
    for col, titolo in enumerate(intestazioni, 1):
        ws.cell(row=riga, column=col, value=titolo)
    _applica_stile_intestazione(ws, riga, len(intestazioni))

    # Pre-popola con voci sede e driver predefiniti
    riga = 4
    for codice, desc in VOCI_COSTI_SEDE.items():
        ws.cell(row=riga, column=1, value=codice)
        ws.cell(row=riga, column=2, value=desc)
        driver = DRIVER_PREDEFINITI.get(codice)
        ws.cell(row=riga, column=3, value=driver.value if driver else "")
        # Celle input per valori driver per UO
        for col_idx in range(4, 4 + len(uo_list)):
            ws.cell(row=riga, column=col_idx).fill = FILL_INPUT
            ws.cell(row=riga, column=col_idx).border = BORDO_SOTTILE
        riga += 1

    _imposta_larghezza_colonne(ws, {1: 14, 2: 35, 3: 20})
    for i in range(4, 4 + len(uo_list)):
        ws.column_dimensions[get_column_letter(i)].width = 12

    logger.info("Foglio Driver_Allocazione creato")


def _crea_foglio_benchmark(wb):
    """Crea il foglio Benchmark_Settore con dati di riferimento."""
    ws = wb.create_sheet("Benchmark_Settore")
    ws.cell(row=1, column=1, value="BENCHMARK DI SETTORE - COSTI STANDARD")
    ws.cell(row=1, column=1).font = FONT_TITOLO
    ws.merge_cells("A1:F1")

    intestazioni = [
        "Tipologia", "Costo Pers./Ricavi Min%", "Costo Pers./Ricavi Max%",
        "MOL% Target Min", "MOL% Target Max",
        "Costo/Giornata Min €", "Costo/Giornata Max €"
    ]
    riga = 3
    for col, titolo in enumerate(intestazioni, 1):
        ws.cell(row=riga, column=col, value=titolo)
    _applica_stile_intestazione(ws, riga, len(intestazioni))

    for i, (key, bm) in enumerate(BENCHMARK.items(), 4):
        ws.cell(row=i, column=1, value=bm.tipologia.value)
        ws.cell(row=i, column=2, value=bm.costo_personale_su_ricavi_min)
        ws.cell(row=i, column=3, value=bm.costo_personale_su_ricavi_max)
        ws.cell(row=i, column=4, value=bm.mol_percentuale_target_min)
        ws.cell(row=i, column=5, value=bm.mol_percentuale_target_max)
        ws.cell(row=i, column=6, value=bm.costo_giornata_degenza_min or "n.a.")
        ws.cell(row=i, column=7, value=bm.costo_giornata_degenza_max or "n.a.")
        for col in range(1, 8):
            ws.cell(row=i, column=col).font = FONT_NORMALE
            ws.cell(row=i, column=col).border = BORDO_SOTTILE

    _imposta_larghezza_colonne(ws, {
        1: 25, 2: 22, 3: 22, 4: 18, 5: 18, 6: 22, 7: 22
    })
    logger.info("Foglio Benchmark_Settore creato")


def _crea_foglio_parametri_scenari(wb):
    """Crea il foglio Parametri_Scenari."""
    ws = wb.create_sheet("Parametri_Scenari")
    ws.cell(row=1, column=1, value="PARAMETRI SCENARI CASH FLOW")
    ws.cell(row=1, column=1).font = FONT_TITOLO
    ws.merge_cells("A1:D1")

    intestazioni = ["Parametro", "Ottimistico", "Base", "Pessimistico"]
    riga = 3
    for col, titolo in enumerate(intestazioni, 1):
        ws.cell(row=riga, column=col, value=titolo)
    _applica_stile_intestazione(ws, riga, len(intestazioni))

    parametri = [
        ("DSO ASP (giorni)", "dso_asp_giorni"),
        ("Delta Occupancy", "occupancy_delta"),
        ("Costi imprevisti %", "costi_imprevisti_pct"),
        ("Crescita ricavi %", "crescita_ricavi_pct"),
    ]

    for i, (label, key) in enumerate(parametri, 4):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=1).font = FONT_NORMALE
        for j, scenario in enumerate(["ottimistico", "base", "pessimistico"], 2):
            valore = SCENARI_CASH_FLOW[scenario].get(key, "")
            ws.cell(row=i, column=j, value=valore)
            ws.cell(row=i, column=j).fill = FILL_INPUT
            ws.cell(row=i, column=j).border = BORDO_SOTTILE

    _imposta_larghezza_colonne(ws, {1: 25, 2: 16, 3: 16, 4: 16})
    logger.info("Foglio Parametri_Scenari creato")


def _crea_foglio_soglie_alert(wb):
    """Crea il foglio Soglie_Alert."""
    ws = wb.create_sheet("Soglie_Alert")
    ws.cell(row=1, column=1, value="CONFIGURAZIONE SOGLIE ALERT")
    ws.cell(row=1, column=1).font = FONT_TITOLO
    ws.merge_cells("A1:C1")

    intestazioni = ["Parametro", "Valore Soglia", "Descrizione"]
    riga = 3
    for col, titolo in enumerate(intestazioni, 1):
        ws.cell(row=riga, column=col, value=titolo)
    _applica_stile_intestazione(ws, riga, len(intestazioni))

    descrizioni = {
        "cassa_minima": "Alert se cassa disponibile < questo valore (€)",
        "copertura_minima_giorni": "Alert se copertura cassa < questi giorni",
        "dso_massimo_asp": "Alert se DSO clienti ASP > questi giorni",
        "dso_massimo_privati": "Alert se DSO clienti privati > questi giorni",
        "dpo_massimo_fornitori": "Alert se DPO fornitori > questi giorni",
        "dscr_minimo": "Alert se DSCR < questo valore",
        "mol_minimo_uo": "Alert se MOL% UO < questo valore",
        "mol_minimo_consolidato": "Alert se MOL% consolidato < questo valore",
        "costo_personale_max_pct": "Alert se costo personale/ricavi > questo %",
        "peso_costi_sede_max_pct": "Alert se costi sede/ricavi > questo %",
        "scostamento_budget_max": "Alert se scostamento budget > questo %",
        "occupancy_minima": "Alert se occupancy < questo %",
        "scostamento_ricavo_giornata": "Alert se scostamento ricavo/giornata > questo %",
        "scostamento_costo_giornata": "Alert se scostamento costo/giornata > questo %",
    }

    for i, (chiave, valore) in enumerate(ALERT_CONFIG.items(), 4):
        ws.cell(row=i, column=1, value=chiave)
        ws.cell(row=i, column=2, value=valore)
        ws.cell(row=i, column=2).fill = FILL_INPUT
        ws.cell(row=i, column=3, value=descrizioni.get(chiave, ""))
        for col in range(1, 4):
            ws.cell(row=i, column=col).font = FONT_NORMALE
            ws.cell(row=i, column=col).border = BORDO_SOTTILE

    _imposta_larghezza_colonne(ws, {1: 30, 2: 18, 3: 50})
    logger.info("Foglio Soglie_Alert creato")


# ============================================================================
# FOGLI SCENARI
# ============================================================================

def _crea_fogli_scenari(wb):
    """Crea i fogli Scenario_1, Scenario_2, Scenario_3."""
    for num in range(1, 4):
        ws = wb.create_sheet(f"Scenario_{num}")
        ws.cell(row=1, column=1, value=f"SCENARIO {num}")
        ws.cell(row=1, column=1).font = FONT_TITOLO
        ws.merge_cells("A1:D1")

        # Intestazione scenario
        info = [
            ("Nome:", ""), ("Descrizione:", ""), ("Stato:", "Bozza"),
            ("Data Creazione:", ""), ("", ""),
        ]
        for i, (label, valore) in enumerate(info, 3):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=1).font = FONT_TOTALE
            ws.cell(row=i, column=2, value=valore)
            ws.cell(row=i, column=2).fill = FILL_INPUT

        # Tabella interventi
        riga = 9
        ws.cell(row=riga, column=1, value="INTERVENTI")
        ws.cell(row=riga, column=1).font = FONT_SEZIONE

        int_intestazioni = [
            "Tipo (Sede/UO)", "Unità Operativa", "Leva",
            "Valore Attuale", "Valore Target",
            "Costo Implementazione", "Tempo (mesi)", "Note"
        ]
        riga += 1
        for col, titolo in enumerate(int_intestazioni, 1):
            ws.cell(row=riga, column=col, value=titolo)
        _applica_stile_intestazione(ws, riga, len(int_intestazioni))

        for r in range(riga + 1, riga + 11):
            for col in range(1, len(int_intestazioni) + 1):
                ws.cell(row=r, column=col).fill = FILL_INPUT
                ws.cell(row=r, column=col).border = BORDO_SOTTILE

        # Impatto economico
        riga_impatto = riga + 12
        ws.cell(row=riga_impatto, column=1, value="IMPATTO ECONOMICO")
        ws.cell(row=riga_impatto, column=1).font = FONT_SEZIONE

        impatti = [
            "Delta Ricavi", "Delta Costi", "Delta MOL", "Delta MOL %",
            "", "Investimento Richiesto", "Payback (mesi)", "Impatto Cassa Anno 1"
        ]
        for i, voce in enumerate(impatti, riga_impatto + 1):
            ws.cell(row=i, column=1, value=voce)
            ws.cell(row=i, column=2).fill = FILL_INPUT

        _imposta_larghezza_colonne(ws, {
            1: 22, 2: 18, 3: 25, 4: 16, 5: 16, 6: 22, 7: 14, 8: 30
        })

    logger.info("Fogli Scenario_1/2/3 creati")


# ============================================================================
# FOGLI DASHBOARD
# ============================================================================

def _crea_fogli_dashboard(wb):
    """Crea i fogli Dashboard (Operativa, Economica, Finanziaria)."""
    for nome, titolo in [
        ("Dashboard_Operativa", "DASHBOARD OPERATIVA - KPI PER UNITÀ OPERATIVA"),
        ("Dashboard_Economica", "DASHBOARD ECONOMICA - CE E MARGINI"),
        ("Dashboard_Finanziaria", "DASHBOARD FINANZIARIA - CASH FLOW E POSIZIONE"),
    ]:
        ws = wb.create_sheet(nome)
        ws.cell(row=1, column=1, value=titolo)
        ws.cell(row=1, column=1).font = FONT_TITOLO
        ws.merge_cells("A1:J1")

        ws.cell(row=3, column=1, value="Periodo: [Aggiornato automaticamente]")
        ws.cell(row=3, column=1).font = Font(italic=True, size=10)

        ws.cell(row=5, column=1, value="[Dashboard generata automaticamente dall'elaborazione]")
        ws.cell(row=5, column=1).font = Font(italic=True, color="999999")

        _imposta_larghezza_colonne(ws, {1: 30})

    logger.info("Fogli Dashboard creati")


# ============================================================================
# FUNZIONE PRINCIPALE
# ============================================================================

def genera_excel_master(output_path: Path = None) -> Path:
    """
    Genera il file Excel Master con tutta la struttura.

    Args:
        output_path: Percorso output. Se None, usa il percorso di default da config.

    Returns:
        Path del file generato.
    """
    if output_path is None:
        output_path = EXCEL_MASTER

    # Crea directory se non esiste
    output_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info(f"Generazione Excel Master: {output_path}")

    wb = openpyxl.Workbook()

    # Rimuovi il foglio di default
    ws_default = wb.active
    wb.remove(ws_default)

    # --- FOGLI DATI ---
    _crea_foglio_anagrafiche_uo(wb)
    _crea_foglio_piano_conti(wb)
    _crea_foglio_personale(wb)
    _crea_foglio_produzione_mensile(wb)
    _crea_foglio_costi_mensili(wb)
    _crea_foglio_costi_sede(wb)
    _crea_foglio_scadenzario(wb)

    # --- FOGLI ELABORAZIONI ---
    _crea_foglio_ce_industriale(wb)
    _crea_foglio_ce_gestionale(wb)
    _crea_foglio_ce_consolidato(wb)
    _crea_foglio_cash_flow_operativo(wb)
    _crea_foglio_cash_flow_strategico(wb)
    _crea_foglio_kpi(wb)

    # --- FOGLI CONFIGURAZIONE ---
    _crea_foglio_driver_allocazione(wb)
    _crea_foglio_benchmark(wb)
    _crea_foglio_parametri_scenari(wb)
    _crea_foglio_soglie_alert(wb)

    # --- FOGLI SCENARI ---
    _crea_fogli_scenari(wb)

    # --- FOGLI DASHBOARD ---
    _crea_fogli_dashboard(wb)

    # Salva
    wb.save(output_path)
    logger.info(f"Excel Master generato con successo: {output_path}")
    logger.info(f"Fogli creati: {wb.sheetnames}")

    return output_path


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
    percorso = genera_excel_master()
    print(f"\nFile generato: {percorso}")
    print(f"Dimensione: {percorso.stat().st_size / 1024:.1f} KB")
