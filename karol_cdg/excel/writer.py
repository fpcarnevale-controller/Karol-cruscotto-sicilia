"""
Scrittura risultati calcolati su file Excel (KAROL_CDG_MASTER.xlsx).

Modulo per la scrittura dei dati elaborati dal sistema di Controllo di
Gestione Karol CDG: Conto Economico Industriale, Gestionale, Consolidato,
Cash Flow, KPI e scenari.

Prima di ogni scrittura viene creato un backup con timestamp del file
originale. Le funzioni utilizzano openpyxl per la formattazione avanzata
delle celle (valuta, percentuali, colori semaforo).
"""

import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from karol_cdg.config import (
    BACKUP_DIR,
    FORMATO_DATA,
    MESI_IT,
    SIMBOLO_VALUTA,
    VOCI_COSTI_DIRETTI,
    VOCI_COSTI_SEDE,
    VOCI_RICAVI,
)

logger = logging.getLogger(__name__)

# ============================================================================
# NOMI FOGLI DI OUTPUT
# ============================================================================

FOGLIO_CE_INDUSTRIALE = "CE_Industriale"
FOGLIO_CE_GESTIONALE = "CE_Gestionale"
FOGLIO_CE_CONSOLIDATO = "CE_Consolidato"
FOGLIO_CASH_FLOW_OPERATIVO = "CF_Operativo"
FOGLIO_CASH_FLOW_STRATEGICO = "CF_Strategico"
FOGLIO_KPI = "KPI"

# ============================================================================
# STILI DI FORMATTAZIONE
# ============================================================================

# Colori semaforo
_COLORE_VERDE = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
_COLORE_GIALLO = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
_COLORE_ROSSO = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
_COLORE_GRIGIO = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# Font predefiniti
_FONT_INTESTAZIONE = Font(name="Calibri", size=11, bold=True)
_FONT_TITOLO = Font(name="Calibri", size=14, bold=True)
_FONT_NORMALE = Font(name="Calibri", size=10)
_FONT_TOTALE = Font(name="Calibri", size=10, bold=True)

# Bordo sottile
_BORDO_SOTTILE = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Formato numerico valuta italiana (euro con separatore migliaia)
_FORMATO_VALUTA_IT = '#.##0,00 "' + SIMBOLO_VALUTA + '"'
# Formato percentuale
_FORMATO_PERCENTUALE = "0,00%"


# ============================================================================
# BACKUP
# ============================================================================


def backup_file(file_path: Path) -> Path:
    """
    Crea un backup con timestamp del file Excel prima della scrittura.

    Il file di backup viene salvato nella directory configurata in
    BACKUP_DIR con il formato: nome_YYYYMMDD_HHMMSS.xlsx

    Parametri
    ---------
    file_path : Path
        Percorso del file Excel da salvaguardare.

    Ritorna
    -------
    Path
        Percorso del file di backup creato.

    Raises
    ------
    FileNotFoundError
        Se il file originale non esiste.
    """
    percorso = Path(file_path)

    if not percorso.exists():
        raise FileNotFoundError(
            f"Impossibile creare il backup: file non trovato {percorso}"
        )

    # Crea la directory di backup se non esiste
    cartella_backup = Path(BACKUP_DIR)
    cartella_backup.mkdir(parents=True, exist_ok=True)

    # Genera il nome del backup con timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_backup = f"{percorso.stem}_{timestamp}{percorso.suffix}"
    percorso_backup = cartella_backup / nome_backup

    shutil.copy2(percorso, percorso_backup)

    logger.info("Backup creato: %s", percorso_backup)
    return percorso_backup


# ============================================================================
# FUNZIONI AUSILIARIE INTERNE
# ============================================================================


def _ottieni_o_crea_foglio(wb, nome_foglio: str) -> Worksheet:
    """
    Restituisce il foglio richiesto dal workbook, creandolo se non esiste.

    Parametri
    ---------
    wb : openpyxl.Workbook
        Workbook aperto.
    nome_foglio : str
        Nome del foglio da ottenere o creare.

    Ritorna
    -------
    Worksheet
        Il foglio di lavoro richiesto.
    """
    if nome_foglio in wb.sheetnames:
        ws = wb[nome_foglio]
        # Pulisce il contenuto esistente per riscrittura completa
        for riga in ws.iter_rows():
            for cella in riga:
                cella.value = None
                cella.fill = PatternFill()
        logger.info("Foglio '%s' esistente svuotato per riscrittura.", nome_foglio)
    else:
        ws = wb.create_sheet(title=nome_foglio)
        logger.info("Foglio '%s' creato.", nome_foglio)

    return ws


def _applica_formattazione_valuta(ws: Worksheet, cella, valore: float) -> None:
    """
    Formatta una cella come importo in valuta italiana (euro).

    Imposta il valore numerico, il formato valuta con separatore migliaia
    italiano, l'allineamento a destra e il font standard.

    Parametri
    ---------
    ws : Worksheet
        Foglio di lavoro corrente (necessario per il contesto).
    cella : Cell
        Cella openpyxl da formattare.
    valore : float
        Importo da inserire nella cella.
    """
    cella.value = valore
    cella.number_format = _FORMATO_VALUTA_IT
    cella.alignment = Alignment(horizontal="right")
    cella.font = _FONT_NORMALE


def _applica_formattazione_percentuale(ws: Worksheet, cella, valore: float) -> None:
    """
    Formatta una cella come percentuale.

    Il valore viene memorizzato come decimale (es. 0.15 per 15%) e
    visualizzato con il formato percentuale italiano.

    Parametri
    ---------
    ws : Worksheet
        Foglio di lavoro corrente.
    cella : Cell
        Cella openpyxl da formattare.
    valore : float
        Valore percentuale come decimale (es. 0.15 = 15%).
    """
    cella.value = valore
    cella.number_format = _FORMATO_PERCENTUALE
    cella.alignment = Alignment(horizontal="right")
    cella.font = _FONT_NORMALE


def _applica_colore_semaforo(ws: Worksheet, cella, colore: str) -> None:
    """
    Applica un colore di sfondo semaforo (verde/giallo/rosso) a una cella.

    Parametri
    ---------
    ws : Worksheet
        Foglio di lavoro corrente.
    cella : Cell
        Cella openpyxl da colorare.
    colore : str
        Livello del semaforo: "verde", "giallo" o "rosso".
    """
    mappa_colori = {
        "verde": _COLORE_VERDE,
        "giallo": _COLORE_GIALLO,
        "rosso": _COLORE_ROSSO,
    }

    riempimento = mappa_colori.get(colore.lower(), _COLORE_GRIGIO)
    cella.fill = riempimento


def _scrivi_intestazione_ce(
    ws: Worksheet,
    titolo: str,
    periodo: str,
    codice_uo: Optional[str] = None,
) -> int:
    """
    Scrive l'intestazione del Conto Economico nel foglio.

    Parametri
    ---------
    ws : Worksheet
        Foglio di lavoro.
    titolo : str
        Titolo del CE (es. "CE Industriale").
    periodo : str
        Periodo di riferimento (es. "03/2026").
    codice_uo : str, opzionale
        Codice dell'Unita' Operativa.

    Ritorna
    -------
    int
        Numero della prossima riga disponibile dopo l'intestazione.
    """
    riga = 1

    # Titolo
    ws.cell(row=riga, column=1, value=titolo).font = _FONT_TITOLO
    riga += 1

    # Periodo
    try:
        parti = periodo.split("/")
        mese_num = int(parti[0])
        anno = parti[1]
        nome_mese = MESI_IT.get(mese_num, parti[0])
        ws.cell(row=riga, column=1, value=f"Periodo: {nome_mese} {anno}")
    except (ValueError, IndexError):
        ws.cell(row=riga, column=1, value=f"Periodo: {periodo}")
    ws.cell(row=riga, column=1).font = _FONT_INTESTAZIONE
    riga += 1

    # UO se specificata
    if codice_uo:
        ws.cell(row=riga, column=1, value=f"Unita' Operativa: {codice_uo}")
        ws.cell(row=riga, column=1).font = _FONT_INTESTAZIONE
        riga += 1

    # Data generazione
    ws.cell(
        row=riga,
        column=1,
        value=f"Generato il: {datetime.now().strftime(FORMATO_DATA)}",
    )
    ws.cell(row=riga, column=1).font = Font(name="Calibri", size=8, italic=True)
    riga += 2  # Riga vuota di separazione

    return riga


def _scrivi_sezione_ce(
    ws: Worksheet,
    riga_inizio: int,
    titolo_sezione: str,
    voci: dict,
    dati: dict,
) -> int:
    """
    Scrive una sezione del Conto Economico (ricavi, costi diretti, ecc.).

    Parametri
    ---------
    ws : Worksheet
        Foglio di lavoro.
    riga_inizio : int
        Riga da cui iniziare a scrivere.
    titolo_sezione : str
        Titolo della sezione (es. "RICAVI").
    voci : dict
        Dizionario {codice: descrizione} delle voci della sezione.
    dati : dict
        Dizionario {codice: importo} con i valori calcolati.

    Ritorna
    -------
    int
        Numero della prossima riga disponibile.
    """
    riga = riga_inizio

    # Intestazione di sezione
    cella_titolo = ws.cell(row=riga, column=1, value=titolo_sezione)
    cella_titolo.font = _FONT_INTESTAZIONE
    cella_titolo.fill = _COLORE_GRIGIO

    # Intestazioni colonne
    ws.cell(row=riga, column=2, value="Descrizione").font = _FONT_INTESTAZIONE
    ws.cell(row=riga, column=3, value="Importo").font = _FONT_INTESTAZIONE
    ws.cell(row=riga, column=3).alignment = Alignment(horizontal="right")
    riga += 1

    # Righe dettaglio
    totale_sezione = 0.0
    for codice, descrizione in voci.items():
        ws.cell(row=riga, column=1, value=codice).font = _FONT_NORMALE
        ws.cell(row=riga, column=2, value=descrizione).font = _FONT_NORMALE
        importo = dati.get(codice, 0.0)
        _applica_formattazione_valuta(ws, ws.cell(row=riga, column=3), importo)
        ws.cell(row=riga, column=3).border = _BORDO_SOTTILE
        totale_sezione += importo
        riga += 1

    # Riga totale di sezione
    ws.cell(row=riga, column=1).font = _FONT_TOTALE
    ws.cell(row=riga, column=2, value=f"TOTALE {titolo_sezione}").font = _FONT_TOTALE
    cella_totale = ws.cell(row=riga, column=3)
    _applica_formattazione_valuta(ws, cella_totale, totale_sezione)
    cella_totale.font = _FONT_TOTALE
    cella_totale.border = Border(
        top=Side(style="thin"),
        bottom=Side(style="double"),
    )
    riga += 2  # Riga vuota di separazione

    return riga


# ============================================================================
# SCRITTURA CONTO ECONOMICO INDUSTRIALE
# ============================================================================


def scrivi_ce_industriale(
    file_path: Path,
    codice_uo: str,
    ce_data: dict,
    periodo: str,
) -> None:
    """
    Scrive il Conto Economico Industriale per una singola UO.

    Il CE Industriale comprende:
      - Ricavi (voci R01-R07)
      - Costi diretti (voci CD01-CD30)
      - Margine Operativo Lordo Industriale

    I dati vengono scritti nel foglio "CE_Industriale" del file Excel.
    Se il foglio esiste gia', viene sovrascritto. Prima della scrittura
    viene creato un backup del file.

    Parametri
    ---------
    file_path : Path
        Percorso del file Excel master.
    codice_uo : str
        Codice dell'Unita' Operativa (es. "VLB").
    ce_data : dict
        Dizionario con i valori calcolati. Le chiavi corrispondono ai
        codici delle voci (es. "R01", "CD01", ecc.). Deve contenere
        anche "mol_industriale" e "mol_industriale_pct".
    periodo : str
        Periodo di riferimento nel formato "MM/YYYY".
    """
    logger.info(
        "Scrittura CE Industriale per UO '%s', periodo %s",
        codice_uo,
        periodo,
    )

    percorso = Path(file_path)
    backup_file(percorso)

    wb = load_workbook(percorso)
    nome_foglio = f"{FOGLIO_CE_INDUSTRIALE}_{codice_uo}"
    ws = _ottieni_o_crea_foglio(wb, nome_foglio)

    # Intestazione
    riga = _scrivi_intestazione_ce(
        ws,
        titolo="Conto Economico Industriale",
        periodo=periodo,
        codice_uo=codice_uo,
    )

    # Sezione Ricavi
    riga = _scrivi_sezione_ce(ws, riga, "RICAVI", VOCI_RICAVI, ce_data)

    # Sezione Costi Diretti
    riga = _scrivi_sezione_ce(ws, riga, "COSTI DIRETTI", VOCI_COSTI_DIRETTI, ce_data)

    # MOL Industriale
    ws.cell(row=riga, column=2, value="MARGINE OPERATIVO LORDO INDUSTRIALE").font = _FONT_TOTALE
    mol = ce_data.get("mol_industriale", 0.0)
    _applica_formattazione_valuta(ws, ws.cell(row=riga, column=3), mol)
    ws.cell(row=riga, column=3).font = _FONT_TOTALE
    riga += 1

    # MOL % su ricavi
    ws.cell(row=riga, column=2, value="MOL Industriale % su ricavi").font = _FONT_NORMALE
    mol_pct = ce_data.get("mol_industriale_pct", 0.0)
    _applica_formattazione_percentuale(ws, ws.cell(row=riga, column=3), mol_pct)

    # Imposta larghezza colonne
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 18

    wb.save(percorso)
    logger.info("CE Industriale salvato nel foglio '%s'.", nome_foglio)


# ============================================================================
# SCRITTURA CONTO ECONOMICO GESTIONALE
# ============================================================================


def scrivi_ce_gestionale(
    file_path: Path,
    codice_uo: str,
    ce_data: dict,
    periodo: str,
) -> None:
    """
    Scrive il Conto Economico Gestionale per una singola UO.

    Il CE Gestionale comprende il CE Industriale piu':
      - Costi sede allocati (voci CS01-CS20)
      - Margine Operativo Lordo Gestionale

    Parametri
    ---------
    file_path : Path
        Percorso del file Excel master.
    codice_uo : str
        Codice dell'Unita' Operativa.
    ce_data : dict
        Dizionario con tutti i valori calcolati (ricavi, costi diretti,
        costi sede allocati, MOL industriale e gestionale).
    periodo : str
        Periodo di riferimento nel formato "MM/YYYY".
    """
    logger.info(
        "Scrittura CE Gestionale per UO '%s', periodo %s",
        codice_uo,
        periodo,
    )

    percorso = Path(file_path)
    backup_file(percorso)

    wb = load_workbook(percorso)
    nome_foglio = f"{FOGLIO_CE_GESTIONALE}_{codice_uo}"
    ws = _ottieni_o_crea_foglio(wb, nome_foglio)

    # Intestazione
    riga = _scrivi_intestazione_ce(
        ws,
        titolo="Conto Economico Gestionale",
        periodo=periodo,
        codice_uo=codice_uo,
    )

    # Sezione Ricavi
    riga = _scrivi_sezione_ce(ws, riga, "RICAVI", VOCI_RICAVI, ce_data)

    # Sezione Costi Diretti
    riga = _scrivi_sezione_ce(ws, riga, "COSTI DIRETTI", VOCI_COSTI_DIRETTI, ce_data)

    # MOL Industriale (subtotale)
    ws.cell(row=riga, column=2, value="MARGINE OPERATIVO LORDO INDUSTRIALE").font = _FONT_TOTALE
    mol_ind = ce_data.get("mol_industriale", 0.0)
    _applica_formattazione_valuta(ws, ws.cell(row=riga, column=3), mol_ind)
    ws.cell(row=riga, column=3).font = _FONT_TOTALE
    riga += 2

    # Sezione Costi Sede Allocati
    riga = _scrivi_sezione_ce(
        ws, riga, "COSTI SEDE ALLOCATI", VOCI_COSTI_SEDE, ce_data
    )

    # MOL Gestionale
    ws.cell(row=riga, column=2, value="MARGINE OPERATIVO LORDO GESTIONALE").font = _FONT_TOTALE
    mol_gest = ce_data.get("mol_gestionale", 0.0)
    _applica_formattazione_valuta(ws, ws.cell(row=riga, column=3), mol_gest)
    ws.cell(row=riga, column=3).font = _FONT_TOTALE
    riga += 1

    # MOL Gestionale %
    ws.cell(row=riga, column=2, value="MOL Gestionale % su ricavi").font = _FONT_NORMALE
    mol_gest_pct = ce_data.get("mol_gestionale_pct", 0.0)
    _applica_formattazione_percentuale(ws, ws.cell(row=riga, column=3), mol_gest_pct)

    # Imposta larghezza colonne
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 18

    wb.save(percorso)
    logger.info("CE Gestionale salvato nel foglio '%s'.", nome_foglio)


# ============================================================================
# SCRITTURA CONTO ECONOMICO CONSOLIDATO
# ============================================================================


def scrivi_ce_consolidato(
    file_path: Path,
    ce_data: dict,
    periodo: str,
) -> None:
    """
    Scrive il Conto Economico Consolidato (somma di tutte le UO).

    Il CE Consolidato include tutte le voci aggregate a livello di gruppo
    con il dettaglio dei costi comuni non allocabili.

    Parametri
    ---------
    file_path : Path
        Percorso del file Excel master.
    ce_data : dict
        Dizionario con i valori consolidati. Contiene le stesse chiavi
        del CE gestionale piu' le voci AC01-AC03.
    periodo : str
        Periodo di riferimento nel formato "MM/YYYY".
    """
    logger.info("Scrittura CE Consolidato, periodo %s", periodo)

    percorso = Path(file_path)
    backup_file(percorso)

    wb = load_workbook(percorso)
    ws = _ottieni_o_crea_foglio(wb, FOGLIO_CE_CONSOLIDATO)

    # Intestazione
    riga = _scrivi_intestazione_ce(
        ws,
        titolo="Conto Economico Consolidato - Gruppo Karol",
        periodo=periodo,
    )

    # Ricavi
    riga = _scrivi_sezione_ce(ws, riga, "RICAVI", VOCI_RICAVI, ce_data)

    # Costi Diretti
    riga = _scrivi_sezione_ce(ws, riga, "COSTI DIRETTI", VOCI_COSTI_DIRETTI, ce_data)

    # MOL Industriale
    ws.cell(row=riga, column=2, value="TOTALE MOL INDUSTRIALE").font = _FONT_TOTALE
    _applica_formattazione_valuta(
        ws, ws.cell(row=riga, column=3), ce_data.get("mol_industriale", 0.0)
    )
    ws.cell(row=riga, column=3).font = _FONT_TOTALE
    riga += 2

    # Costi Sede
    riga = _scrivi_sezione_ce(ws, riga, "COSTI SEDE", VOCI_COSTI_SEDE, ce_data)

    # MOL Gestionale
    ws.cell(row=riga, column=2, value="TOTALE MOL GESTIONALE").font = _FONT_TOTALE
    _applica_formattazione_valuta(
        ws, ws.cell(row=riga, column=3), ce_data.get("mol_gestionale", 0.0)
    )
    ws.cell(row=riga, column=3).font = _FONT_TOTALE
    riga += 2

    # Altri costi (ammortamenti, oneri finanziari, imposte)
    from karol_cdg.config import VOCI_ALTRI_COSTI

    riga = _scrivi_sezione_ce(ws, riga, "ALTRI COSTI", VOCI_ALTRI_COSTI, ce_data)

    # Risultato Netto
    ws.cell(row=riga, column=2, value="RISULTATO NETTO").font = _FONT_TOTALE
    risultato_netto = ce_data.get("risultato_netto", 0.0)
    cella_rn = ws.cell(row=riga, column=3)
    _applica_formattazione_valuta(ws, cella_rn, risultato_netto)
    cella_rn.font = Font(name="Calibri", size=12, bold=True)
    cella_rn.border = Border(
        top=Side(style="thin"),
        bottom=Side(style="double"),
    )

    # Larghezza colonne
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20

    wb.save(percorso)
    logger.info("CE Consolidato salvato nel foglio '%s'.", FOGLIO_CE_CONSOLIDATO)


# ============================================================================
# SCRITTURA CASH FLOW OPERATIVO
# ============================================================================


def scrivi_cash_flow_operativo(
    file_path: Path,
    cf_data: pd.DataFrame,
) -> None:
    """
    Scrive il cash flow operativo mensile nel foglio "CF_Operativo".

    Il DataFrame cf_data deve contenere le colonne: periodo, incassi,
    pagamenti, saldo_netto, saldo_cumulato, cassa_fine_mese.

    Parametri
    ---------
    file_path : Path
        Percorso del file Excel master.
    cf_data : pd.DataFrame
        DataFrame con i dati di cash flow operativo mensili.
    """
    logger.info("Scrittura cash flow operativo")

    percorso = Path(file_path)
    backup_file(percorso)

    wb = load_workbook(percorso)
    ws = _ottieni_o_crea_foglio(wb, FOGLIO_CASH_FLOW_OPERATIVO)

    # Titolo
    ws.cell(row=1, column=1, value="Cash Flow Operativo Mensile").font = _FONT_TITOLO
    ws.cell(
        row=2,
        column=1,
        value=f"Generato il: {datetime.now().strftime(FORMATO_DATA)}",
    ).font = Font(name="Calibri", size=8, italic=True)

    # Intestazioni colonne
    riga_intestazione = 4
    intestazioni = ["Periodo", "Incassi", "Pagamenti", "Saldo Netto",
                     "Saldo Cumulato", "Cassa Fine Mese"]
    for col_idx, intestazione in enumerate(intestazioni, start=1):
        cella = ws.cell(row=riga_intestazione, column=col_idx, value=intestazione)
        cella.font = _FONT_INTESTAZIONE
        cella.fill = _COLORE_GRIGIO
        cella.border = _BORDO_SOTTILE
        cella.alignment = Alignment(horizontal="center")

    # Dati
    riga = riga_intestazione + 1
    for _, record in cf_data.iterrows():
        ws.cell(row=riga, column=1, value=record.get("periodo", "")).font = _FONT_NORMALE
        ws.cell(row=riga, column=1).border = _BORDO_SOTTILE

        for col_idx, campo in enumerate(
            ["incassi", "pagamenti", "saldo_netto", "saldo_cumulato", "cassa_fine_mese"],
            start=2,
        ):
            valore = record.get(campo, 0.0)
            cella = ws.cell(row=riga, column=col_idx)
            _applica_formattazione_valuta(ws, cella, valore)
            cella.border = _BORDO_SOTTILE

        riga += 1

    # Larghezza colonne
    for col_idx in range(1, len(intestazioni) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    wb.save(percorso)
    logger.info("Cash flow operativo salvato nel foglio '%s'.", FOGLIO_CASH_FLOW_OPERATIVO)


# ============================================================================
# SCRITTURA CASH FLOW STRATEGICO
# ============================================================================


def scrivi_cash_flow_strategico(
    file_path: Path,
    cf_data: pd.DataFrame,
) -> None:
    """
    Scrive il cash flow strategico (proiezione a 12 mesi per scenario)
    nel foglio "CF_Strategico".

    Il DataFrame cf_data deve contenere le colonne: periodo, scenario,
    incassi, pagamenti, saldo_netto, cassa_fine_mese.

    Parametri
    ---------
    file_path : Path
        Percorso del file Excel master.
    cf_data : pd.DataFrame
        DataFrame con i dati di cash flow per scenario.
    """
    logger.info("Scrittura cash flow strategico")

    percorso = Path(file_path)
    backup_file(percorso)

    wb = load_workbook(percorso)
    ws = _ottieni_o_crea_foglio(wb, FOGLIO_CASH_FLOW_STRATEGICO)

    # Titolo
    ws.cell(
        row=1, column=1,
        value="Cash Flow Strategico - Proiezione per Scenario",
    ).font = _FONT_TITOLO
    ws.cell(
        row=2,
        column=1,
        value=f"Generato il: {datetime.now().strftime(FORMATO_DATA)}",
    ).font = Font(name="Calibri", size=8, italic=True)

    # Intestazioni
    riga_intestazione = 4
    intestazioni = ["Periodo", "Scenario", "Incassi", "Pagamenti",
                     "Saldo Netto", "Cassa Fine Mese"]
    for col_idx, intestazione in enumerate(intestazioni, start=1):
        cella = ws.cell(row=riga_intestazione, column=col_idx, value=intestazione)
        cella.font = _FONT_INTESTAZIONE
        cella.fill = _COLORE_GRIGIO
        cella.border = _BORDO_SOTTILE
        cella.alignment = Alignment(horizontal="center")

    # Dati
    riga = riga_intestazione + 1
    for _, record in cf_data.iterrows():
        ws.cell(row=riga, column=1, value=record.get("periodo", "")).font = _FONT_NORMALE
        ws.cell(row=riga, column=1).border = _BORDO_SOTTILE

        ws.cell(row=riga, column=2, value=record.get("scenario", "")).font = _FONT_NORMALE
        ws.cell(row=riga, column=2).border = _BORDO_SOTTILE

        for col_idx, campo in enumerate(
            ["incassi", "pagamenti", "saldo_netto", "cassa_fine_mese"],
            start=3,
        ):
            valore = record.get(campo, 0.0)
            cella = ws.cell(row=riga, column=col_idx)
            _applica_formattazione_valuta(ws, cella, valore)
            cella.border = _BORDO_SOTTILE

        riga += 1

    # Larghezza colonne
    for col_idx in range(1, len(intestazioni) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    wb.save(percorso)
    logger.info("Cash flow strategico salvato nel foglio '%s'.", FOGLIO_CASH_FLOW_STRATEGICO)


# ============================================================================
# SCRITTURA KPI
# ============================================================================


def scrivi_kpi(
    file_path: Path,
    kpi_list: list,
    periodo: str,
) -> None:
    """
    Scrive la tabella dei KPI calcolati nel foglio "KPI".

    Ogni elemento della lista kpi_list deve essere un dizionario con
    le chiavi: nome, valore, unita, livello (verde/giallo/rosso),
    benchmark_min, benchmark_max, codice_uo (opzionale).

    Parametri
    ---------
    file_path : Path
        Percorso del file Excel master.
    kpi_list : list
        Lista di dizionari con i KPI calcolati.
    periodo : str
        Periodo di riferimento nel formato "MM/YYYY".
    """
    logger.info("Scrittura KPI, periodo %s (%d indicatori)", periodo, len(kpi_list))

    percorso = Path(file_path)
    backup_file(percorso)

    wb = load_workbook(percorso)
    ws = _ottieni_o_crea_foglio(wb, FOGLIO_KPI)

    # Titolo
    ws.cell(row=1, column=1, value="Indicatori Chiave di Performance (KPI)").font = _FONT_TITOLO

    try:
        parti = periodo.split("/")
        mese_num = int(parti[0])
        anno = parti[1]
        nome_mese = MESI_IT.get(mese_num, parti[0])
        ws.cell(row=2, column=1, value=f"Periodo: {nome_mese} {anno}")
    except (ValueError, IndexError):
        ws.cell(row=2, column=1, value=f"Periodo: {periodo}")
    ws.cell(row=2, column=1).font = _FONT_INTESTAZIONE

    # Intestazioni colonne
    riga_intestazione = 4
    intestazioni_kpi = [
        "UO", "Indicatore", "Valore", "Unita'",
        "Semaforo", "Benchmark Min", "Benchmark Max",
    ]
    for col_idx, intestazione in enumerate(intestazioni_kpi, start=1):
        cella = ws.cell(row=riga_intestazione, column=col_idx, value=intestazione)
        cella.font = _FONT_INTESTAZIONE
        cella.fill = _COLORE_GRIGIO
        cella.border = _BORDO_SOTTILE

    # Righe KPI
    riga = riga_intestazione + 1
    for kpi in kpi_list:
        ws.cell(
            row=riga, column=1, value=kpi.get("codice_uo", "CONS")
        ).font = _FONT_NORMALE
        ws.cell(row=riga, column=1).border = _BORDO_SOTTILE

        ws.cell(row=riga, column=2, value=kpi.get("nome", "")).font = _FONT_NORMALE
        ws.cell(row=riga, column=2).border = _BORDO_SOTTILE

        # Valore: formato in base all'unita'
        valore = kpi.get("valore", 0.0)
        unita = kpi.get("unita", "")
        cella_valore = ws.cell(row=riga, column=3)

        if unita == "%":
            _applica_formattazione_percentuale(ws, cella_valore, valore)
        elif unita in (SIMBOLO_VALUTA, "euro"):
            _applica_formattazione_valuta(ws, cella_valore, valore)
        else:
            cella_valore.value = valore
            cella_valore.font = _FONT_NORMALE
        cella_valore.border = _BORDO_SOTTILE

        ws.cell(row=riga, column=4, value=unita).font = _FONT_NORMALE
        ws.cell(row=riga, column=4).border = _BORDO_SOTTILE

        # Semaforo
        livello = kpi.get("livello", "")
        cella_semaforo = ws.cell(row=riga, column=5, value=livello.upper())
        cella_semaforo.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cella_semaforo.alignment = Alignment(horizontal="center")
        cella_semaforo.border = _BORDO_SOTTILE
        _applica_colore_semaforo(ws, cella_semaforo, livello)

        # Benchmark min e max
        for col_bm, chiave_bm in [(6, "benchmark_min"), (7, "benchmark_max")]:
            valore_bm = kpi.get(chiave_bm)
            cella_bm = ws.cell(row=riga, column=col_bm)
            if valore_bm is not None:
                cella_bm.value = valore_bm
                cella_bm.font = _FONT_NORMALE
            cella_bm.border = _BORDO_SOTTILE

        riga += 1

    # Larghezza colonne
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15

    wb.save(percorso)
    logger.info("KPI salvati nel foglio '%s'.", FOGLIO_KPI)


# ============================================================================
# SCRITTURA SCENARIO
# ============================================================================


def scrivi_scenario(
    file_path: Path,
    scenario_data: dict,
    nome_foglio: str,
) -> None:
    """
    Scrive i risultati di uno scenario what-if in un foglio dedicato.

    Il dizionario scenario_data deve contenere:
      - "nome": nome dello scenario
      - "descrizione": descrizione dello scenario
      - "parametri": dizionario dei parametri utilizzati
      - "risultati": dizionario con i risultati calcolati
        (chiavi: ricavi, costi, mol, cash_flow_12m, ecc.)

    Parametri
    ---------
    file_path : Path
        Percorso del file Excel master.
    scenario_data : dict
        Dizionario con i dati completi dello scenario.
    nome_foglio : str
        Nome del foglio di destinazione (es. "Scenario_Pessimistico").
    """
    logger.info("Scrittura scenario '%s' nel foglio '%s'", scenario_data.get("nome", ""), nome_foglio)

    percorso = Path(file_path)
    backup_file(percorso)

    wb = load_workbook(percorso)
    ws = _ottieni_o_crea_foglio(wb, nome_foglio)

    # Titolo
    nome_scenario = scenario_data.get("nome", "Scenario")
    ws.cell(
        row=1, column=1,
        value=f"Scenario: {nome_scenario}",
    ).font = _FONT_TITOLO

    # Descrizione
    descrizione = scenario_data.get("descrizione", "")
    if descrizione:
        ws.cell(row=2, column=1, value=descrizione).font = Font(
            name="Calibri", size=10, italic=True
        )

    ws.cell(
        row=3,
        column=1,
        value=f"Generato il: {datetime.now().strftime(FORMATO_DATA)}",
    ).font = Font(name="Calibri", size=8, italic=True)

    # Sezione parametri
    riga = 5
    ws.cell(row=riga, column=1, value="PARAMETRI SCENARIO").font = _FONT_INTESTAZIONE
    ws.cell(row=riga, column=1).fill = _COLORE_GRIGIO
    ws.cell(row=riga, column=2).fill = _COLORE_GRIGIO
    riga += 1

    parametri = scenario_data.get("parametri", {})
    for chiave, valore in parametri.items():
        ws.cell(row=riga, column=1, value=chiave).font = _FONT_NORMALE
        ws.cell(row=riga, column=1).border = _BORDO_SOTTILE

        cella_val = ws.cell(row=riga, column=2)
        if isinstance(valore, float) and abs(valore) < 1.0:
            _applica_formattazione_percentuale(ws, cella_val, valore)
        elif isinstance(valore, (int, float)):
            _applica_formattazione_valuta(ws, cella_val, float(valore))
        else:
            cella_val.value = valore
            cella_val.font = _FONT_NORMALE
        cella_val.border = _BORDO_SOTTILE

        riga += 1

    # Sezione risultati
    riga += 1
    ws.cell(row=riga, column=1, value="RISULTATI SCENARIO").font = _FONT_INTESTAZIONE
    ws.cell(row=riga, column=1).fill = _COLORE_GRIGIO
    ws.cell(row=riga, column=2).fill = _COLORE_GRIGIO
    riga += 1

    risultati = scenario_data.get("risultati", {})
    for chiave, valore in risultati.items():
        ws.cell(row=riga, column=1, value=chiave).font = _FONT_NORMALE
        ws.cell(row=riga, column=1).border = _BORDO_SOTTILE

        cella_ris = ws.cell(row=riga, column=2)
        if isinstance(valore, float):
            if "pct" in chiave or "percentuale" in chiave:
                _applica_formattazione_percentuale(ws, cella_ris, valore)
            else:
                _applica_formattazione_valuta(ws, cella_ris, valore)
        elif isinstance(valore, int):
            cella_ris.value = valore
            cella_ris.font = _FONT_NORMALE
        else:
            cella_ris.value = str(valore)
            cella_ris.font = _FONT_NORMALE
        cella_ris.border = _BORDO_SOTTILE

        riga += 1

    # Larghezza colonne
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    wb.save(percorso)
    logger.info("Scenario salvato nel foglio '%s'.", nome_foglio)
