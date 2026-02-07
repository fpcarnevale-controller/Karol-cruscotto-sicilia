"""
Aggiornamento dashboard Excel con dati formattati e tabelle semaforo.

Modulo per la generazione e aggiornamento delle dashboard nel file Excel
master del sistema di Controllo di Gestione Karol CDG. Le dashboard
presentano KPI, trend mensili e confronti tra Unita' Operative con
formattazione a semaforo (verde/giallo/rosso).

Dashboard disponibili:
  - Dashboard Operativa: KPI operativi, occupancy, produzione
  - Dashboard Economica: CE consolidato, confronto MOL per UO
  - Dashboard Finanziaria: cash flow, DSO/DPO, copertura cassa
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from karol_cdg.config import (
    FORMATO_DATA,
    LivelliAlert,
    MESI_BREVI_IT,
    MESI_IT,
    SIMBOLO_VALUTA,
    SOGLIE_SEMAFORO,
    UNITA_OPERATIVE,
)

logger = logging.getLogger(__name__)

# ============================================================================
# NOMI FOGLI DASHBOARD
# ============================================================================

FOGLIO_DASHBOARD_OPERATIVA = "Dashboard_Operativa"
FOGLIO_DASHBOARD_ECONOMICA = "Dashboard_Economica"
FOGLIO_DASHBOARD_FINANZIARIA = "Dashboard_Finanziaria"

# ============================================================================
# COSTANTI DI STILE
# ============================================================================

# Colori semaforo
_VERDE = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
_GIALLO = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
_ROSSO = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
_GRIGIO_CHIARO = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_GRIGIO_SCURO = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_BLU_INTESTAZIONE = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# Font
_FONT_TITOLO_DASH = Font(name="Calibri", size=16, bold=True, color="1F4E79")
_FONT_SOTTOTITOLO = Font(name="Calibri", size=12, bold=True, color="2E75B6")
_FONT_INTESTAZIONE = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_FONT_INTESTAZIONE_SCURA = Font(name="Calibri", size=10, bold=True)
_FONT_NORMALE = Font(name="Calibri", size=10)
_FONT_VALORE_GRANDE = Font(name="Calibri", size=14, bold=True)
_FONT_SEMAFORO = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_FONT_DATA = Font(name="Calibri", size=8, italic=True, color="808080")

# Bordi
_BORDO_SOTTILE = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_BORDO_INFERIORE = Border(bottom=Side(style="thin"))

# Formato numerico
_FORMATO_VALUTA = '#.##0,00 "' + SIMBOLO_VALUTA + '"'
_FORMATO_PERCENTUALE = "0,00%"
_FORMATO_NUMERO = "#.##0"


# ============================================================================
# FUNZIONE AUSILIARIA: COLORE DA LIVELLO ALERT
# ============================================================================


def _colore_da_livello(livello: str) -> PatternFill:
    """
    Restituisce il PatternFill openpyxl corrispondente al livello di alert.

    Parametri
    ---------
    livello : str
        Livello del semaforo: "verde", "giallo" o "rosso".
        Accetta anche i valori dell'enum LivelliAlert.

    Ritorna
    -------
    PatternFill
        Oggetto PatternFill con il colore corrispondente.
        Restituisce grigio chiaro per livelli non riconosciuti.
    """
    # Normalizza il livello a stringa minuscola
    if isinstance(livello, LivelliAlert):
        livello_str = livello.value
    else:
        livello_str = str(livello).strip().lower()

    mappa = {
        "verde": _VERDE,
        "giallo": _GIALLO,
        "rosso": _ROSSO,
    }

    return mappa.get(livello_str, _GRIGIO_CHIARO)


# ============================================================================
# FUNZIONI AUSILIARIE INTERNE
# ============================================================================


def _ottieni_o_crea_foglio(wb, nome_foglio: str) -> Worksheet:
    """
    Restituisce il foglio richiesto, creandolo se non esiste.
    Se esiste, ne cancella il contenuto per la riscrittura.

    Parametri
    ---------
    wb : openpyxl.Workbook
        Workbook aperto.
    nome_foglio : str
        Nome del foglio.

    Ritorna
    -------
    Worksheet
        Il foglio di lavoro pronto per la scrittura.
    """
    if nome_foglio in wb.sheetnames:
        ws = wb[nome_foglio]
        for riga in ws.iter_rows():
            for cella in riga:
                cella.value = None
                cella.fill = PatternFill()
                cella.font = Font()
                cella.border = Border()
        logger.info("Foglio '%s' svuotato per riscrittura dashboard.", nome_foglio)
    else:
        ws = wb.create_sheet(title=nome_foglio)
        logger.info("Foglio '%s' creato per dashboard.", nome_foglio)

    return ws


def _scrivi_intestazione_dashboard(
    ws: Worksheet,
    titolo: str,
    periodo: str,
) -> int:
    """
    Scrive l'intestazione standard di una dashboard.

    Parametri
    ---------
    ws : Worksheet
        Foglio di lavoro.
    titolo : str
        Titolo della dashboard.
    periodo : str
        Periodo di riferimento (es. "03/2026").

    Ritorna
    -------
    int
        Numero della prossima riga disponibile.
    """
    riga = 1

    # Titolo principale
    ws.cell(row=riga, column=1, value=titolo).font = _FONT_TITOLO_DASH
    riga += 1

    # Periodo formattato
    try:
        parti = periodo.split("/")
        mese_num = int(parti[0])
        anno = parti[1]
        nome_mese = MESI_IT.get(mese_num, parti[0])
        periodo_esteso = f"Periodo: {nome_mese} {anno}"
    except (ValueError, IndexError):
        periodo_esteso = f"Periodo: {periodo}"

    ws.cell(row=riga, column=1, value=periodo_esteso).font = _FONT_SOTTOTITOLO
    riga += 1

    # Data di generazione
    ws.cell(
        row=riga,
        column=1,
        value=f"Aggiornamento: {datetime.now().strftime(FORMATO_DATA)}",
    ).font = _FONT_DATA
    riga += 2  # Riga vuota di separazione

    return riga


def _formatta_cella_valuta(ws: Worksheet, cella, valore: float) -> None:
    """Imposta valore, formato valuta italiana e allineamento a destra."""
    cella.value = valore
    cella.number_format = _FORMATO_VALUTA
    cella.alignment = Alignment(horizontal="right")
    cella.font = _FONT_NORMALE


def _formatta_cella_percentuale(ws: Worksheet, cella, valore: float) -> None:
    """Imposta valore, formato percentuale e allineamento a destra."""
    cella.value = valore
    cella.number_format = _FORMATO_PERCENTUALE
    cella.alignment = Alignment(horizontal="right")
    cella.font = _FONT_NORMALE


# ============================================================================
# TABELLA SEMAFORI
# ============================================================================


def crea_tabella_semafori(
    ws: Worksheet,
    kpi_list: list,
    riga_inizio: int,
    colonna_inizio: int,
) -> int:
    """
    Crea una tabella con indicatori e semafori nel foglio di lavoro.

    Ogni KPI viene visualizzato su una riga con: nome indicatore, valore
    formattato, cella colorata con il livello di allarme (semaforo).

    Parametri
    ---------
    ws : Worksheet
        Foglio di lavoro su cui scrivere.
    kpi_list : list
        Lista di dizionari KPI. Ogni dizionario deve contenere:
        - "nome" (str): nome dell'indicatore
        - "valore" (float): valore calcolato
        - "unita" (str): unita' di misura ("%", "euro", "gg", ecc.)
        - "livello" (str): "verde", "giallo" o "rosso"
        - "codice_uo" (str, opzionale): codice UO di riferimento
    riga_inizio : int
        Riga da cui iniziare la tabella.
    colonna_inizio : int
        Colonna da cui iniziare la tabella.

    Ritorna
    -------
    int
        Numero della prossima riga disponibile dopo la tabella.
    """
    riga = riga_inizio
    col = colonna_inizio

    # Intestazione tabella
    intestazioni = ["Indicatore", "UO", "Valore", "Stato"]
    for offset, intestazione in enumerate(intestazioni):
        cella = ws.cell(row=riga, column=col + offset, value=intestazione)
        cella.font = _FONT_INTESTAZIONE
        cella.fill = _BLU_INTESTAZIONE
        cella.border = _BORDO_SOTTILE
        cella.alignment = Alignment(horizontal="center")
    riga += 1

    # Righe dati
    for kpi in kpi_list:
        # Nome indicatore
        ws.cell(row=riga, column=col, value=kpi.get("nome", "")).font = _FONT_NORMALE
        ws.cell(row=riga, column=col).border = _BORDO_SOTTILE

        # Codice UO
        codice_uo = kpi.get("codice_uo", "CONS")
        ws.cell(row=riga, column=col + 1, value=codice_uo).font = _FONT_NORMALE
        ws.cell(row=riga, column=col + 1).border = _BORDO_SOTTILE
        ws.cell(row=riga, column=col + 1).alignment = Alignment(horizontal="center")

        # Valore formattato
        valore = kpi.get("valore", 0.0)
        unita = kpi.get("unita", "")
        cella_valore = ws.cell(row=riga, column=col + 2)

        if unita == "%":
            _formatta_cella_percentuale(ws, cella_valore, valore)
        elif unita in (SIMBOLO_VALUTA, "euro"):
            _formatta_cella_valuta(ws, cella_valore, valore)
        elif unita == "gg":
            cella_valore.value = valore
            cella_valore.number_format = _FORMATO_NUMERO
            cella_valore.font = _FONT_NORMALE
            cella_valore.alignment = Alignment(horizontal="right")
        else:
            cella_valore.value = valore
            cella_valore.font = _FONT_NORMALE
            cella_valore.alignment = Alignment(horizontal="right")
        cella_valore.border = _BORDO_SOTTILE

        # Semaforo colorato
        livello = kpi.get("livello", "")
        cella_stato = ws.cell(row=riga, column=col + 3)
        cella_stato.value = livello.upper() if livello else ""
        cella_stato.font = _FONT_SEMAFORO
        cella_stato.fill = _colore_da_livello(livello)
        cella_stato.alignment = Alignment(horizontal="center")
        cella_stato.border = _BORDO_SOTTILE

        # Alternanza colore righe per leggibilita'
        if (riga - riga_inizio) % 2 == 0:
            for offset in range(len(intestazioni)):
                if ws.cell(row=riga, column=col + offset).fill == PatternFill():
                    ws.cell(row=riga, column=col + offset).fill = _GRIGIO_CHIARO

        riga += 1

    return riga


# ============================================================================
# TABELLA CONFRONTO UNITA' OPERATIVE
# ============================================================================


def crea_tabella_confronto_uo(
    ws: Worksheet,
    dati_uo: dict,
    riga_inizio: int,
) -> int:
    """
    Crea una tabella di confronto tra le Unita' Operative.

    La tabella mostra, per ciascuna UO: ricavi, costi diretti,
    MOL industriale, MOL%, occupancy e il livello semaforo complessivo.

    Parametri
    ---------
    ws : Worksheet
        Foglio di lavoro su cui scrivere.
    dati_uo : dict
        Dizionario con chiavi = codice UO e valori = dizionario contenente:
        - "ricavi" (float): totale ricavi
        - "costi_diretti" (float): totale costi diretti
        - "mol_industriale" (float): MOL industriale
        - "mol_industriale_pct" (float): MOL industriale percentuale
        - "occupancy" (float, opzionale): tasso di occupazione
        - "livello" (str): livello semaforo complessivo
    riga_inizio : int
        Riga da cui iniziare la tabella.

    Ritorna
    -------
    int
        Numero della prossima riga disponibile dopo la tabella.
    """
    riga = riga_inizio

    # Titolo sezione
    ws.cell(row=riga, column=1, value="Confronto Unita' Operative").font = _FONT_SOTTOTITOLO
    riga += 1

    # Intestazioni colonne
    intestazioni = [
        "Codice UO", "Nome", "Ricavi", "Costi Diretti",
        "MOL Industriale", "MOL %", "Occupancy", "Stato",
    ]
    for col_idx, intestazione in enumerate(intestazioni, start=1):
        cella = ws.cell(row=riga, column=col_idx, value=intestazione)
        cella.font = _FONT_INTESTAZIONE
        cella.fill = _BLU_INTESTAZIONE
        cella.border = _BORDO_SOTTILE
        cella.alignment = Alignment(horizontal="center")
    riga += 1

    # Ordina le UO per codice
    codici_ordinati = sorted(dati_uo.keys())

    # Righe dettaglio per ogni UO
    for idx, codice_uo in enumerate(codici_ordinati):
        dati = dati_uo[codice_uo]

        # Colore alternato per le righe
        colore_riga = _GRIGIO_CHIARO if idx % 2 == 0 else PatternFill()

        # Codice UO
        cella_cod = ws.cell(row=riga, column=1, value=codice_uo)
        cella_cod.font = Font(name="Calibri", size=10, bold=True)
        cella_cod.border = _BORDO_SOTTILE
        cella_cod.alignment = Alignment(horizontal="center")
        if colore_riga != PatternFill():
            cella_cod.fill = colore_riga

        # Nome UO (dall'anagrafica di config, se disponibile)
        uo_config = UNITA_OPERATIVE.get(codice_uo)
        nome_uo = uo_config.nome if uo_config else codice_uo
        cella_nome = ws.cell(row=riga, column=2, value=nome_uo)
        cella_nome.font = _FONT_NORMALE
        cella_nome.border = _BORDO_SOTTILE
        if colore_riga != PatternFill():
            cella_nome.fill = colore_riga

        # Ricavi
        cella_ricavi = ws.cell(row=riga, column=3)
        _formatta_cella_valuta(ws, cella_ricavi, dati.get("ricavi", 0.0))
        cella_ricavi.border = _BORDO_SOTTILE
        if colore_riga != PatternFill():
            cella_ricavi.fill = colore_riga

        # Costi Diretti
        cella_costi = ws.cell(row=riga, column=4)
        _formatta_cella_valuta(ws, cella_costi, dati.get("costi_diretti", 0.0))
        cella_costi.border = _BORDO_SOTTILE
        if colore_riga != PatternFill():
            cella_costi.fill = colore_riga

        # MOL Industriale
        cella_mol = ws.cell(row=riga, column=5)
        _formatta_cella_valuta(ws, cella_mol, dati.get("mol_industriale", 0.0))
        cella_mol.border = _BORDO_SOTTILE
        if colore_riga != PatternFill():
            cella_mol.fill = colore_riga

        # MOL %
        cella_mol_pct = ws.cell(row=riga, column=6)
        _formatta_cella_percentuale(ws, cella_mol_pct, dati.get("mol_industriale_pct", 0.0))
        cella_mol_pct.border = _BORDO_SOTTILE
        if colore_riga != PatternFill():
            cella_mol_pct.fill = colore_riga

        # Occupancy
        cella_occ = ws.cell(row=riga, column=7)
        occupancy = dati.get("occupancy")
        if occupancy is not None:
            _formatta_cella_percentuale(ws, cella_occ, occupancy)
        else:
            cella_occ.value = "N/A"
            cella_occ.font = _FONT_NORMALE
            cella_occ.alignment = Alignment(horizontal="center")
        cella_occ.border = _BORDO_SOTTILE
        if colore_riga != PatternFill():
            cella_occ.fill = colore_riga

        # Stato semaforo
        livello = dati.get("livello", "")
        cella_stato = ws.cell(row=riga, column=8)
        cella_stato.value = livello.upper() if livello else ""
        cella_stato.font = _FONT_SEMAFORO
        cella_stato.fill = _colore_da_livello(livello)
        cella_stato.alignment = Alignment(horizontal="center")
        cella_stato.border = _BORDO_SOTTILE

        riga += 1

    # Riga totale/media
    riga += 1  # Riga vuota

    return riga


# ============================================================================
# DASHBOARD OPERATIVA
# ============================================================================


def aggiorna_dashboard_operativa(
    file_path: Path,
    periodo: str,
    kpi_operativi: list,
    dati_uo: dict,
) -> None:
    """
    Aggiorna la dashboard operativa con KPI, semafori e confronto UO.

    La dashboard operativa mostra:
      - Tabella KPI operativi con semaforo (occupancy, produzione, ecc.)
      - Tabella di confronto tra Unita' Operative
      - Dati di trend (se presenti nei dati UO)

    Parametri
    ---------
    file_path : Path
        Percorso del file Excel master.
    periodo : str
        Periodo di riferimento nel formato "MM/YYYY".
    kpi_operativi : list
        Lista di dizionari KPI operativi (occupancy, giornate degenza,
        prestazioni, ecc.). Ogni dizionario contiene: nome, valore,
        unita, livello, codice_uo.
    dati_uo : dict
        Dizionario {codice_uo: dati} con i dati per ogni UO.
    """
    logger.info("Aggiornamento dashboard operativa, periodo %s", periodo)

    percorso = Path(file_path)
    wb = load_workbook(percorso)
    ws = _ottieni_o_crea_foglio(wb, FOGLIO_DASHBOARD_OPERATIVA)

    # Intestazione
    riga = _scrivi_intestazione_dashboard(
        ws,
        titolo="Dashboard Operativa",
        periodo=periodo,
    )

    # Sezione 1: Tabella KPI operativi con semafori
    ws.cell(row=riga, column=1, value="KPI Operativi").font = _FONT_SOTTOTITOLO
    riga += 1

    if kpi_operativi:
        riga = crea_tabella_semafori(ws, kpi_operativi, riga, colonna_inizio=1)
    else:
        ws.cell(
            row=riga, column=1,
            value="Nessun KPI operativo disponibile per il periodo selezionato.",
        ).font = Font(name="Calibri", size=10, italic=True, color="808080")
        riga += 1

    riga += 1  # Separazione

    # Sezione 2: Confronto Unita' Operative
    if dati_uo:
        riga = crea_tabella_confronto_uo(ws, dati_uo, riga)
    else:
        ws.cell(
            row=riga, column=1,
            value="Nessun dato UO disponibile per il confronto.",
        ).font = Font(name="Calibri", size=10, italic=True, color="808080")
        riga += 1

    # Imposta larghezza colonne
    larghezze = {
        "A": 30, "B": 30, "C": 18, "D": 18,
        "E": 18, "F": 12, "G": 12, "H": 12,
    }
    for lettera, larghezza in larghezze.items():
        ws.column_dimensions[lettera].width = larghezza

    wb.save(percorso)
    logger.info("Dashboard operativa aggiornata nel foglio '%s'.", FOGLIO_DASHBOARD_OPERATIVA)


# ============================================================================
# DASHBOARD ECONOMICA
# ============================================================================


def aggiorna_dashboard_economica(
    file_path: Path,
    periodo: str,
    ce_consolidato: dict,
    ce_per_uo: dict,
) -> None:
    """
    Aggiorna la dashboard economica con CE consolidato e dettaglio per UO.

    La dashboard economica mostra:
      - Riepilogo CE consolidato (ricavi, costi, MOL)
      - Tabella confronto MOL per UO
      - Incidenza costi sede su ricavi

    Parametri
    ---------
    file_path : Path
        Percorso del file Excel master.
    periodo : str
        Periodo di riferimento nel formato "MM/YYYY".
    ce_consolidato : dict
        Dizionario con i dati del CE consolidato. Chiavi principali:
        totale_ricavi, totale_costi_diretti, mol_industriale,
        mol_industriale_pct, totale_costi_sede, mol_gestionale,
        mol_gestionale_pct, risultato_netto.
    ce_per_uo : dict
        Dizionario {codice_uo: ce_data} con i dati CE per ogni UO.
    """
    logger.info("Aggiornamento dashboard economica, periodo %s", periodo)

    percorso = Path(file_path)
    wb = load_workbook(percorso)
    ws = _ottieni_o_crea_foglio(wb, FOGLIO_DASHBOARD_ECONOMICA)

    # Intestazione
    riga = _scrivi_intestazione_dashboard(
        ws,
        titolo="Dashboard Economica",
        periodo=periodo,
    )

    # Sezione 1: Riepilogo CE Consolidato
    ws.cell(row=riga, column=1, value="Conto Economico Consolidato").font = _FONT_SOTTOTITOLO
    riga += 1

    # Card riepilogative
    voci_riepilogo = [
        ("Totale Ricavi", ce_consolidato.get("totale_ricavi", 0.0), "valuta"),
        ("Totale Costi Diretti", ce_consolidato.get("totale_costi_diretti", 0.0), "valuta"),
        ("MOL Industriale", ce_consolidato.get("mol_industriale", 0.0), "valuta"),
        ("MOL Industriale %", ce_consolidato.get("mol_industriale_pct", 0.0), "pct"),
        ("Totale Costi Sede", ce_consolidato.get("totale_costi_sede", 0.0), "valuta"),
        ("MOL Gestionale", ce_consolidato.get("mol_gestionale", 0.0), "valuta"),
        ("MOL Gestionale %", ce_consolidato.get("mol_gestionale_pct", 0.0), "pct"),
        ("Risultato Netto", ce_consolidato.get("risultato_netto", 0.0), "valuta"),
    ]

    # Intestazioni card
    ws.cell(row=riga, column=1, value="Voce").font = _FONT_INTESTAZIONE_SCURA
    ws.cell(row=riga, column=1).fill = _GRIGIO_SCURO
    ws.cell(row=riga, column=1).border = _BORDO_SOTTILE
    ws.cell(row=riga, column=2, value="Importo").font = _FONT_INTESTAZIONE_SCURA
    ws.cell(row=riga, column=2).fill = _GRIGIO_SCURO
    ws.cell(row=riga, column=2).border = _BORDO_SOTTILE
    riga += 1

    for voce_nome, voce_valore, voce_tipo in voci_riepilogo:
        ws.cell(row=riga, column=1, value=voce_nome).font = _FONT_NORMALE
        ws.cell(row=riga, column=1).border = _BORDO_SOTTILE

        cella_val = ws.cell(row=riga, column=2)
        if voce_tipo == "pct":
            _formatta_cella_percentuale(ws, cella_val, voce_valore)
        else:
            _formatta_cella_valuta(ws, cella_val, voce_valore)
        cella_val.border = _BORDO_SOTTILE

        # Evidenzia le righe di subtotale/totale
        if "MOL" in voce_nome or "Risultato" in voce_nome:
            ws.cell(row=riga, column=1).font = Font(name="Calibri", size=10, bold=True)
            cella_val.font = Font(name="Calibri", size=10, bold=True)

        riga += 1

    riga += 2  # Separazione

    # Sezione 2: Confronto MOL per UO
    if ce_per_uo:
        ws.cell(
            row=riga, column=1,
            value="Confronto Margine per Unita' Operativa",
        ).font = _FONT_SOTTOTITOLO
        riga += 1

        # Intestazioni
        intestazioni_uo = ["Codice UO", "Nome", "Ricavi", "MOL Ind.", "MOL %", "Stato"]
        for col_idx, intestazione in enumerate(intestazioni_uo, start=1):
            cella = ws.cell(row=riga, column=col_idx, value=intestazione)
            cella.font = _FONT_INTESTAZIONE
            cella.fill = _BLU_INTESTAZIONE
            cella.border = _BORDO_SOTTILE
            cella.alignment = Alignment(horizontal="center")
        riga += 1

        for idx, (codice_uo, ce_uo) in enumerate(sorted(ce_per_uo.items())):
            colore_riga = _GRIGIO_CHIARO if idx % 2 == 0 else PatternFill()

            # Codice
            cella_cod = ws.cell(row=riga, column=1, value=codice_uo)
            cella_cod.font = Font(name="Calibri", size=10, bold=True)
            cella_cod.border = _BORDO_SOTTILE
            cella_cod.alignment = Alignment(horizontal="center")
            if colore_riga != PatternFill():
                cella_cod.fill = colore_riga

            # Nome
            uo_config = UNITA_OPERATIVE.get(codice_uo)
            nome_uo = uo_config.nome if uo_config else codice_uo
            cella_nome = ws.cell(row=riga, column=2, value=nome_uo)
            cella_nome.font = _FONT_NORMALE
            cella_nome.border = _BORDO_SOTTILE
            if colore_riga != PatternFill():
                cella_nome.fill = colore_riga

            # Ricavi
            cella_ricavi = ws.cell(row=riga, column=3)
            _formatta_cella_valuta(ws, cella_ricavi, ce_uo.get("totale_ricavi", 0.0))
            cella_ricavi.border = _BORDO_SOTTILE
            if colore_riga != PatternFill():
                cella_ricavi.fill = colore_riga

            # MOL
            cella_mol = ws.cell(row=riga, column=4)
            _formatta_cella_valuta(ws, cella_mol, ce_uo.get("mol_industriale", 0.0))
            cella_mol.border = _BORDO_SOTTILE
            if colore_riga != PatternFill():
                cella_mol.fill = colore_riga

            # MOL %
            cella_mol_pct = ws.cell(row=riga, column=5)
            _formatta_cella_percentuale(ws, cella_mol_pct, ce_uo.get("mol_industriale_pct", 0.0))
            cella_mol_pct.border = _BORDO_SOTTILE
            if colore_riga != PatternFill():
                cella_mol_pct.fill = colore_riga

            # Semaforo
            livello = ce_uo.get("livello", "")
            cella_stato = ws.cell(row=riga, column=6)
            cella_stato.value = livello.upper() if livello else ""
            cella_stato.font = _FONT_SEMAFORO
            cella_stato.fill = _colore_da_livello(livello)
            cella_stato.alignment = Alignment(horizontal="center")
            cella_stato.border = _BORDO_SOTTILE

            riga += 1
    else:
        ws.cell(
            row=riga, column=1,
            value="Nessun dato CE per UO disponibile.",
        ).font = Font(name="Calibri", size=10, italic=True, color="808080")

    # Larghezza colonne
    larghezze = {"A": 30, "B": 30, "C": 18, "D": 18, "E": 12, "F": 12}
    for lettera, larghezza in larghezze.items():
        ws.column_dimensions[lettera].width = larghezza

    wb.save(percorso)
    logger.info("Dashboard economica aggiornata nel foglio '%s'.", FOGLIO_DASHBOARD_ECONOMICA)


# ============================================================================
# DASHBOARD FINANZIARIA
# ============================================================================


def aggiorna_dashboard_finanziaria(
    file_path: Path,
    periodo: str,
    cash_flow: dict,
    kpi_finanziari: list,
) -> None:
    """
    Aggiorna la dashboard finanziaria con cash flow e KPI finanziari.

    La dashboard finanziaria mostra:
      - KPI finanziari con semaforo (DSO, DPO, DSCR, copertura cassa)
      - Riepilogo cash flow del periodo
      - Proiezione cassa a 3/6/12 mesi

    Parametri
    ---------
    file_path : Path
        Percorso del file Excel master.
    periodo : str
        Periodo di riferimento nel formato "MM/YYYY".
    cash_flow : dict
        Dizionario con i dati di cash flow. Chiavi principali:
        incassi_mese, pagamenti_mese, saldo_netto_mese,
        cassa_attuale, proiezione_3m, proiezione_6m, proiezione_12m.
    kpi_finanziari : list
        Lista di dizionari KPI finanziari: DSO ASP, DSO privati,
        DPO fornitori, DSCR, copertura cassa in mesi.
    """
    logger.info("Aggiornamento dashboard finanziaria, periodo %s", periodo)

    percorso = Path(file_path)
    wb = load_workbook(percorso)
    ws = _ottieni_o_crea_foglio(wb, FOGLIO_DASHBOARD_FINANZIARIA)

    # Intestazione
    riga = _scrivi_intestazione_dashboard(
        ws,
        titolo="Dashboard Finanziaria",
        periodo=periodo,
    )

    # Sezione 1: KPI Finanziari con semafori
    ws.cell(row=riga, column=1, value="KPI Finanziari").font = _FONT_SOTTOTITOLO
    riga += 1

    if kpi_finanziari:
        riga = crea_tabella_semafori(ws, kpi_finanziari, riga, colonna_inizio=1)
    else:
        ws.cell(
            row=riga, column=1,
            value="Nessun KPI finanziario disponibile.",
        ).font = Font(name="Calibri", size=10, italic=True, color="808080")
        riga += 1

    riga += 1  # Separazione

    # Sezione 2: Riepilogo Cash Flow
    ws.cell(row=riga, column=1, value="Riepilogo Cash Flow").font = _FONT_SOTTOTITOLO
    riga += 1

    voci_cf = [
        ("Incassi del mese", cash_flow.get("incassi_mese", 0.0)),
        ("Pagamenti del mese", cash_flow.get("pagamenti_mese", 0.0)),
        ("Saldo netto del mese", cash_flow.get("saldo_netto_mese", 0.0)),
        ("Cassa attuale", cash_flow.get("cassa_attuale", 0.0)),
    ]

    # Intestazione tabella CF
    ws.cell(row=riga, column=1, value="Voce").font = _FONT_INTESTAZIONE_SCURA
    ws.cell(row=riga, column=1).fill = _GRIGIO_SCURO
    ws.cell(row=riga, column=1).border = _BORDO_SOTTILE
    ws.cell(row=riga, column=2, value="Importo").font = _FONT_INTESTAZIONE_SCURA
    ws.cell(row=riga, column=2).fill = _GRIGIO_SCURO
    ws.cell(row=riga, column=2).border = _BORDO_SOTTILE
    riga += 1

    for voce_nome, voce_valore in voci_cf:
        ws.cell(row=riga, column=1, value=voce_nome).font = _FONT_NORMALE
        ws.cell(row=riga, column=1).border = _BORDO_SOTTILE

        cella_val = ws.cell(row=riga, column=2)
        _formatta_cella_valuta(ws, cella_val, voce_valore)
        cella_val.border = _BORDO_SOTTILE

        # Evidenzia il saldo netto e la cassa attuale
        if "Saldo netto" in voce_nome or "Cassa attuale" in voce_nome:
            ws.cell(row=riga, column=1).font = Font(name="Calibri", size=10, bold=True)
            cella_val.font = Font(name="Calibri", size=10, bold=True)

        riga += 1

    riga += 1  # Separazione

    # Sezione 3: Proiezioni cassa
    ws.cell(row=riga, column=1, value="Proiezione Cassa").font = _FONT_SOTTOTITOLO
    riga += 1

    proiezioni = [
        ("Proiezione a 3 mesi", cash_flow.get("proiezione_3m", 0.0)),
        ("Proiezione a 6 mesi", cash_flow.get("proiezione_6m", 0.0)),
        ("Proiezione a 12 mesi", cash_flow.get("proiezione_12m", 0.0)),
    ]

    ws.cell(row=riga, column=1, value="Orizzonte").font = _FONT_INTESTAZIONE_SCURA
    ws.cell(row=riga, column=1).fill = _GRIGIO_SCURO
    ws.cell(row=riga, column=1).border = _BORDO_SOTTILE
    ws.cell(row=riga, column=2, value="Cassa Prevista").font = _FONT_INTESTAZIONE_SCURA
    ws.cell(row=riga, column=2).fill = _GRIGIO_SCURO
    ws.cell(row=riga, column=2).border = _BORDO_SOTTILE
    riga += 1

    for proiezione_nome, proiezione_valore in proiezioni:
        ws.cell(row=riga, column=1, value=proiezione_nome).font = _FONT_NORMALE
        ws.cell(row=riga, column=1).border = _BORDO_SOTTILE

        cella_proj = ws.cell(row=riga, column=2)
        _formatta_cella_valuta(ws, cella_proj, proiezione_valore)
        cella_proj.border = _BORDO_SOTTILE

        # Colora la cella in base al segno
        if proiezione_valore < 0:
            cella_proj.font = Font(name="Calibri", size=10, bold=True, color="FF0000")
        elif proiezione_valore > 0:
            cella_proj.font = Font(name="Calibri", size=10, bold=True, color="006100")

        riga += 1

    # Larghezza colonne
    larghezze = {"A": 30, "B": 20, "C": 18, "D": 12}
    for lettera, larghezza in larghezze.items():
        ws.column_dimensions[lettera].width = larghezza

    wb.save(percorso)
    logger.info(
        "Dashboard finanziaria aggiornata nel foglio '%s'.",
        FOGLIO_DASHBOARD_FINANZIARIA,
    )
