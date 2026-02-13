"""
Pagina Gestione Dati.

Permette di verificare lo stato dei dati, caricare nuovi file Excel,
rielaborare i dati e scaricare il report Excel Master.
"""

import os
from datetime import datetime
from pathlib import Path

import streamlit as st
import pandas as pd
import openpyxl

from karol_cdg.config import EXCEL_MASTER, DATA_DIR
from karol_cdg.elabora import elabora_completo


# Fogli attesi nel file Excel Master
FOGLI_ATTESI = [
    "Produzione_Mensile",
    "Costi_Mensili",
    "Costi_Sede_Dettaglio",
    "Driver_Allocazione",
    "Anagrafiche_Personale",
]


def mostra_carica_dati():
    """
    Pagina di gestione dati: stato attuale, caricamento,
    rielaborazione e scaricamento report.
    """
    st.title("Gestione Dati")
    st.markdown("---")

    # =========================================================================
    # SEZIONE 1: STATO ATTUALE DATI
    # =========================================================================
    st.subheader("1. Stato attuale dati")

    if EXCEL_MASTER.exists():
        # Informazioni sul file
        dimensione_file = EXCEL_MASTER.stat().st_size
        data_modifica = datetime.fromtimestamp(EXCEL_MASTER.stat().st_mtime)

        col_info1, col_info2, col_info3 = st.columns(3)

        with col_info1:
            st.metric(
                label="File Master",
                value="Presente",
            )
        with col_info2:
            st.metric(
                label="Dimensione",
                value=f"{dimensione_file / 1024:.1f} KB",
            )
        with col_info3:
            st.metric(
                label="Ultima modifica",
                value=data_modifica.strftime("%d/%m/%Y %H:%M"),
            )

        st.markdown(f"**Percorso**: `{EXCEL_MASTER}`")

        # Conteggio righe per foglio
        st.markdown("**Conteggio righe per foglio:**")

        try:
            wb = openpyxl.load_workbook(EXCEL_MASTER, read_only=True)
            nomi_fogli = wb.sheetnames

            righe_fogli = []
            for nome_foglio in nomi_fogli:
                ws = wb[nome_foglio]
                n_righe = ws.max_row if ws.max_row else 0
                # Sottrai l'header
                n_righe_dati = max(0, n_righe - 1)

                stato = "OK" if n_righe_dati > 0 else "Vuoto"

                righe_fogli.append({
                    "Foglio": nome_foglio,
                    "Righe dati": n_righe_dati,
                    "Stato": stato,
                })

            wb.close()

            df_fogli = pd.DataFrame(righe_fogli)
            st.dataframe(df_fogli, use_container_width=True, hide_index=True)

            # Verifica fogli attesi
            fogli_mancanti = [
                f for f in FOGLI_ATTESI if f not in nomi_fogli
            ]
            if fogli_mancanti:
                st.warning(
                    f"Fogli mancanti nel Master: {', '.join(fogli_mancanti)}"
                )
            else:
                st.success("Tutti i fogli richiesti sono presenti nel Master.")

        except Exception as errore:
            st.error(f"Errore nella lettura del file Master: {errore}")

    else:
        st.warning(
            f"File Master non trovato: `{EXCEL_MASTER}`\n\n"
            "Caricare un file Excel nella sezione sottostante."
        )

    st.markdown("---")

    # =========================================================================
    # SEZIONE 2: CARICA NUOVO FILE EXCEL
    # =========================================================================
    st.subheader("2. Carica nuovo file Excel")
    st.markdown(
        "Carica un file `.xlsx` con i dati aggiornati. "
        "Il file deve contenere i fogli richiesti per l'elaborazione."
    )

    file_caricato = st.file_uploader(
        "Seleziona file Excel (.xlsx)",
        type=["xlsx"],
        help="Il file deve contenere i fogli: "
             + ", ".join(FOGLI_ATTESI),
    )

    if file_caricato is not None:
        st.markdown(
            f"**File selezionato**: {file_caricato.name} "
            f"({file_caricato.size / 1024:.1f} KB)"
        )

        # Validazione struttura
        try:
            wb_caricato = openpyxl.load_workbook(file_caricato, read_only=True)
            fogli_presenti = wb_caricato.sheetnames
            wb_caricato.close()

            # Verifica fogli richiesti
            fogli_mancanti = [
                f for f in FOGLI_ATTESI if f not in fogli_presenti
            ]

            if fogli_mancanti:
                st.error(
                    f"Il file non contiene i fogli richiesti: "
                    f"{', '.join(fogli_mancanti)}\n\n"
                    f"Fogli presenti: {', '.join(fogli_presenti)}"
                )
            else:
                st.success(
                    f"Struttura valida. Fogli trovati: "
                    f"{', '.join(fogli_presenti)}"
                )

                # Anteprima dei dati
                with st.expander("Anteprima dati", expanded=False):
                    # Ricarica il file per la lettura con pandas
                    file_caricato.seek(0)
                    for nome_foglio in FOGLI_ATTESI:
                        if nome_foglio in fogli_presenti:
                            try:
                                df_anteprima = pd.read_excel(
                                    file_caricato,
                                    sheet_name=nome_foglio,
                                    nrows=5,
                                )
                                st.markdown(f"**{nome_foglio}** (prime 5 righe):")
                                st.dataframe(
                                    df_anteprima,
                                    use_container_width=True,
                                    hide_index=True,
                                )
                                file_caricato.seek(0)
                            except Exception as errore:
                                st.warning(
                                    f"Impossibile leggere anteprima di "
                                    f"{nome_foglio}: {errore}"
                                )

                # Pulsante salvataggio
                col_salva, col_annulla = st.columns(2)

                with col_salva:
                    if st.button(
                        "Salva nella cartella dati",
                        type="primary",
                        use_container_width=True,
                    ):
                        try:
                            # Assicurati che la cartella dati esista
                            DATA_DIR.mkdir(parents=True, exist_ok=True)

                            # Salva il file
                            percorso_destinazione = DATA_DIR / file_caricato.name
                            file_caricato.seek(0)
                            contenuto = file_caricato.read()

                            with open(percorso_destinazione, "wb") as f:
                                f.write(contenuto)

                            st.success(
                                f"File salvato con successo in: "
                                f"`{percorso_destinazione}`"
                            )

                            # Se il nome e' diverso dal Master, suggerisci rinomina
                            if percorso_destinazione != EXCEL_MASTER:
                                st.info(
                                    f"Il file e' stato salvato come "
                                    f"`{file_caricato.name}`. "
                                    f"Il file Master atteso e': "
                                    f"`{EXCEL_MASTER.name}`. "
                                    f"Rinominare se necessario."
                                )

                        except Exception as errore:
                            st.error(
                                f"Errore nel salvataggio: {errore}"
                            )

                with col_annulla:
                    st.button(
                        "Annulla",
                        use_container_width=True,
                    )

        except Exception as errore:
            st.error(
                f"Errore nella validazione del file: {errore}\n\n"
                "Assicurarsi che il file sia un valido .xlsx"
            )

    st.markdown("---")

    # =========================================================================
    # SEZIONE 3: RIELABORA DATI
    # =========================================================================
    st.subheader("3. Rielabora dati")
    st.markdown(
        "Esegue l'elaborazione completa: lettura dati, calcolo CE Industriale, "
        "allocazione costi sede, CE Gestionale, KPI e scrittura risultati "
        "nell'Excel Master."
    )

    if not EXCEL_MASTER.exists():
        st.warning(
            "Impossibile elaborare: il file Master non esiste. "
            "Caricare prima un file nella sezione precedente."
        )
    else:
        if st.button(
            "Esegui elaborazione completa",
            type="primary",
            use_container_width=True,
        ):
            # Barra di progresso
            barra_progresso = st.progress(0, text="Avvio elaborazione...")

            try:
                barra_progresso.progress(10, text="Lettura dati dal Master...")

                # Esegui l'elaborazione
                barra_progresso.progress(30, text="Calcolo CE Industriale...")
                barra_progresso.progress(50, text="Allocazione costi sede...")
                barra_progresso.progress(70, text="Calcolo CE Gestionale e KPI...")

                risultati_elaborazione = elabora_completo(EXCEL_MASTER)

                barra_progresso.progress(90, text="Scrittura risultati...")

                # Risultati
                barra_progresso.progress(100, text="Elaborazione completata!")

                st.success("Elaborazione completata con successo!")

                # Mostra riepilogo
                if risultati_elaborazione:
                    ce_ind = risultati_elaborazione.get("ce_industriale", {})
                    ce_gest = risultati_elaborazione.get("ce_gestionale", {})
                    kpi_calcolati = risultati_elaborazione.get("kpi", [])

                    col_r1, col_r2, col_r3, col_r4 = st.columns(4)

                    with col_r1:
                        n_uo = len(ce_ind)
                        st.metric("UO elaborate", n_uo)

                    with col_r2:
                        tot_r = sum(
                            ce_ind[uo]["totale_ricavi"] for uo in ce_ind
                        )
                        st.metric("Ricavi totali", f"{tot_r:,.0f}")

                    with col_r3:
                        tot_mol = sum(
                            ce_gest[uo]["mol_gestionale"] for uo in ce_gest
                        )
                        st.metric("MOL-G totale", f"{tot_mol:,.0f}")

                    with col_r4:
                        st.metric("KPI calcolati", len(kpi_calcolati))

                    # Dimensione aggiornata del file
                    if EXCEL_MASTER.exists():
                        dim_aggiornata = EXCEL_MASTER.stat().st_size
                        st.info(
                            f"File Master aggiornato: "
                            f"{dim_aggiornata / 1024:.1f} KB"
                        )

            except FileNotFoundError as errore:
                barra_progresso.progress(0, text="Errore!")
                st.error(
                    f"File non trovato durante l'elaborazione: {errore}\n\n"
                    "Verificare che il file Master e tutti i dati sorgente "
                    "siano presenti."
                )

            except Exception as errore:
                barra_progresso.progress(0, text="Errore!")
                st.error(
                    f"Errore durante l'elaborazione: {errore}\n\n"
                    "Controllare i log per maggiori dettagli."
                )

    st.markdown("---")

    # =========================================================================
    # SEZIONE 4: SCARICA REPORT
    # =========================================================================
    st.subheader("4. Scarica report")

    if EXCEL_MASTER.exists():
        # Pulsante download Excel Master
        with open(EXCEL_MASTER, "rb") as file_master:
            contenuto_master = file_master.read()

        st.download_button(
            label="Scarica Excel Master",
            data=contenuto_master,
            file_name=EXCEL_MASTER.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        st.markdown(
            f"Il file contiene tutti i fogli aggiornati con i risultati "
            f"dell'ultima elaborazione."
        )
    else:
        st.info("Nessun file Master disponibile per il download.")

    # Placeholder per export PDF (funzionalita' futura)
    st.markdown("---")
    st.info(
        "**Export PDF**: funzionalita' in fase di sviluppo. "
        "Sara' possibile generare un report PDF completo con "
        "grafici, tabelle e analisi."
    )
