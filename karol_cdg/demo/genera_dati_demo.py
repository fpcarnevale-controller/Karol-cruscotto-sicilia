"""
Generatore dati demo per il sistema CDG Karol.
Crea dati fittizi ma realistici per 12 mesi (Gen-Dic 2025) per tutte le UO operative.
Popola il file Excel Master con produzione, costi, personale, costi sede e scadenzario.
"""

import logging
import random
from datetime import datetime, date, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from karol_cdg.config import (
    UNITA_OPERATIVE, UO_OPERATIVE, UO_CLIENTI,
    VOCI_RICAVI, VOCI_COSTI_DIRETTI, VOCI_COSTI_SEDE,
    BENCHMARK, ALERT_CONFIG, MESI_IT, MESI_BREVI_IT,
    CategoriaCostoSede, DriverAllocazione, StatoUO,
    EXCEL_MASTER,
)

logger = logging.getLogger(__name__)

# Seed per riproducibilità
random.seed(42)
np.random.seed(42)

ANNO = 2025
MESI = list(range(1, 13))


# ============================================================================
# PARAMETRI DEMO PER UO
# ============================================================================

# Parametri realistici per ogni UO operativa
PARAMETRI_UO = {
    "VLB": {
        "ricavi_annui": 1_900_000,
        "posti_letto": 44,
        "occupancy_media": 0.93,
        "tariffa_giornaliera": 128,
        "pct_ricavi_convenzione": 0.95,
        "pct_costo_personale": 0.58,
        "pct_acquisti": 0.08,
        "pct_servizi": 0.12,
        "pct_ammortamenti": 0.03,
        "n_dipendenti": 42,
        "qualifiche": {"Medico": 2, "Infermiere": 10, "OSS": 22, "Tecnico": 0, "Amministrativo": 2, "Altro": 6},
    },
    "CTA": {
        "ricavi_annui": 2_280_000,
        "posti_letto": 40,
        "occupancy_media": 0.88,
        "tariffa_giornaliera": 156,
        "pct_ricavi_convenzione": 0.98,
        "pct_costo_personale": 0.57,
        "pct_acquisti": 0.06,
        "pct_servizi": 0.10,
        "pct_ammortamenti": 0.02,
        "n_dipendenti": 38,
        "qualifiche": {"Medico": 4, "Infermiere": 12, "OSS": 14, "Tecnico": 0, "Amministrativo": 2, "Altro": 6},
    },
    "COS": {
        "ricavi_annui": 3_950_000,
        "posti_letto": 50,
        "occupancy_media": 0.85,
        "tariffa_giornaliera": 220,
        "pct_ricavi_convenzione": 0.88,
        "pct_costo_personale": 0.48,
        "pct_acquisti": 0.15,
        "pct_servizi": 0.10,
        "pct_ammortamenti": 0.05,
        "n_dipendenti": 55,
        "qualifiche": {"Medico": 8, "Infermiere": 15, "OSS": 12, "Tecnico": 8, "Amministrativo": 4, "Altro": 8},
    },
    "LAB": {
        "ricavi_annui": 1_100_000,
        "posti_letto": 0,
        "occupancy_media": 0,
        "tariffa_giornaliera": 0,
        "pct_ricavi_convenzione": 0.75,
        "pct_costo_personale": 0.33,
        "pct_acquisti": 0.25,
        "pct_servizi": 0.08,
        "pct_ammortamenti": 0.06,
        "n_dipendenti": 12,
        "qualifiche": {"Medico": 2, "Infermiere": 0, "OSS": 0, "Tecnico": 6, "Amministrativo": 2, "Altro": 2},
    },
    "KCP": {
        "ricavi_annui": 1_400_000,
        "posti_letto": 40,
        "occupancy_media": 0.91,
        "tariffa_giornaliera": 105,
        "pct_ricavi_convenzione": 0.96,
        "pct_costo_personale": 0.55,
        "pct_acquisti": 0.07,
        "pct_servizi": 0.11,
        "pct_ammortamenti": 0.03,
        "n_dipendenti": 30,
        "qualifiche": {"Medico": 2, "Infermiere": 8, "OSS": 14, "Tecnico": 0, "Amministrativo": 2, "Altro": 4},
    },
}

# Costi sede totali ~2.5M suddivisi per voce
COSTI_SEDE_DEMO = {
    "CS01": {"desc": "Contabilità/Amministrazione", "importo": 380_000, "cat": "SERVIZI", "driver": "FATTURE"},
    "CS02": {"desc": "Paghe/HR", "importo": 220_000, "cat": "SERVIZI", "driver": "CEDOLINI"},
    "CS03": {"desc": "Acquisti centralizzati", "importo": 180_000, "cat": "SERVIZI", "driver": "ACQUISTI"},
    "CS04": {"desc": "IT/Sistemi informativi", "importo": 250_000, "cat": "SERVIZI", "driver": "POSTAZIONI"},
    "CS05": {"desc": "Qualità/Compliance", "importo": 150_000, "cat": "SERVIZI", "driver": "POSTI_LETTO"},
    "CS10": {"desc": "Direzione Generale", "importo": 450_000, "cat": "GOVERNANCE", "driver": "RICAVI"},
    "CS11": {"desc": "Affari Legali", "importo": 180_000, "cat": "GOVERNANCE", "driver": "QUOTA_FISSA"},
    "CS12": {"desc": "Strategia/Sviluppo", "importo": 280_000, "cat": "GOVERNANCE", "driver": "NON_ALLOCABILE"},
    "CS20": {"desc": "Costi comuni non allocabili", "importo": 160_000, "cat": "STORICO", "driver": "NON_ALLOCABILE"},
    # Voci aggiuntive realistiche
    "CS30": {"desc": "Affitto sede centrale", "importo": 120_000, "cat": "GOVERNANCE", "driver": "RICAVI"},
    "CS31": {"desc": "Consulenze fiscali/tributarie", "importo": 85_000, "cat": "SERVIZI", "driver": "FATTURE"},
    "CS32": {"desc": "Assicurazioni gruppo", "importo": 95_000, "cat": "SERVIZI", "driver": "RICAVI"},
    "CS33": {"desc": "Progetto Roma (sviluppo)", "importo": 150_000, "cat": "SVILUPPO", "driver": "NON_ALLOCABILE"},
    "CS34": {"desc": "Progetto KMC (sviluppo)", "importo": 120_000, "cat": "SVILUPPO", "driver": "NON_ALLOCABILE"},
    "CS35": {"desc": "Software gestionale legacy", "importo": 45_000, "cat": "STORICO", "driver": "NON_ALLOCABILE"},
    "CS36": {"desc": "Personale sede non assegnato", "importo": 130_000, "cat": "STORICO", "driver": "NON_ALLOCABILE"},
}

# Driver allocazione per UO (valori assoluti)
DRIVER_VALORI_UO = {
    "FATTURE": {"VLB": 180, "CTA": 150, "COS": 350, "LAB": 280, "KCP": 160},
    "CEDOLINI": {"VLB": 42, "CTA": 38, "COS": 55, "LAB": 12, "KCP": 30},
    "ACQUISTI": {"VLB": 150_000, "CTA": 130_000, "COS": 580_000, "LAB": 270_000, "KCP": 100_000},
    "POSTAZIONI": {"VLB": 8, "CTA": 7, "COS": 15, "LAB": 10, "KCP": 6},
    "POSTI_LETTO": {"VLB": 44, "CTA": 40, "COS": 50, "LAB": 0, "KCP": 40},
    "RICAVI": {"VLB": 1_900_000, "CTA": 2_280_000, "COS": 3_950_000, "LAB": 1_100_000, "KCP": 1_400_000},
}


def _variazione_mensile(base, mese, stagionalita=True):
    """Applica variazione mensile realistica con stagionalità."""
    if not stagionalita:
        return base * (1 + np.random.normal(0, 0.03))

    # Stagionalità sanità: calo agosto, leggero calo dicembre
    fattori_stagione = {
        1: 0.98, 2: 0.97, 3: 1.02, 4: 1.01, 5: 1.03, 6: 1.02,
        7: 0.99, 8: 0.80, 9: 1.01, 10: 1.03, 11: 1.02, 12: 0.92,
    }
    fattore = fattori_stagione.get(mese, 1.0)
    rumore = np.random.normal(0, 0.02)
    return base * fattore * (1 + rumore)


# ============================================================================
# GENERAZIONE DATI
# ============================================================================

def genera_produzione_mensile():
    """Genera dati di produzione mensile per tutte le UO operative."""
    righe = []
    for codice_uo in UO_OPERATIVE:
        params = PARAMETRI_UO.get(codice_uo)
        if not params:
            continue

        ricavi_mese_base = params["ricavi_annui"] / 12

        for mese in MESI:
            ricavi_mese = _variazione_mensile(ricavi_mese_base, mese)

            if params["posti_letto"] > 0:
                # Struttura con degenza
                gg_disponibili = params["posti_letto"] * _giorni_mese(mese, ANNO)
                occupancy_mese = params["occupancy_media"] * (
                    0.80 if mese == 8 else (0.93 if mese == 12 else 1.0)
                ) + np.random.normal(0, 0.02)
                occupancy_mese = max(0.5, min(1.0, occupancy_mese))
                giornate = int(gg_disponibili * occupancy_mese)

                ricavi_conv = ricavi_mese * params["pct_ricavi_convenzione"]
                ricavi_priv = ricavi_mese * (1 - params["pct_ricavi_convenzione"])

                # Degenza convenzionata
                righe.append({
                    "unita_operativa": codice_uo, "mese": mese, "anno": ANNO,
                    "tipo_prestazione": "Degenza convenzionata",
                    "codice_prestazione": "R01",
                    "descrizione": f"Giornate degenza conv. {MESI_BREVI_IT[mese]}",
                    "quantita": giornate,
                    "tariffa_unitaria": round(ricavi_conv / max(giornate, 1), 2),
                    "importo_totale": round(ricavi_conv, 2),
                    "giornate_degenza": giornate,
                    "posti_letto_occupati": int(giornate / _giorni_mese(mese, ANNO)),
                    "fonte_dati": "Demo",
                })

                # Privati/solvenza
                if ricavi_priv > 1000:
                    righe.append({
                        "unita_operativa": codice_uo, "mese": mese, "anno": ANNO,
                        "tipo_prestazione": "Degenza privata/solvenza",
                        "codice_prestazione": "R04",
                        "descrizione": f"Prestazioni solvenza {MESI_BREVI_IT[mese]}",
                        "quantita": int(ricavi_priv / (params["tariffa_giornaliera"] * 1.5)),
                        "tariffa_unitaria": round(params["tariffa_giornaliera"] * 1.5, 2),
                        "importo_totale": round(ricavi_priv, 2),
                        "giornate_degenza": 0,
                        "posti_letto_occupati": 0,
                        "fonte_dati": "Demo",
                    })
            else:
                # Laboratorio / Ambulatorio
                ricavi_conv = ricavi_mese * params["pct_ricavi_convenzione"]
                ricavi_priv = ricavi_mese * (1 - params["pct_ricavi_convenzione"])
                n_prestazioni = int(ricavi_conv / 15)  # tariffa media ~15€

                righe.append({
                    "unita_operativa": codice_uo, "mese": mese, "anno": ANNO,
                    "tipo_prestazione": "Prestazioni convenzionate",
                    "codice_prestazione": "R03",
                    "descrizione": f"Analisi/prestazioni conv. {MESI_BREVI_IT[mese]}",
                    "quantita": n_prestazioni,
                    "tariffa_unitaria": round(ricavi_conv / max(n_prestazioni, 1), 2),
                    "importo_totale": round(ricavi_conv, 2),
                    "giornate_degenza": 0,
                    "posti_letto_occupati": 0,
                    "fonte_dati": "Demo",
                })

                if ricavi_priv > 500:
                    righe.append({
                        "unita_operativa": codice_uo, "mese": mese, "anno": ANNO,
                        "tipo_prestazione": "Prestazioni solvenza",
                        "codice_prestazione": "R06",
                        "descrizione": f"Analisi/prestazioni solv. {MESI_BREVI_IT[mese]}",
                        "quantita": int(ricavi_priv / 25),
                        "tariffa_unitaria": 25.0,
                        "importo_totale": round(ricavi_priv, 2),
                        "giornate_degenza": 0,
                        "posti_letto_occupati": 0,
                        "fonte_dati": "Demo",
                    })

    return pd.DataFrame(righe)


def genera_costi_mensili():
    """Genera costi diretti mensili per tutte le UO operative."""
    righe = []
    for codice_uo in UO_OPERATIVE:
        params = PARAMETRI_UO.get(codice_uo)
        if not params:
            continue

        ricavi_annui = params["ricavi_annui"]

        for mese in MESI:
            # Personale (più stabile, meno stagionale)
            costo_pers_mese = (ricavi_annui * params["pct_costo_personale"] / 12)
            # 13a e 14a: extra a giugno e dicembre
            if mese == 6:
                costo_pers_mese *= 1.08  # anticipo 14ma
            elif mese == 12:
                costo_pers_mese *= 1.10  # 13ma

            qualifiche_voci = {
                "Medico": ("CD01", 0.25),
                "Infermiere": ("CD02", 0.30),
                "OSS": ("CD03", 0.30),
                "Tecnico": ("CD04", 0.10),
                "Amministrativo": ("CD05", 0.05),
            }
            for qual, (codice_voce, pct) in qualifiche_voci.items():
                n = params["qualifiche"].get(qual, 0)
                if n > 0:
                    importo = costo_pers_mese * pct * (1 + np.random.normal(0, 0.01))
                    righe.append({
                        "unita_operativa": codice_uo, "mese": mese, "anno": ANNO,
                        "codice_voce": codice_voce,
                        "descrizione_voce": VOCI_COSTI_DIRETTI[codice_voce],
                        "categoria": "Personale Diretto",
                        "importo": round(importo, 2),
                        "note": f"{n} {qual}",
                        "fonte_dati": "Demo",
                    })

            # Acquisti diretti
            acquisti_mese = ricavi_annui * params["pct_acquisti"] / 12
            acquisti_mese = _variazione_mensile(acquisti_mese, mese)
            voci_acquisti = [("CD10", 0.45), ("CD11", 0.25), ("CD12", 0.15), ("CD13", 0.15)]
            for codice_voce, pct in voci_acquisti:
                importo = acquisti_mese * pct
                if importo > 100:
                    righe.append({
                        "unita_operativa": codice_uo, "mese": mese, "anno": ANNO,
                        "codice_voce": codice_voce,
                        "descrizione_voce": VOCI_COSTI_DIRETTI[codice_voce],
                        "categoria": "Acquisti Diretti",
                        "importo": round(importo, 2),
                        "note": "",
                        "fonte_dati": "Demo",
                    })

            # Servizi diretti
            servizi_mese = ricavi_annui * params["pct_servizi"] / 12
            servizi_mese = _variazione_mensile(servizi_mese, mese)
            voci_servizi = [("CD20", 0.15), ("CD21", 0.25), ("CD22", 0.20), ("CD23", 0.30), ("CD24", 0.10)]
            for codice_voce, pct in voci_servizi:
                importo = servizi_mese * pct
                if importo > 100:
                    righe.append({
                        "unita_operativa": codice_uo, "mese": mese, "anno": ANNO,
                        "codice_voce": codice_voce,
                        "descrizione_voce": VOCI_COSTI_DIRETTI[codice_voce],
                        "categoria": "Servizi Diretti",
                        "importo": round(importo, 2),
                        "note": "",
                        "fonte_dati": "Demo",
                    })

            # Ammortamenti
            amm_mese = ricavi_annui * params["pct_ammortamenti"] / 12
            righe.append({
                "unita_operativa": codice_uo, "mese": mese, "anno": ANNO,
                "codice_voce": "CD30",
                "descrizione_voce": VOCI_COSTI_DIRETTI["CD30"],
                "categoria": "Ammortamenti Diretti",
                "importo": round(amm_mese, 2),
                "note": "",
                "fonte_dati": "Demo",
            })

    return pd.DataFrame(righe)


def genera_personale():
    """Genera anagrafica personale fittizia."""
    righe = []
    matricola = 1001

    nomi_m = ["Marco", "Giuseppe", "Antonio", "Francesco", "Luigi", "Salvatore", "Pietro", "Giovanni"]
    nomi_f = ["Maria", "Anna", "Rosa", "Giuseppina", "Carmela", "Francesca", "Teresa", "Angela"]
    cognomi = ["Rossi", "Russo", "Ferrari", "Esposito", "Bianchi", "Romano", "Greco", "Bruno",
               "Gallo", "Conti", "Marino", "Ricci", "Lombardi", "Moretti", "Barbieri"]

    costi_medi = {
        "Medico": 75_000, "Infermiere": 38_000, "OSS": 28_000,
        "Tecnico": 35_000, "Amministrativo": 32_000, "Altro": 26_000,
    }

    for codice_uo in UO_OPERATIVE:
        params = PARAMETRI_UO.get(codice_uo)
        if not params:
            continue

        for qualifica, n in params["qualifiche"].items():
            for i in range(n):
                nome = random.choice(nomi_m if random.random() > 0.6 else nomi_f)
                cognome = random.choice(cognomi)
                costo_base = costi_medi.get(qualifica, 30_000)
                costo_lordo = costo_base * (1 + np.random.normal(0, 0.15))
                contributi = costo_lordo * 0.30
                tfr = costo_lordo * 0.07

                for mese in MESI:
                    righe.append({
                        "matricola": f"M{matricola:05d}",
                        "cognome": cognome,
                        "nome": nome,
                        "qualifica": qualifica,
                        "unita_operativa": codice_uo,
                        "costo_lordo": round(costo_lordo / 12, 2),
                        "contributi": round(contributi / 12, 2),
                        "tfr": round(tfr / 12, 2),
                        "costo_totale": round((costo_lordo + contributi + tfr) / 12, 2),
                        "ore_ordinarie": 160,
                        "ore_straordinarie": random.randint(0, 20),
                        "fte": 1.0,
                        "mese": mese,
                        "anno": ANNO,
                    })
                matricola += 1

    return pd.DataFrame(righe)


def genera_scadenzario():
    """Genera scadenzario incassi/pagamenti per 3 mesi futuri."""
    righe = []
    data_base = date(2026, 1, 1)

    for codice_uo in UO_OPERATIVE:
        params = PARAMETRI_UO.get(codice_uo)
        if not params:
            continue

        ricavi_mese = params["ricavi_annui"] / 12

        for mese_offset in range(3):
            mese_ref = data_base + timedelta(days=30 * mese_offset)

            # Incasso ASP (ritardo 90-120 gg)
            data_incasso = mese_ref + timedelta(days=random.randint(90, 130))
            righe.append({
                "data_scadenza": data_incasso,
                "tipo": "Incasso",
                "categoria": "Incassi ASP/SSN",
                "importo": round(ricavi_mese * params["pct_ricavi_convenzione"], 2),
                "controparte": f"ASP - {UNITA_OPERATIVE[codice_uo].regione.value}",
                "unita_operativa": codice_uo,
                "stato": "Previsto",
                "note": f"Competenza {mese_ref.strftime('%m/%Y')}",
            })

            # Incasso privati (30 gg)
            ricavi_priv = ricavi_mese * (1 - params["pct_ricavi_convenzione"])
            if ricavi_priv > 1000:
                data_incasso_priv = mese_ref + timedelta(days=random.randint(20, 40))
                righe.append({
                    "data_scadenza": data_incasso_priv,
                    "tipo": "Incasso",
                    "categoria": "Incassi privati",
                    "importo": round(ricavi_priv, 2),
                    "controparte": "Clienti privati",
                    "unita_operativa": codice_uo,
                    "stato": "Previsto",
                    "note": "",
                })

            # Stipendi (27 del mese)
            data_stip = date(mese_ref.year, mese_ref.month, 27)
            costo_pers = ricavi_mese * params["pct_costo_personale"]
            righe.append({
                "data_scadenza": data_stip,
                "tipo": "Pagamento",
                "categoria": "Stipendi",
                "importo": round(costo_pers * 0.70, 2),  # netto ~70% del lordo
                "controparte": "Dipendenti",
                "unita_operativa": codice_uo,
                "stato": "Confermato",
                "note": "",
            })

            # Contributi (16 del mese successivo)
            mese_succ = mese_ref + timedelta(days=35)
            data_contr = date(mese_succ.year, mese_succ.month, 16)
            righe.append({
                "data_scadenza": data_contr,
                "tipo": "Pagamento",
                "categoria": "Contributi",
                "importo": round(costo_pers * 0.30, 2),
                "controparte": "INPS/INAIL",
                "unita_operativa": codice_uo,
                "stato": "Confermato",
                "note": "",
            })

            # Fornitori
            data_forn = mese_ref + timedelta(days=random.randint(60, 90))
            costo_forn = ricavi_mese * (params["pct_acquisti"] + params["pct_servizi"])
            righe.append({
                "data_scadenza": data_forn,
                "tipo": "Pagamento",
                "categoria": "Fornitori",
                "importo": round(costo_forn, 2),
                "controparte": "Fornitori vari",
                "unita_operativa": codice_uo,
                "stato": "Previsto",
                "note": "",
            })

    # Pagamenti sede (unici, non per UO)
    for mese_offset in range(3):
        mese_ref = data_base + timedelta(days=30 * mese_offset)
        costi_sede_mese = sum(v["importo"] for v in COSTI_SEDE_DEMO.values()) / 12

        righe.append({
            "data_scadenza": date(mese_ref.year, mese_ref.month, 27),
            "tipo": "Pagamento",
            "categoria": "Stipendi sede",
            "importo": round(costi_sede_mese * 0.55, 2),
            "controparte": "Personale sede",
            "unita_operativa": "SEDE",
            "stato": "Confermato",
            "note": "",
        })

        righe.append({
            "data_scadenza": mese_ref + timedelta(days=60),
            "tipo": "Pagamento",
            "categoria": "Servizi sede",
            "importo": round(costi_sede_mese * 0.45, 2),
            "controparte": "Fornitori sede",
            "unita_operativa": "SEDE",
            "stato": "Previsto",
            "note": "",
        })

    return pd.DataFrame(righe)


def _giorni_mese(mese, anno):
    """Numero di giorni nel mese."""
    import calendar
    return calendar.monthrange(anno, mese)[1]


# ============================================================================
# SCRITTURA SU EXCEL
# ============================================================================

def popola_excel_master(file_path: Path = None):
    """Popola il file Excel Master con tutti i dati demo."""
    if file_path is None:
        file_path = EXCEL_MASTER

    logger.info(f"Caricamento file Excel: {file_path}")
    wb = openpyxl.load_workbook(file_path)

    # --- PRODUZIONE MENSILE ---
    logger.info("Generazione produzione mensile...")
    df_prod = genera_produzione_mensile()
    ws = wb["Produzione_Mensile"]
    # Rimuovi riga placeholder se presente
    if ws.cell(2, 1).value and "Import" in str(ws.cell(2, 1).value):
        ws.delete_rows(2)
    _scrivi_dataframe(ws, df_prod, riga_inizio=2)
    logger.info(f"  {len(df_prod)} righe produzione scritte")

    # --- COSTI MENSILI ---
    logger.info("Generazione costi mensili...")
    df_costi = genera_costi_mensili()
    ws = wb["Costi_Mensili"]
    # Rimuovi righe placeholder
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, 1).value and "---" in str(ws.cell(r, 1).value):
            ws.delete_rows(r)
        elif ws.cell(r, 4).value and str(ws.cell(r, 4).value).startswith("CD"):
            ws.delete_rows(r)
    _scrivi_dataframe(ws, df_costi, riga_inizio=2)
    logger.info(f"  {len(df_costi)} righe costi scritte")

    # --- PERSONALE ---
    logger.info("Generazione anagrafica personale...")
    df_pers = genera_personale()
    ws = wb["Anagrafiche_Personale"]
    if ws.cell(2, 1).value and "Import" in str(ws.cell(2, 1).value):
        ws.delete_rows(2)
    _scrivi_dataframe(ws, df_pers, riga_inizio=2)
    logger.info(f"  {len(df_pers)} righe personale scritte")

    # --- COSTI SEDE ---
    logger.info("Popolamento costi sede dettaglio...")
    ws = wb["Costi_Sede_Dettaglio"]
    _popola_costi_sede(ws)
    logger.info(f"  {len(COSTI_SEDE_DEMO)} voci costi sede scritte")

    # --- DRIVER ALLOCAZIONE ---
    logger.info("Popolamento driver allocazione...")
    ws = wb["Driver_Allocazione"]
    _popola_driver_allocazione(ws)

    # --- SCADENZARIO ---
    logger.info("Generazione scadenzario...")
    df_scad = genera_scadenzario()
    ws = wb["Scadenzario"]
    _scrivi_dataframe(ws, df_scad, riga_inizio=2)
    logger.info(f"  {len(df_scad)} righe scadenzario scritte")

    # --- SALVATAGGIO ---
    wb.save(file_path)
    logger.info(f"Excel Master popolato e salvato: {file_path}")
    logger.info(f"Dimensione: {file_path.stat().st_size / 1024:.1f} KB")

    # Stampa riepilogo
    _stampa_riepilogo(df_prod, df_costi, df_pers)

    return file_path


def _scrivi_dataframe(ws, df, riga_inizio=2):
    """Scrive un DataFrame su un foglio Excel a partire dalla riga indicata."""
    for i, (_, row) in enumerate(df.iterrows()):
        for j, val in enumerate(row):
            cella = ws.cell(row=riga_inizio + i, column=j + 1)
            if isinstance(val, (np.integer,)):
                cella.value = int(val)
            elif isinstance(val, (np.floating,)):
                cella.value = float(val)
            else:
                cella.value = val
            # Formato numeri
            if isinstance(val, float) and val > 100:
                cella.number_format = '#,##0.00'


def _popola_costi_sede(ws):
    """Popola il foglio Costi_Sede_Dettaglio con i dati demo."""
    # Riscrivi dalla riga 2 (riga 1 = header)
    riga = 2
    for codice, dati in COSTI_SEDE_DEMO.items():
        ws.cell(row=riga, column=1, value=codice)
        ws.cell(row=riga, column=2, value=dati["desc"])
        ws.cell(row=riga, column=3, value=dati["importo"])
        ws.cell(row=riga, column=3).number_format = '#,##0.00'
        ws.cell(row=riga, column=4, value=dati["cat"])
        ws.cell(row=riga, column=5, value="")
        ws.cell(row=riga, column=6, value=dati["driver"])
        ws.cell(row=riga, column=7, value="")
        ws.cell(row=riga, column=8, value="No")
        riga += 1


def _popola_driver_allocazione(ws):
    """Popola i valori dei driver per ogni UO nel foglio Driver_Allocazione."""
    # Le colonne UO partono dalla 4 (A=codice, B=desc, C=driver, D..=UO)
    uo_list = list(UNITA_OPERATIVE.keys())

    riga = 4  # prima riga dati
    for codice, dati in COSTI_SEDE_DEMO.items():
        ws.cell(row=riga, column=1, value=codice)
        ws.cell(row=riga, column=2, value=dati["desc"])
        ws.cell(row=riga, column=3, value=dati["driver"])

        driver = dati["driver"]
        if driver in DRIVER_VALORI_UO:
            valori = DRIVER_VALORI_UO[driver]
            for col_idx, uo_cod in enumerate(uo_list, 4):
                valore = valori.get(uo_cod, 0)
                ws.cell(row=riga, column=col_idx, value=valore)
                if isinstance(valore, (int, float)) and valore > 1000:
                    ws.cell(row=riga, column=col_idx).number_format = '#,##0'

        riga += 1


def _stampa_riepilogo(df_prod, df_costi, df_pers):
    """Stampa riepilogo dati generati."""
    print("\n" + "=" * 70)
    print("RIEPILOGO DATI DEMO GENERATI")
    print("=" * 70)

    print(f"\nPeriodo: Gennaio - Dicembre {ANNO}")
    print(f"UO operative: {', '.join(UO_OPERATIVE)}")

    print("\n--- RICAVI PER UO (12 mesi) ---")
    ricavi_uo = df_prod.groupby("unita_operativa")["importo_totale"].sum()
    for uo, ricavi in ricavi_uo.items():
        print(f"  {uo:5s}: € {ricavi:>12,.2f}")
    print(f"  {'TOTALE':5s}: € {ricavi_uo.sum():>12,.2f}")

    print("\n--- COSTI DIRETTI PER UO (12 mesi) ---")
    costi_uo = df_costi.groupby("unita_operativa")["importo"].sum()
    for uo, costi in costi_uo.items():
        print(f"  {uo:5s}: € {costi:>12,.2f}")
    print(f"  {'TOTALE':5s}: € {costi_uo.sum():>12,.2f}")

    print("\n--- MOL INDUSTRIALE PER UO ---")
    for uo in UO_OPERATIVE:
        if uo in ricavi_uo.index and uo in costi_uo.index:
            r = ricavi_uo[uo]
            c = costi_uo[uo]
            mol = r - c
            pct = (mol / r * 100) if r > 0 else 0
            print(f"  {uo:5s}: € {mol:>12,.2f}  ({pct:5.1f}%)")

    print(f"\n--- COSTI SEDE ---")
    totale_sede = sum(v["importo"] for v in COSTI_SEDE_DEMO.values())
    for cat in ["SERVIZI", "GOVERNANCE", "SVILUPPO", "STORICO"]:
        tot_cat = sum(v["importo"] for v in COSTI_SEDE_DEMO.values() if v["cat"] == cat)
        print(f"  {cat:12s}: € {tot_cat:>12,.2f}  ({tot_cat/totale_sede*100:.1f}%)")
    print(f"  {'TOTALE':12s}: € {totale_sede:>12,.2f}")

    print(f"\n--- PERSONALE ---")
    pers_uo = df_pers[df_pers["mese"] == 1].groupby("unita_operativa")["matricola"].nunique()
    for uo, n in pers_uo.items():
        print(f"  {uo:5s}: {n:3d} dipendenti")
    print(f"  {'TOTALE':5s}: {pers_uo.sum():3d} dipendenti")

    print("\n" + "=" * 70)


# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
    )
    popola_excel_master()
