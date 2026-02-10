"""
Genera template Excel vuoti per l'import dei dati reali.

Crea file Excel con le intestazioni corrette e un foglio di istruzioni,
pronti per essere compilati dal controller. I template vengono salvati
nella cartella dati/template/.

Uso:
    python tools/genera_template_import.py
"""

import sys
from pathlib import Path

# Aggiungi root progetto al path
ROOT_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT_DIR))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURAZIONE
# ============================================================================

TEMPLATE_DIR = ROOT_DIR / "dati" / "template"

# Stili
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
INSTRUCTION_FONT = Font(italic=True, color="666666", size=10)
EXAMPLE_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _formatta_intestazione(ws, colonne):
    """Formatta la riga di intestazione."""
    for col_idx, nome in enumerate(colonne, 1):
        cell = ws.cell(row=1, column=col_idx, value=nome)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
        # Larghezza colonna
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(nome) + 4, 15)


def _aggiungi_esempio(ws, riga, valori):
    """Aggiunge una riga di esempio con sfondo azzurro."""
    for col_idx, valore in enumerate(valori, 1):
        cell = ws.cell(row=riga, column=col_idx, value=valore)
        cell.fill = EXAMPLE_FILL
        cell.border = THIN_BORDER


def _aggiungi_foglio_istruzioni(wb, testo_istruzioni):
    """Aggiunge un foglio 'Istruzioni' al workbook."""
    ws = wb.create_sheet("Istruzioni", 0)
    ws.sheet_properties.tabColor = "FF6600"
    ws.column_dimensions["A"].width = 100

    for idx, riga in enumerate(testo_istruzioni.split("\n"), 1):
        cell = ws.cell(row=idx, column=1, value=riga)
        if riga.startswith("#"):
            cell.font = Font(bold=True, size=13, color="1F4E79")
        elif riga.startswith("*"):
            cell.font = Font(bold=True, size=11)
        else:
            cell.font = Font(size=10)


# ============================================================================
# TEMPLATE: COSTI MENSILI (da E-Solver)
# ============================================================================

def genera_template_costi_mensili():
    """Genera template per il foglio Costi_Mensili."""
    wb = openpyxl.Workbook()

    istruzioni = """# Template Import Costi Mensili (da E-Solver)
#
* Fonte: Export saldi per centro di costo da E-Solver (SISTEMI)
* Frequenza: Mensile
* Destinazione: Foglio 'Costi_Mensili' del file KAROL_CDG_MASTER.xlsx
*
* ISTRUZIONI:
1. Esportate i saldi per CDC da E-Solver in formato CSV
2. Compilate il foglio 'Dati' con le righe corrispondenti
3. Usate i codici UO: VLB, CTA, COS, LAB, KCP
4. Usate i codici voce: CD01-CD30 per costi diretti
5. Gli importi devono essere numeri positivi (no simbolo EUR)
6. Una riga per ogni combinazione UO + Voce + Mese
7. Copiate poi i dati nel foglio Costi_Mensili del file Master
*
* CODICI VOCE COSTI DIRETTI:
CD01 = Personale - Medici
CD02 = Personale - Infermieri
CD03 = Personale - OSS/Ausiliari
CD04 = Personale - Tecnici
CD05 = Personale - Amministrativi di struttura
CD10 = Farmaci e presidi sanitari
CD11 = Materiale diagnostico
CD12 = Vitto
CD13 = Altri materiali di consumo
CD20 = Lavanderia
CD21 = Pulizie
CD22 = Manutenzioni ordinarie
CD23 = Utenze
CD24 = Consulenze sanitarie esterne
CD30 = Ammortamenti attrezzature e arredi"""

    _aggiungi_foglio_istruzioni(wb, istruzioni)

    # Foglio dati
    ws = wb.create_sheet("Dati")
    ws.sheet_properties.tabColor = "1F4E79"
    colonne = ["Codice UO", "Codice Voce", "Descrizione", "Mese", "Anno", "Importo"]
    _formatta_intestazione(ws, colonne)

    # Righe esempio
    _aggiungi_esempio(ws, 2, ["VLB", "CD01", "Personale - Medici", 1, 2026, 45678.90])
    _aggiungi_esempio(ws, 3, ["VLB", "CD02", "Personale - Infermieri", 1, 2026, 62340.00])
    _aggiungi_esempio(ws, 4, ["VLB", "CD03", "Personale - OSS/Ausiliari", 1, 2026, 38900.00])
    _aggiungi_esempio(ws, 5, ["VLB", "CD10", "Farmaci e presidi sanitari", 1, 2026, 12500.00])
    _aggiungi_esempio(ws, 6, ["COS", "CD01", "Personale - Medici", 1, 2026, 85000.00])

    # Rimuovi foglio default
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    percorso = TEMPLATE_DIR / "template_costi_mensili.xlsx"
    wb.save(percorso)
    print(f"  Creato: {percorso.name}")


# ============================================================================
# TEMPLATE: PRODUZIONE MENSILE (da Caremed / HT Sang)
# ============================================================================

def genera_template_produzione():
    """Genera template per il foglio Produzione_Mensile."""
    wb = openpyxl.Workbook()

    istruzioni = """# Template Import Produzione Mensile
#
* Fonte Caremed/INNOGEA: strutture COS, LAB, BET
* Fonte HT Sang: strutture VLB, CTA, BRG
* Frequenza: Mensile
* Destinazione: Foglio 'Produzione_Mensile' del file KAROL_CDG_MASTER.xlsx
*
* ISTRUZIONI:
1. Esportate la produzione mensile dal gestionale
2. Aggregate i dati per tipo di ricavo (R01-R07) e UO
3. Compilate il foglio 'Dati' con i totali mensili
4. Una riga per ogni combinazione UO + Codice Ricavo + Mese
5. Copiate poi i dati nel foglio Produzione_Mensile del file Master
*
* CODICI RICAVO:
R01 = Ricavi da convenzione SSN/ASP - Degenza
R02 = Ricavi da convenzione SSN/ASP - Ambulatoriale
R03 = Ricavi da convenzione SSN/ASP - Laboratorio
R04 = Ricavi privati/solvenza - Degenza
R05 = Ricavi privati/solvenza - Pacchetti comfort
R06 = Ricavi privati/solvenza - Ambulatoriale/Laboratorio
R07 = Altri ricavi (affitti, rimborsi, contributi)
*
* NOTE:
- Per le RSA (VLB, CTA): il ricavo principale e' R01 (retta giornaliera SSN)
- Per la Casa di Cura (COS): R01 (degenza), R02 (ambulatoriale), R04 (privati)
- Per il Laboratorio (LAB): R03 (lab SSN), R06 (lab privati)
- La colonna Quantita' indica le prestazioni/giornate (utile per KPI)"""

    _aggiungi_foglio_istruzioni(wb, istruzioni)

    ws = wb.create_sheet("Dati")
    ws.sheet_properties.tabColor = "9BBB59"
    colonne = ["Codice UO", "Codice Voce", "Descrizione", "Mese", "Anno", "Importo", "Quantita"]
    _formatta_intestazione(ws, colonne)

    _aggiungi_esempio(ws, 2, ["VLB", "R01", "Ricavi SSN/ASP - Degenza", 1, 2026, 136080.00, 1296])
    _aggiungi_esempio(ws, 3, ["CTA", "R01", "Ricavi SSN/ASP - Degenza", 1, 2026, 210000.00, 1178])
    _aggiungi_esempio(ws, 4, ["COS", "R01", "Ricavi SSN/ASP - Degenza", 1, 2026, 285000.00, 450])
    _aggiungi_esempio(ws, 5, ["COS", "R02", "Ricavi SSN/ASP - Ambulatoriale", 1, 2026, 15000.00, 120])
    _aggiungi_esempio(ws, 6, ["LAB", "R03", "Ricavi SSN/ASP - Laboratorio", 1, 2026, 95000.00, 4200])

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    percorso = TEMPLATE_DIR / "template_produzione_mensile.xlsx"
    wb.save(percorso)
    print(f"  Creato: {percorso.name}")


# ============================================================================
# TEMPLATE: ANAGRAFICHE PERSONALE (da Zucchetti)
# ============================================================================

def genera_template_personale():
    """Genera template per il foglio Anagrafiche_Personale."""
    wb = openpyxl.Workbook()

    istruzioni = """# Template Import Anagrafiche Personale (da Zucchetti)
#
* Fonte: Export personale da Zucchetti Paghe
* Frequenza: Quando cambia l'organico (assunzioni, cessazioni, trasferimenti)
* Destinazione: Foglio 'Anagrafiche_Personale' del file KAROL_CDG_MASTER.xlsx
*
* ISTRUZIONI:
1. Esportate l'elenco dipendenti in forza da Zucchetti
2. Per ogni dipendente, compilate una riga
3. Usate il codice UO di appartenenza (VLB, CTA, COS, LAB, KCP)
4. Per la qualifica, usate ESATTAMENTE uno dei valori accettati (vedi sotto)
5. Il Costo Mensile Azienda e' il costo totale (lordo + contributi + TFR)
6. Copiate poi i dati nel foglio Anagrafiche_Personale del file Master
*
* QUALIFICHE ACCETTATE:
- Medico
- Infermiere
- OSS/Ausiliario
- Tecnico Laboratorio
- Tecnico Radiologia
- Fisioterapista
- Amministrativo
- Dirigente
- Altro
*
* NOTE:
- Inserite TUTTI i dipendenti, inclusi quelli a tempo determinato
- Il costo mensile deve includere TUTTI gli oneri (lordo + contributi + TFR)
- Le date devono essere nel formato GG/MM/AAAA"""

    _aggiungi_foglio_istruzioni(wb, istruzioni)

    ws = wb.create_sheet("Dati")
    ws.sheet_properties.tabColor = "F79646"
    colonne = [
        "Matricola", "Cognome", "Nome", "Qualifica", "Unita Operativa",
        "Data Assunzione", "Tipo Contratto", "Ore Settimanali",
        "RAL", "Costo Mensile Azienda", "CCNL", "Livello"
    ]
    _formatta_intestazione(ws, colonne)

    _aggiungi_esempio(ws, 2, [
        "001234", "ROSSI", "Mario", "Medico", "VLB",
        "15/03/2020", "Tempo Indeterminato", 36,
        55000, 6360.00, "AIOP", "D"
    ])
    _aggiungi_esempio(ws, 3, [
        "001235", "BIANCHI", "Anna", "Infermiere", "VLB",
        "01/09/2021", "Tempo Indeterminato", 36,
        28000, 3230.00, "AIOP", "D1"
    ])
    _aggiungi_esempio(ws, 4, [
        "001236", "VERDI", "Giuseppe", "OSS/Ausiliario", "CTA",
        "10/01/2022", "Tempo Determinato", 36,
        22000, 2540.00, "AIOP", "C"
    ])

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    percorso = TEMPLATE_DIR / "template_anagrafiche_personale.xlsx"
    wb.save(percorso)
    print(f"  Creato: {percorso.name}")


# ============================================================================
# TEMPLATE: COSTI SEDE (da E-Solver)
# ============================================================================

def genera_template_costi_sede():
    """Genera template per il foglio Costi_Sede_Dettaglio."""
    wb = openpyxl.Workbook()

    istruzioni = """# Template Import Costi Sede (da E-Solver)
#
* Fonte: Export saldi per CDC Sede/HQ da E-Solver
* Frequenza: Mensile
* Destinazione: Foglio 'Costi_Sede_Dettaglio' del file KAROL_CDG_MASTER.xlsx
*
* ISTRUZIONI:
1. Esportate i saldi con centro di costo Sede/HQ da E-Solver
2. Classificate i costi nei codici CS01-CS20
3. Una riga per ogni codice voce + mese
4. Copiate poi i dati nel foglio Costi_Sede_Dettaglio del file Master
*
* CODICI VOCE SEDE:
CS01 = Contabilita'/Amministrazione (driver: Num. fatture)
CS02 = Paghe/HR (driver: Num. cedolini)
CS03 = Acquisti centralizzati (driver: Euro acquistato)
CS04 = IT/Sistemi informativi (driver: Num. postazioni IT)
CS05 = Qualita'/Compliance (driver: Posti letto)
CS10 = Direzione Generale (driver: Ricavi)
CS11 = Affari Legali (driver: Quota fissa)
CS12 = Strategia/Sviluppo (non allocabile)
CS20 = Costi comuni non allocabili"""

    _aggiungi_foglio_istruzioni(wb, istruzioni)

    ws = wb.create_sheet("Dati")
    ws.sheet_properties.tabColor = "C0504D"
    colonne = ["Codice Voce", "Descrizione", "Mese", "Anno", "Importo"]
    _formatta_intestazione(ws, colonne)

    _aggiungi_esempio(ws, 2, ["CS01", "Contabilita'/Amministrazione", 1, 2026, 18500.00])
    _aggiungi_esempio(ws, 3, ["CS02", "Paghe/HR", 1, 2026, 12000.00])
    _aggiungi_esempio(ws, 4, ["CS03", "Acquisti centralizzati", 1, 2026, 8500.00])
    _aggiungi_esempio(ws, 5, ["CS04", "IT/Sistemi informativi", 1, 2026, 15000.00])
    _aggiungi_esempio(ws, 6, ["CS10", "Direzione Generale", 1, 2026, 35000.00])

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    percorso = TEMPLATE_DIR / "template_costi_sede.xlsx"
    wb.save(percorso)
    print(f"  Creato: {percorso.name}")


# ============================================================================
# TEMPLATE: SCADENZARIO (manuale + E-Solver)
# ============================================================================

def genera_template_scadenzario():
    """Genera template per il foglio Scadenzario."""
    wb = openpyxl.Workbook()

    istruzioni = """# Template Import Scadenzario
#
* Fonte: Compilazione manuale + scadenzario fornitori E-Solver
* Frequenza: Mensile (aggiornamento continuo)
* Destinazione: Foglio 'Scadenzario' del file KAROL_CDG_MASTER.xlsx
*
* ISTRUZIONI:
1. Inserite TUTTE le scadenze previste per i prossimi 6-12 mesi
2. Ogni riga e' una scadenza (incasso o pagamento)
3. Il Tipo deve essere esattamente 'Incasso' oppure 'Pagamento'
4. Lo Stato deve essere esattamente 'Previsto', 'Confermato' oppure 'Pagato'
5. Le date devono essere nel formato GG/MM/AAAA
6. Gli importi sono sempre numeri positivi
7. Copiate poi i dati nel foglio Scadenzario del file Master
*
* CATEGORIE INCASSI:
- SSN Convenzione (incassi da ASP/SSN)
- Privato/Solvenza (incassi da pazienti privati)
- Altro (affitti, rimborsi, contributi)
*
* CATEGORIE PAGAMENTI:
- Fornitori (farmaci, materiali, servizi)
- Personale (stipendi, contributi)
- Fiscale (F24, IVA, IRES, IRAP)
- Finanziario (rate mutui, leasing)
- CAPEX (investimenti)
*
* CONSIGLI:
- Incassi ASP: stimate 90-150 giorni dalla fattura
- Stipendi: 27 del mese
- F24: 16 del mese
- IVA trimestrale: fine mese marzo, giugno, settembre, dicembre
- Aggiornate lo stato quando i pagamenti vengono effettuati"""

    _aggiungi_foglio_istruzioni(wb, istruzioni)

    ws = wb.create_sheet("Dati")
    ws.sheet_properties.tabColor = "4BACC6"
    colonne = [
        "Data Scadenza", "Tipo (Incasso/Pagamento)", "Categoria",
        "Importo", "Controparte", "Unita Operativa",
        "Stato (Previsto/Confermato/Pagato)", "Note"
    ]
    _formatta_intestazione(ws, colonne)

    _aggiungi_esempio(ws, 2, [
        "15/03/2026", "Incasso", "SSN Convenzione",
        136080.00, "ASP Palermo", "VLB", "Previsto", "Retta Gennaio 2026"
    ])
    _aggiungi_esempio(ws, 3, [
        "15/03/2026", "Incasso", "SSN Convenzione",
        210000.00, "ASP Palermo", "CTA", "Previsto", "Retta Gennaio 2026"
    ])
    _aggiungi_esempio(ws, 4, [
        "27/03/2026", "Pagamento", "Personale",
        380000.00, "Dipendenti", "VLB", "Confermato", "Stipendi Marzo 2026"
    ])
    _aggiungi_esempio(ws, 5, [
        "16/03/2026", "Pagamento", "Fiscale",
        45000.00, "Erario", "SEDE", "Confermato", "F24 Marzo"
    ])
    _aggiungi_esempio(ws, 6, [
        "31/03/2026", "Pagamento", "Fornitori",
        35000.00, "Fornitore Farmaci SpA", "COS", "Previsto", "Ft. 123/2026"
    ])
    _aggiungi_esempio(ws, 7, [
        "01/04/2026", "Pagamento", "Finanziario",
        15000.00, "Banca XYZ", "SEDE", "Confermato", "Rata mutuo"
    ])

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    percorso = TEMPLATE_DIR / "template_scadenzario.xlsx"
    wb.save(percorso)
    print(f"  Creato: {percorso.name}")


# ============================================================================
# TEMPLATE: MAPPATURA CONTI E-SOLVER -> KAROL CDG
# ============================================================================

def genera_template_mappatura_conti():
    """Genera template per la mappatura conti E-Solver -> Karol CDG."""
    wb = openpyxl.Workbook()

    istruzioni = """# Tabella Mappatura Conti E-Solver -> Codici Karol CDG
#
* Scopo: Tradurre i codici conto E-Solver nei codici voce Karol CDG
* Frequenza: Una tantum (aggiornare solo quando cambia il piano dei conti)
*
* ISTRUZIONI:
1. Elencate TUTTI i conti E-Solver usati per costi e ricavi
2. Per ogni conto, assegnate il codice Karol CDG corrispondente
3. Questa tabella servira' come riferimento per la compilazione mensile
4. In futuro, n8n usera' questa tabella per la conversione automatica
*
* ESEMPIO:
- Conto E-Solver 60.10.001 (Stipendi medici) -> CD01
- Conto E-Solver 61.01.001 (Farmaci) -> CD10
- Conto E-Solver 40.01.001 (Ricavi degenza SSN) -> R01"""

    _aggiungi_foglio_istruzioni(wb, istruzioni)

    ws = wb.create_sheet("Mappatura")
    ws.sheet_properties.tabColor = "2E75B6"
    colonne = [
        "Codice Conto E-Solver", "Descrizione E-Solver",
        "Codice Karol CDG", "Descrizione Karol CDG",
        "Tipo (Ricavo/Costo)", "Note"
    ]
    _formatta_intestazione(ws, colonne)

    # Esempio mappatura
    esempi = [
        ["60.10.001", "Stipendi medici specialisti", "CD01", "Personale - Medici", "Costo", ""],
        ["60.10.002", "Stipendi medici generici", "CD01", "Personale - Medici", "Costo", ""],
        ["60.20.001", "Stipendi infermieri", "CD02", "Personale - Infermieri", "Costo", ""],
        ["60.30.001", "Stipendi OSS/ausiliari", "CD03", "Personale - OSS/Ausiliari", "Costo", ""],
        ["61.01.001", "Acquisto farmaci", "CD10", "Farmaci e presidi sanitari", "Costo", ""],
        ["61.02.001", "Reagenti laboratorio", "CD11", "Materiale diagnostico", "Costo", "Solo LAB"],
        ["64.10.001", "Servizio lavanderia", "CD20", "Lavanderia", "Costo", ""],
        ["64.20.001", "Servizio pulizie", "CD21", "Pulizie", "Costo", ""],
        ["40.01.001", "Ricavi degenza SSN", "R01", "Ricavi conv. SSN - Degenza", "Ricavo", ""],
        ["40.02.001", "Ricavi ambulatoriale SSN", "R02", "Ricavi conv. SSN - Ambulat.", "Ricavo", ""],
        ["40.03.001", "Ricavi laboratorio SSN", "R03", "Ricavi conv. SSN - Lab.", "Ricavo", "Solo LAB"],
    ]

    for idx, valori in enumerate(esempi, 2):
        _aggiungi_esempio(ws, idx, valori)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    percorso = TEMPLATE_DIR / "template_mappatura_conti.xlsx"
    wb.save(percorso)
    print(f"  Creato: {percorso.name}")


# ============================================================================
# MAIN
# ============================================================================

def main():
    """Genera tutti i template di import."""
    print("\n=== Generazione Template Import Karol CDG ===\n")

    # Crea cartella template
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    print(f"Cartella template: {TEMPLATE_DIR}\n")

    # Crea cartelle per import automatizzato (n8n)
    import_dirs = [
        ROOT_DIR / "dati" / "import" / "esolver",
        ROOT_DIR / "dati" / "import" / "zucchetti",
        ROOT_DIR / "dati" / "import" / "caremed",
        ROOT_DIR / "dati" / "import" / "htsang",
        ROOT_DIR / "dati" / "import" / "archivio",
    ]
    print("Creazione cartelle import:")
    for d in import_dirs:
        d.mkdir(parents=True, exist_ok=True)
        # Crea .gitkeep per mantenere le cartelle in git
        gitkeep = d / ".gitkeep"
        if not gitkeep.exists():
            gitkeep.touch()
        print(f"  {d.relative_to(ROOT_DIR)}/")

    print("\nGenerazione template:")
    genera_template_costi_mensili()
    genera_template_produzione()
    genera_template_personale()
    genera_template_costi_sede()
    genera_template_scadenzario()
    genera_template_mappatura_conti()

    print(f"\n=== Completato! {len(list(TEMPLATE_DIR.glob('*.xlsx')))} template generati ===\n")
    print("Prossimi passi:")
    print("1. Aprite i template nella cartella dati/template/")
    print("2. Leggete il foglio 'Istruzioni' di ogni template")
    print("3. Compilate il foglio 'Dati' con i dati reali")
    print("4. Copiate i dati nel file Master KAROL_CDG_MASTER.xlsx")


if __name__ == "__main__":
    main()
